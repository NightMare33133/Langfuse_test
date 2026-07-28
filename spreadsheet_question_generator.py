"""
统一电子表格检索题生成模块。

支持 XLSX、XLS、CSV 三种格式，统一架构：
1. 本地解析表格 → SheetContext 结构化对象
2. 渲染带锚点的 Markdown 表格块 → 发给 LLM
3. LLM 只返回题目 + sheet_name + anchor_range（不输出 reference_answer）
4. 本地按 anchor_range 从原始数据重新渲染金标准证据

核心原则：
- 表格源文件是唯一事实来源
- LLM 不生成、不决定 reference_answer
- reference_answer 必须由本地原始数据按 anchor 渲染得到
"""

import csv
import hashlib
import io
import json
import re
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

import pandas as pd
import requests

from question_generator import deduplicate_questions, MODE_RETRIEVAL

# ─── 常量 ────────────────────────────────────────────────────────────────────

PROMPT_TEMPLATE_PATH = Path(__file__).parent / "prompts" / "qgen_prompt_spreadsheet_retrieval.txt"
_MAX_BLOCK_ROWS = 30       # 单个表格块最大行数
_MAX_EVIDENCE_ROWS = 20    # 单个证据范围最大行数
_MAX_EVIDENCE_COLS = 15    # 单个证据范围最大列数


# ─── 数据结构 ─────────────────────────────────────────────────────────────────

@dataclass
class SheetContext:
    sheet_name: str
    max_row: int
    max_col: int
    headers: list                           # 列标题字符串列表
    rows: list                              # 二维数组，rows[0] = Excel 第 1 行
    merged_cells: list = field(default_factory=list)
    formula_cells_without_cache: list = field(default_factory=list)
    format_warnings: list = field(default_factory=list)
    allowed_anchor_ranges: list = field(default_factory=list)
    table_blocks: list = field(default_factory=list)


@dataclass
class TableBlock:
    block_index: int
    markdown: str
    row_range: tuple                        # (start_row, end_row) 含两端，1-indexed
    col_range: tuple                        # (start_col, end_col) 含两端，1-indexed
    allowed_anchor_ranges: list
    has_formula_warnings: bool
    header_context_range: str = None        # 表头上下文范围（如 B2:D2），用于价格题


@dataclass
class FieldSchema:
    """Phase 1 单字段 schema 分析结果。"""
    source_label: str           # 原始列标题文本
    col_index: int              # 1-indexed 列号
    inferred_role: str          # record_identifier | context | metric | cost | categorical | excluded
    confidence: float           # 0.0-1.0
    needs_confirmation: bool    # True = LLM 不确定或数值字段
    alias: str = ""             # 用户别名（UI 审核时填写）
    user_confirmed: bool = False
    user_action: str = "context_only"  # ask_question | context_only | exclude


@dataclass
class TableSchemaResult:
    """Phase 1 完整 schema 分析结果。"""
    table_purpose: str
    header_row: int
    data_start_row: int
    record_identifier_fields: list = field(default_factory=list)
    context_fields: list = field(default_factory=list)
    metric_fields: list = field(default_factory=list)
    cost_fields: list = field(default_factory=list)
    categorical_fields: list = field(default_factory=list)
    excluded_fields: list = field(default_factory=list)
    excluded_rows: list = field(default_factory=list)
    safe_question_fields: list = field(default_factory=list)
    sheet_name: str = ""
    analysis_model: str = ""
    analysis_timestamp: str = ""


# ─── Schema 缓存 ──────────────────────────────────────────────────────────────

_SCHEMA_CACHE_DIR = Path(__file__).parent / "data" / "schema_cache"
_SCHEMA_ANALYSIS_PROMPT_PATH = Path(__file__).parent / "prompts" / "qgen_schema_analysis_prompt.txt"

# 测试/假模型标识，生产缓存中不允许出现
_FAKE_MODEL_MARKERS = ("fake", "mock", "test", "dummy", "stub")


def _file_content_hash(file_bytes):
    """计算文件内容 SHA-256 哈希。"""
    return hashlib.sha256(file_bytes).hexdigest()


def _is_fake_model(model_name):
    """判断模型名是否为测试/假模型。"""
    if not model_name:
        return True
    lower = model_name.lower()
    return any(lower == marker or lower.startswith(marker + "-") or lower.startswith(marker + "_")
               for marker in _FAKE_MODEL_MARKERS)


def _load_schema_cache(file_hash, expected_model=None, allow_test_model=False):
    """按文件哈希加载 schema 缓存，miss 返回 None。

    Args:
        file_hash: 文件内容哈希
        expected_model: 期望的模型名。如果缓存中的 analysis_model 不匹配，返回 None。
        allow_test_model: 是否允许加载 test/mock 模型生成的缓存（仅测试使用）
    """
    cache_path = _SCHEMA_CACHE_DIR / f"{file_hash}.json"
    if not cache_path.exists():
        return None
    try:
        data = json.loads(cache_path.read_text(encoding="utf-8"))
        if not isinstance(data, dict) or "fields" not in data:
            return None
        # 检测旧格式缓存（使用旧角色名），强制重分析
        if "group_fields" in data and "record_identifier_fields" not in data:
            print(f"  [WARN] 检测到旧格式 schema 缓存，强制重分析")
            return None
        # 拒绝 fake/mock 缓存（除非测试显式允许）
        cached_model = data.get("analysis_model", "")
        if _is_fake_model(cached_model) and not allow_test_model:
            print(f"  [WARN] 拒绝 fake 缓存 (model={cached_model})")
            return None
        # 模型不匹配时拒绝（允许 expected_model 为空则跳过检查）
        if expected_model and cached_model and cached_model != expected_model:
            print(f"  [WARN] 缓存模型不匹配: 缓存={cached_model}, 期望={expected_model}")
            return None
        return data
    except (json.JSONDecodeError, OSError):
        return None


def _save_schema_cache(file_hash, schema_data):
    """保存 schema 分析结果到缓存。"""
    _SCHEMA_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    cache_path = _SCHEMA_CACHE_DIR / f"{file_hash}.json"
    cache_path.write_text(json.dumps(schema_data, ensure_ascii=False, indent=2), encoding="utf-8")


def _delete_schema_cache(file_hash):
    """删除指定文件的 schema 缓存。"""
    cache_path = _SCHEMA_CACHE_DIR / f"{file_hash}.json"
    if cache_path.exists():
        cache_path.unlink()


def _set_schema_cache_dir(path):
    """覆盖 schema 缓存目录（供测试隔离使用）。"""
    global _SCHEMA_CACHE_DIR
    _SCHEMA_CACHE_DIR = Path(path)


def _get_schema_cache_dir():
    """获取当前 schema 缓存目录。"""
    return _SCHEMA_CACHE_DIR


# ─── 列字母转换（独立于 openpyxl） ────────────────────────────────────────────

def _col_letter(n):
    """将 1-indexed 列号转换为 Excel 列字母（1→A, 26→Z, 27→AA）。"""
    result = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _col_index(letter):
    """将 Excel 列字母转换为 1-indexed 列号（A→1, Z→26, AA→27）。"""
    result = 0
    for ch in letter.upper():
        result = result * 26 + (ord(ch) - 64)
    return result


def _parse_range_str(range_str):
    """解析 'A1:C5' 格式的范围字符串，返回 (min_col, min_row, max_col, max_row)，1-indexed。

    Returns None if invalid.
    """
    range_str = range_str.strip().upper()
    match = re.match(r'^([A-Z]+)(\d+):([A-Z]+)(\d+)$', range_str)
    if not match:
        return None
    try:
        min_col = _col_index(match.group(1))
        min_row = int(match.group(2))
        max_col = _col_index(match.group(3))
        max_row = int(match.group(4))
        if min_row < 1 or min_col < 1 or max_row < min_row or max_col < min_col:
            return None
        return (min_col, min_row, max_col, max_row)
    except (ValueError, Exception):
        return None


def _range_to_str(min_col, min_row, max_col, max_row):
    """将行列范围转为 Excel 范围字符串。"""
    return f"{_col_letter(min_col)}{min_row}:{_col_letter(max_col)}{max_row}"


# ─── CSV 编码探测 ─────────────────────────────────────────────────────────────

def _detect_csv_encoding(file_bytes):
    """依次尝试 UTF-8-sig、UTF-8、GBK，最后回退 charset_normalizer。"""
    for enc in ("utf-8-sig", "utf-8", "gbk"):
        try:
            file_bytes.decode(enc)
            return enc
        except (UnicodeDecodeError, LookupError):
            continue
    try:
        import charset_normalizer
        result = charset_normalizer.from_bytes(file_bytes).best()
        if result:
            return result.encoding
    except Exception:
        pass
    return "utf-8"


# ─── 格式解析器 ───────────────────────────────────────────────────────────────

def parse_xlsx_to_sheet_contexts(file_bytes):
    """解析 XLSX 文件为 SheetContext 列表。

    使用 openpyxl 双重打开：data_only=False 获取公式，data_only=True 获取缓存值。
    支持合并单元格填充、公式检测。
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=False)
    try:
        wb_cached = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception:
        wb_cached = None

    contexts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws_cached = wb_cached[sheet_name] if wb_cached and sheet_name in wb_cached.sheetnames else None

        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        if max_row < 1 or max_col < 1:
            continue

        # 读取原始数据（含公式文本）
        rows = []
        for r in range(1, max_row + 1):
            row_vals = []
            for c in range(1, max_col + 1):
                row_vals.append(ws.cell(row=r, column=c).value)
            rows.append(row_vals)

        # 处理合并单元格：将左上角的值填充到范围内所有单元格
        merged_cells = []
        for merge_range in ws.merged_cells.ranges:
            min_col, min_row, max_col_m, max_row_m = (
                merge_range.min_col, merge_range.min_row,
                merge_range.max_col, merge_range.max_row,
            )
            merged_cells.append((min_row, min_col, max_row_m, max_col_m))
            top_left_val = rows[min_row - 1][min_col - 1]
            for r in range(min_row, max_row_m + 1):
                for c in range(min_col, max_col_m + 1):
                    if r != min_row or c != min_col:
                        rows[r - 1][c - 1] = top_left_val

        # 检测公式单元格并获取缓存值
        formula_cells_without_cache = []
        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                val = rows[r - 1][c - 1]
                if isinstance(val, str) and val.startswith("="):
                    # 有公式，尝试从 cached 版本获取缓存值
                    cached_val = None
                    if ws_cached:
                        cached_val = ws_cached.cell(row=r, column=c).value
                    if cached_val is not None:
                        rows[r - 1][c - 1] = cached_val
                    else:
                        rows[r - 1][c - 1] = "[公式未计算]"
                        formula_cells_without_cache.append((r, c))

        # 表头：第 1 行
        headers = [str(v).strip() if v is not None else f"列{_col_letter(c)}" for c, v in enumerate(rows[0], 1)]

        # 计算允许的锚定范围
        allowed = _compute_allowed_anchor_ranges(rows, max_row, max_col)

        ctx = SheetContext(
            sheet_name=sheet_name,
            max_row=max_row,
            max_col=max_col,
            headers=headers,
            rows=rows,
            merged_cells=merged_cells,
            formula_cells_without_cache=formula_cells_without_cache,
            format_warnings=[],
            allowed_anchor_ranges=allowed,
        )
        _split_into_table_blocks(ctx)
        contexts.append(ctx)

    wb.close()
    if wb_cached:
        wb_cached.close()
    return contexts


def parse_xls_to_sheet_contexts(file_bytes):
    """解析 XLS 文件为 SheetContext 列表。

    使用 pandas + xlrd。不支持公式检测和合并单元格。
    """
    try:
        sheets = pd.read_excel(io.BytesIO(file_bytes), sheet_name=None, engine='xlrd', dtype=str)
    except ImportError:
        raise ValueError("XLS 格式需要安装 xlrd 库。请运行: pip install xlrd")
    except Exception as e:
        raise ValueError(f"无法解析 XLS 文件: {e}")

    contexts = []
    for sheet_name, df in sheets.items():
        if df.empty:
            continue
        df = df.fillna("")
        headers = [str(c) for c in df.columns.tolist()]
        rows = [headers] + [[str(v) for v in row] for row in df.values.tolist()]
        max_row = len(rows)
        max_col = len(headers)

        allowed = _compute_allowed_anchor_ranges(rows, max_row, max_col)

        ctx = SheetContext(
            sheet_name=str(sheet_name),
            max_row=max_row,
            max_col=max_col,
            headers=headers,
            rows=rows,
            merged_cells=[],
            formula_cells_without_cache=[],
            format_warnings=["XLS 格式不保留公式信息，单元格显示值可能为缓存计算结果"],
            allowed_anchor_ranges=allowed,
        )
        _split_into_table_blocks(ctx)
        contexts.append(ctx)

    return contexts


def parse_csv_to_sheet_contexts(file_bytes, file_name=""):
    """解析 CSV 文件为 SheetContext 列表（单工作表 "CSV"）。

    支持 UTF-8、UTF-8 BOM、GBK 编码探测。
    """
    encoding = _detect_csv_encoding(file_bytes)
    try:
        df = pd.read_csv(io.BytesIO(file_bytes), encoding=encoding, dtype=str, keep_default_na=False)
    except Exception as e:
        raise ValueError(f"无法解析 CSV 文件 (编码 {encoding}): {e}")

    if df.empty:
        raise ValueError("CSV 文件为空或无有效数据")

    headers = [str(c) for c in df.columns.tolist()]
    rows = [headers] + [[str(v) for v in row] for row in df.values.tolist()]
    max_row = len(rows)
    max_col = len(headers)

    allowed = _compute_allowed_anchor_ranges(rows, max_row, max_col)

    ctx = SheetContext(
        sheet_name="CSV",
        max_row=max_row,
        max_col=max_col,
        headers=headers,
        rows=rows,
        merged_cells=[],
        formula_cells_without_cache=[],
        format_warnings=[],
        allowed_anchor_ranges=allowed,
    )
    _split_into_table_blocks(ctx)
    return [ctx]


# ─── 锚定范围计算 ─────────────────────────────────────────────────────────────

def _compute_allowed_anchor_ranges(rows, max_row, max_col):
    """扫描网格，找出所有连续非空矩形区域，返回允许的锚定范围列表。

    策略：逐行扫描，识别连续非空列组，输出每个连续区域的范围字符串。
    每个数据行（第 2 行起）单独作为一个允许范围，加上表头+数据的组合范围。
    """
    if max_row < 2 or max_col < 1:
        return []

    allowed = []

    # 每个数据行（第 2 行起）单独作为一个允许范围
    for r in range(2, max_row + 1):
        # 找该行的连续非空列组
        non_empty_cols = []
        for c in range(1, max_col + 1):
            val = rows[r - 1][c - 1] if c - 1 < len(rows[r - 1]) else None
            if val is not None and str(val).strip():
                non_empty_cols.append(c)

        if non_empty_cols:
            # 找连续段
            start = non_empty_cols[0]
            prev = non_empty_cols[0]
            for col in non_empty_cols[1:]:
                if col == prev + 1:
                    prev = col
                else:
                    allowed.append(_range_to_str(start, r, prev, r))
                    start = col
                    prev = col
            allowed.append(_range_to_str(start, r, prev, r))

    # 表头+数据的组合范围（整个连续非空区域）
    for c_start in range(1, max_col + 1):
        # 找连续非空列组
        if not any(
            rows[r][c_start - 1] is not None and str(rows[r][c_start - 1]).strip()
            for r in range(min(2, max_row), max_row)
        ):
            continue
        c_end = c_start
        while c_end + 1 <= max_col and any(
            rows[r][c_end] is not None and str(rows[r][c_end]).strip()
            for r in range(min(2, max_row), max_row)
        ):
            c_end += 1
        # 找行范围
        r_start = 1
        r_end = max_row
        range_str = _range_to_str(c_start, r_start, c_end, r_end)
        if range_str not in allowed:
            allowed.append(range_str)
        c_start = c_end + 1

    # 去重保持顺序
    seen = set()
    unique = []
    for r in allowed:
        if r not in seen:
            seen.add(r)
            unique.append(r)
    return unique


# ─── 表格块拆分 ───────────────────────────────────────────────────────────────

def _split_into_table_blocks(sheet_ctx):
    """将 SheetContext 的行拆分为多个 TableBlock，每块最多 _MAX_BLOCK_ROWS 行数据。

    对于表头行+数值行的费率/参数表模式，额外生成语义化二列块。
    """
    max_row = sheet_ctx.max_row
    max_col = sheet_ctx.max_col
    headers = sheet_ctx.headers

    if max_row <= 1:
        # 只有表头，1 个块
        block = TableBlock(
            block_index=0,
            markdown=_render_block_markdown(sheet_ctx.rows[:1], headers, 1, 1),
            row_range=(1, 1),
            col_range=(1, max_col),
            allowed_anchor_ranges=[],
            has_formula_warnings=False,
        )
        sheet_ctx.table_blocks = [block]
        return

    blocks = []
    block_idx = 0
    # 数据从第 2 行开始，每 _MAX_BLOCK_ROWS 行一块
    data_start = 2
    while data_start <= max_row:
        data_end = min(data_start + _MAX_BLOCK_ROWS - 1, max_row)
        # 包含表头行
        block_rows = [sheet_ctx.rows[0]] + sheet_ctx.rows[data_start - 1:data_end]

        # 计算该块的 allowed_anchor_ranges 子集
        block_allowed = []
        for r_str in sheet_ctx.allowed_anchor_ranges:
            bounds = _parse_range_str(r_str)
            if bounds is None:
                continue
            _, r_min, _, r_max = bounds
            if r_min >= data_start and r_max <= data_end:
                block_allowed.append(r_str)

        has_formula = any(
            data_start <= r <= data_end
            for r, _ in sheet_ctx.formula_cells_without_cache
        )

        block = TableBlock(
            block_index=block_idx,
            markdown=_render_block_markdown(block_rows, headers, data_start, data_end),
            row_range=(data_start, data_end),
            col_range=(1, max_col),
            allowed_anchor_ranges=block_allowed,
            has_formula_warnings=has_formula,
        )
        blocks.append(block)
        block_idx += 1

        # 检测表头行+数值行模式，生成语义化二列块
        data_rows = sheet_ctx.rows[data_start - 1:data_end]
        hv_pairs = _detect_header_value_row_pairs(data_rows, headers)
        for label_idx, value_idx in hv_pairs:
            if label_idx == -1:
                # headers 作为标签行
                excel_label_row = 1  # Excel 第 1 行是表头
                excel_value_row = data_start + value_idx
                pair_rows = [headers, data_rows[value_idx]]
            else:
                excel_label_row = data_start + label_idx
                excel_value_row = data_start + value_idx
                pair_rows = [data_rows[label_idx], data_rows[value_idx]]
            semantic_md, field_anchors = _render_semantic_block(
                pair_rows, headers, excel_label_row, sheet_ctx
            )
            if semantic_md and field_anchors:
                # 每个字段-数值对的 anchor 都是合法的
                pair_allowed = [anchor for _, anchor in field_anchors]
                # 加上整行范围
                pair_range = f"{_col_letter(1)}{excel_label_row}:{_col_letter(max_col)}{excel_value_row}"
                pair_allowed.append(pair_range)

                sem_block = TableBlock(
                    block_index=block_idx,
                    markdown=semantic_md,
                    row_range=(excel_label_row, excel_value_row),
                    col_range=(1, max_col),
                    allowed_anchor_ranges=pair_allowed,
                    has_formula_warnings=has_formula,
                )
                blocks.append(sem_block)
                block_idx += 1

        data_start = data_end + 1

    # 检测"标准表头 + 单条业务数据行"模式（价格题语义块）
    # 真正的字段名在 row 2（sheet_ctx.rows[1]），不是 row 1 的通用列名
    field_header_row_idx = 1  # 字段名行在 sheet_ctx.rows 中的索引
    field_headers = sheet_ctx.rows[field_header_row_idx]  # 真正的字段名行
    all_rows_from_field_header = sheet_ctx.rows[field_header_row_idx:]  # 从字段名行开始
    hb_pairs = _detect_header_business_row_pairs(all_rows_from_field_header, field_headers)
    field_header_excel_row = field_header_row_idx + 1  # Excel 中的字段名行号 = 2

    # 找出表头中的文本列（用于 header_context_range）
    text_cols = [c for c, h in enumerate(field_headers) if h and str(h).strip() and not _is_numeric_value(h)]
    header_context_range = None
    if text_cols:
        min_tc = min(text_cols) + 1
        max_tc = max(text_cols) + 1
        header_context_range = f"{_col_letter(min_tc)}{field_header_excel_row}:{_col_letter(max_tc)}{field_header_excel_row}"

    for _, biz_idx in hb_pairs:
        biz_row = all_rows_from_field_header[biz_idx]
        biz_excel_row = field_header_excel_row + biz_idx  # 转为 Excel 行号

        # 业务行锚点：仅包含业务行本身（如 B4:D4）
        if text_cols:
            min_c = min(text_cols) + 1
            max_c = max(text_cols) + 1
            biz_anchor = f"{_col_letter(min_c)}{biz_excel_row}:{_col_letter(max_c)}{biz_excel_row}"
        else:
            continue

        # 渲染语义块 Markdown（表头+业务行，仅用于 LLM 理解）
        hb_md = _render_header_business_markdown(field_headers, biz_row, text_cols)

        if hb_md:
            hb_block = TableBlock(
                block_index=block_idx,
                markdown=hb_md,
                row_range=(biz_excel_row, biz_excel_row),
                col_range=(1, max_col),
                allowed_anchor_ranges=[biz_anchor],
                has_formula_warnings=False,
                header_context_range=header_context_range,
            )
            blocks.append(hb_block)
            block_idx += 1

    sheet_ctx.table_blocks = blocks

    # 将语义块的 field anchors 添加到 sheet 级白名单，确保验证通过
    for block in blocks:
        if block.allowed_anchor_ranges:
            for anchor in block.allowed_anchor_ranges:
                if anchor not in sheet_ctx.allowed_anchor_ranges:
                    sheet_ctx.allowed_anchor_ranges.append(anchor)


def _render_block_markdown(rows, headers, data_start_row, data_end_row):
    """渲染一个表格块为 Markdown，带 Excel 行号列。"""
    lines = []

    # 表头行
    header_cells = [str(h) if h else "" for h in headers]
    lines.append("| 行号 | " + " | ".join(header_cells) + " |")
    lines.append("|---:|" + "|".join(["---"] * len(header_cells)) + "|")

    # 数据行（第 2 个元素起是数据行，对应 Excel 行 data_start_row）
    for i, row in enumerate(rows[1:], start=0):
        excel_row = data_start_row + i
        cells = []
        for v in row:
            if v is None:
                cells.append("")
            else:
                cells.append(str(v))
        # 补齐列数
        while len(cells) < len(header_cells):
            cells.append("")
        lines.append(f"| {excel_row} | " + " | ".join(cells[:len(header_cells)]) + " |")

    return "\n".join(lines)


# ─── 语义化渲染：表头行+数值行 费率/参数表 ─────────────────────────────────────

def _is_numeric_value(val):
    """判断值是否为数值（数字、百分比、货币等）。"""
    if val is None:
        return False
    s = str(val).strip()
    if not s:
        return False
    # 纯数字
    try:
        float(s.replace(",", ""))
        return True
    except ValueError:
        pass
    # 百分比
    if s.endswith("%") and s[:-1].replace(".", "").replace(",", "").isdigit():
        return True
    # 货币
    for prefix in ("¥", "$", "€", "￥"):
        if s.startswith(prefix):
            try:
                float(s[len(prefix):].replace(",", ""))
                return True
            except ValueError:
                pass
    return False


def _detect_header_value_row_pairs(rows, headers=None):
    """检测表头行+数值行的费率/参数表模式。

    模式：某行是字段名（文本），下一行是对应数值（数字）。
    常见于：费率表、参数表、配置表等。
    也检测 Excel 表头（headers 参数）与第一行数据的配对。

    Returns:
        list of (label_row_idx, value_row_idx) — 0-indexed in rows array
        其中 label_row_idx=-1 表示使用 headers 作为标签行
    """
    if len(rows) < 1:
        return []

    pairs = []

    # 检查 headers + rows[0] 是否构成表头+数值对（费率/参数表模式）
    # 严格条件：数据行中只有第一列是文本（行标签），其余列全部是数值
    if headers and len(rows) >= 1:
        row_value = rows[0]
        total_cols = max(len(headers), len(row_value))
        # 统计数据行中的文本列和数值列
        text_val_cols = []
        numeric_val_cols = []
        for c in range(total_cols):
            v = row_value[c] if c < len(row_value) else None
            if v is not None and str(v).strip():
                if _is_numeric_value(v):
                    numeric_val_cols.append(c)
                else:
                    text_val_cols.append(c)
        # 条件：只有 1 个文本列（行标签）+ 至少 2 个数值列
        # 且文本列对应的 header 也是文本
        # 且只有 1 行数据（单行费率表；多行由内部行对检测处理）
        if (len(text_val_cols) == 1 and len(numeric_val_cols) >= 2
                and len(rows) == 1):
            label_col = text_val_cols[0]
            h = headers[label_col] if label_col < len(headers) else None
            if h and str(h).strip() and not _is_numeric_value(h):
                pairs.append((-1, 0))

    # 检查 rows 内部的相邻行对
    i = 0
    while i < len(rows) - 1:
        # 跳过已作为 value 被配对的行
        if pairs and pairs[-1][1] == i:
            i += 1
            continue

        row_label = rows[i]
        row_value = rows[i + 1]
        total_cols = max(len(row_label), len(row_value))

        # 统计 label 行和 value 行的文本/数值列
        text_label = []
        text_val = []
        numeric_val = []
        for c in range(total_cols):
            lv = row_label[c] if c < len(row_label) else None
            vv = row_value[c] if c < len(row_value) else None
            if lv is not None and str(lv).strip() and not _is_numeric_value(lv):
                text_label.append(c)
            if isinstance(vv, str) and vv == "[公式未计算]":
                continue
            if vv is not None and str(vv).strip():
                if _is_numeric_value(vv):
                    numeric_val.append(c)
                else:
                    text_val.append(c)

        matched = False

        # 模式 1：value 行只有 1 个文本列（行标签）+ 至少 2 个数值列
        if len(text_val) == 1 and len(numeric_val) >= 2:
            label_col = text_val[0]
            lv = row_label[label_col] if label_col < len(row_label) else None
            if lv is not None and str(lv).strip() and not _is_numeric_value(lv):
                label_val_text = str(lv).strip()
                repeat_count = sum(
                    1 for r in rows
                    if label_col < len(r) and str(r[label_col]).strip() == label_val_text
                )
                if repeat_count >= 2:
                    matched = True

        # 模式 2：label 行文本列明显多于 value 行，且 value 行有足够数值列
        # 例如：label 行 12 个文本列，value 行 3 个文本列 + 10 个数值列
        if not matched and len(text_label) >= len(text_val) + 2 and len(numeric_val) >= 3:
            # 且 label 行的文本列在 value 行中大部分变为数值
            converted = sum(1 for c in text_label if c in numeric_val)
            if converted >= 3:
                matched = True

        if matched:
            pairs.append((i, i + 1))
            i += 2
            continue
        i += 1

    return pairs


def _detect_header_business_row_pairs(rows, headers):
    """检测"标准表头 + 单条业务数据行"模式。

    适用于报价表等结构：
    - Row 2: 字段表头（功能模块, 产品功能, 未税价, 项目经理, ...）
    - Row 3: 费率数值行（B/C/D 为空，E-M 有数值）
    - Row 4+: 业务数据行（B/C 有文本，D 有价格数值，E-M 有人数）

    rows 参数是 sheet_ctx.rows[1:]（即从第 2 行开始的数据）。
    rows[0] 是表头行，rows[1] 是第一行数据（可能是费率行），rows[2]+ 是业务行。

    返回：(header_row_idx, business_row_idx) 对列表，0-indexed in rows
    header_row_idx=-1 表示使用 headers 作为表头行。

    识别条件：
    - 业务行有至少 1 个文本列 + 至少 1 个数值列
    - 业务行的所有非空列都在表头有对应字段名
    - 业务行之前有间隔行（空行或数值行，或表头行本身）
    """
    if not headers or len(rows) < 3:
        return []

    # 识别表头中有字段名的列
    header_cols = set()
    for c, h in enumerate(headers):
        if h and str(h).strip():
            header_cols.add(c)

    if not header_cols:
        return []

    # 检查是否存在"间隔行"（表头与业务数据之间的数值行或空行）
    # 间隔行特征：
    # 1. 大部分列为空或纯数值（如费率行）
    # 2. 部分列重复表头文本、部分列有数值（如报价表的费率行）
    # 3. 表头文本列在该行中重复出现，且有独立数值列
    has_gap = False
    if len(rows) >= 3:
        row1 = rows[1]  # 表头后的第一行
        row1_non_empty = sum(1 for v in row1 if v is not None and str(v).strip())
        row1_numeric = sum(1 for v in row1 if v is not None and _is_numeric_value(v))
        # 统计表头有字段名的列中，该行有多少重复表头文本
        same_as_header = sum(
            1 for c in header_cols
            if c < len(row1) and row1[c] is not None and headers[c] is not None
            and str(row1[c]).strip() == str(headers[c]).strip()
        )
        # 统计表头有字段名的列中，该行有多少为空
        empty_in_header = sum(
            1 for c in header_cols
            if c >= len(row1) or row1[c] is None or not str(row1[c]).strip()
        )
        # 间隔行条件（满足任一即可）：
        # a) 大部分为空
        # b) 纯数值
        # c) 部分重复表头+部分数值（至少 2 个重复 + 至少 2 个数值）
        # d) 表头列中至少 20% 为空（如费率行的 B/C/D 列）
        has_gap = (
            (row1_non_empty < len(header_cols) * 0.5)
            or (row1_non_empty > 0 and row1_numeric == row1_non_empty)
            or (same_as_header >= 2 and row1_numeric >= 2)
            or (empty_in_header >= max(2, len(header_cols) * 0.2))
        )

    if not has_gap:
        return []

    # 找到第一个间隔行之后的所有业务行
    # 间隔行是 rows[1]（已确认），业务行从 rows[2] 开始
    pairs = []
    for i in range(2, len(rows)):
        row = rows[i]
        row_text = set()
        row_num = set()
        for c, v in enumerate(row):
            if v is None or not str(v).strip():
                continue
            if _is_numeric_value(v):
                row_num.add(c)
            else:
                row_text.add(c)

        # 条件：有文本列 + 有数值列，且所有非空列都在表头中有字段名
        non_empty = row_text | row_num
        if row_text and row_num and non_empty.issubset(header_cols):
            pairs.append((-1, i))

    return pairs


def _render_header_business_block(headers, business_row, header_excel_row, business_excel_row, sheet_ctx):
    """渲染"表头 + 单条业务数据行"为语义块（保留用于向后兼容）。"""
    text_cols = [c for c, h in enumerate(headers) if h and str(h).strip() and not _is_numeric_value(h)]
    if not text_cols:
        return None, []

    pairs = []
    for c in text_cols:
        field_name = str(headers[c]).strip()
        value = business_row[c] if c < len(business_row) else None
        if value is not None and str(value).strip():
            pairs.append((field_name, value, c))

    if not pairs:
        return None, []

    lines = ["| 字段 | 值 |", "|---|---|"]
    for field_name, value, _ in pairs:
        lines.append(f"| {field_name} | {value} |")

    field_anchors = []
    for field_name, value, c in pairs:
        col_letter = _col_letter(c + 1)
        anchor = f"{col_letter}{header_excel_row}:{col_letter}{business_excel_row}"
        field_anchors.append((field_name, anchor))

    return "\n".join(lines), field_anchors


def _render_header_business_markdown(headers, business_row, text_cols):
    """渲染"表头 + 业务行"Markdown（仅用于 LLM 理解，不用于锚定）。

    格式：
    | 字段 | 值 |
    |---|---|
    | 功能模块 | CICD工具规范... |
    | 产品功能 | 集成发布流水线梳理 |
    | 未税价（元） | 73900 |
    """
    pairs = []
    for c in text_cols:
        field_name = str(headers[c]).strip() if c < len(headers) and headers[c] else ""
        value = business_row[c] if c < len(business_row) else None
        if field_name and value is not None and str(value).strip():
            pairs.append((field_name, value))

    if not pairs:
        return None

    lines = ["| 字段 | 值 |", "|---|---|"]
    for field_name, value in pairs:
        lines.append(f"| {field_name} | {value} |")
    return "\n".join(lines)


def _render_semantic_block(rows, headers, data_start_row, sheet_ctx):
    """将表头行+数值行渲染为规范化二列 Markdown 表格。

    输入：rows[0] 是字段名行，rows[1] 是数值行。
    输出：| 字段名 | 数值 | 格式的 Markdown 表格。

    同时返回每个字段-数值对的 anchor_range（如 E2:E3）。
    如果第一列是行标签（非数值），自动跳过该列。
    """
    label_row = rows[0]
    value_row = rows[1]
    total_cols = max(len(label_row), len(value_row))

    # 检测第一列是否为行标签（标签行的第一列是文本，数值行的第一列也是文本/非数值）
    start_col = 0
    if total_cols > 1:
        label_first = label_row[0] if len(label_row) > 0 else None
        value_first = value_row[0] if len(value_row) > 0 else None
        if (label_first is not None and str(label_first).strip() and
                not _is_numeric_value(label_first) and
                value_first is not None and str(value_first).strip() and
                not _is_numeric_value(value_first)):
            start_col = 1  # 跳过第一列（行标签列）

    pairs = []  # (field_name, value, col_idx)
    for c in range(start_col, total_cols):
        label = label_row[c] if c < len(label_row) else None
        value = value_row[c] if c < len(value_row) else None
        # 只包含 label 是文本且 value 是数值的配对（真正的字段-数值对）
        # 跳过 label 和 value 都是文本的情况（如 "功能模块 | 功能模块"）
        if (label is not None and str(label).strip() and not _is_numeric_value(label)
                and value is not None and _is_numeric_value(value)):
            pairs.append((str(label).strip(), value, c))

    if not pairs:
        return None, []

    # 渲染二列 Markdown
    lines = []
    lines.append("| 字段名 | 数值 |")
    lines.append("|---|---|")
    for field_name, value, _ in pairs:
        value_str = str(value) if value is not None else ""
        if isinstance(value, str) and value == "[公式未计算]":
            value_str = "[公式未计算]"
        lines.append(f"| {field_name} | {value_str} |")

    # 生成每个字段-数值对的 anchor_range
    label_excel_row = data_start_row
    value_excel_row = data_start_row + 1
    field_anchors = []
    for field_name, value, col_idx in pairs:
        col_letter = _col_letter(col_idx + 1)
        # 单列锚定：label_row:value_row
        anchor = f"{col_letter}{label_excel_row}:{col_letter}{value_excel_row}"
        field_anchors.append((field_name, anchor))

    return "\n".join(lines), field_anchors


# ─── 题意-锚点一致性校验 ──────────────────────────────────────────────────────

# 数值/角色相关关键词
_NUMERIC_KEYWORDS = (
    "未税价", "价格", "报价", "费率", "投入", "人月", "人数", "工期",
    "比例", "金额", "配置", "人力", "开发", "工时",
)
_AGGREGATE_KEYWORDS = (
    "各角色", "各模块", "所有", "明细", "汇总", "总计", "配置清单",
    "人力配置", "开发投入",
)


def _extract_semantic_field_names(sheets):
    """从所有语义块中提取字段名集合。"""
    field_names = set()
    for sheet in sheets:
        for block in sheet.table_blocks:
            if block.block_index > 0:  # 语义块
                for line in block.markdown.split("\n"):
                    if line.startswith("|") and "字段名" not in line and "---" not in line:
                        parts = [p.strip() for p in line.split("|") if p.strip()]
                        if parts:
                            field_names.add(parts[0])
    return field_names


def _extract_semantic_anchors(sheets):
    """从所有语义块中提取 field anchor 集合。

    包括：
    - 费率表垂直锚点（如 E2:E3，2行）
    - 业务行锚点（如 B4:D4，1行，来自有 header_context_range 的语义块）
    """
    anchors = set()
    for sheet in sheets:
        for block in sheet.table_blocks:
            if block.block_index > 0:  # 语义块
                for anchor in block.allowed_anchor_ranges:
                    bounds = _parse_range_str(anchor)
                    if bounds:
                        min_col, min_row, max_col_r, max_row_r = bounds
                        row_count = max_row_r - min_row + 1
                        # 2 行：字段名+数值（如 E2:E3）
                        # 1 行 + 有 header_context_range：业务行锚点（如 B4:D4）
                        if row_count >= 2 or block.header_context_range:
                            anchors.add(anchor)
    return anchors


def _validate_question_anchor_consistency(question, semantic_field_names, semantic_anchors, sheets_by_name):
    """校验题意与锚点的一致性（通用结构检查）。

    Returns (is_valid, reason).
    """
    q_text = (question.get("question") or "").strip()
    anchor_range = (question.get("anchor_range") or "").strip()

    # 1. 检测聚合型题目
    for kw in _AGGREGATE_KEYWORDS:
        if kw in q_text:
            return False, f"聚合型题目（含'{kw}'），每题只能考一个知识点"

    # 2. 数值题必须使用语义块锚点
    is_numeric_question = any(kw in q_text for kw in _NUMERIC_KEYWORDS)

    if is_numeric_question and semantic_anchors:
        anchor_bounds = _parse_range_str(anchor_range)
        is_in_semantic = anchor_range in semantic_anchors
        if not is_in_semantic and anchor_bounds:
            a_min_col, a_min_row, a_max_col, a_max_row = anchor_bounds
            for sa in semantic_anchors:
                sa_bounds = _parse_range_str(sa)
                if sa_bounds:
                    s_min_col, s_min_row, s_max_col, s_max_row = sa_bounds
                    if (a_min_col >= s_min_col and a_max_col <= s_max_col and
                            a_min_row >= s_min_row and a_max_row <= s_max_row):
                        is_in_semantic = True
                        break
        if not is_in_semantic:
            return False, (
                f"数值题'{q_text}'的 anchor '{anchor_range}' "
                f"不在语义块中，数值类题只能使用语义块的字段名+数值锚点"
            )

    return True, ""


# ─── 汇总行检测 ──────────────────────────────────────────────────────────────

_SUMMARY_KEYWORDS = (
    "总计", "合计", "小计", "汇总", "总成本", "总费用", "总金额", "小合计",
    "total", "subtotal", "grand total", "sum", "Total",
)


def _is_summary_row(row_data):
    """检测行是否为汇总/合计行。

    判断逻辑：
    1. 第一个非空单元格包含汇总关键词
    2. 行中所有非空单元格都是数值（纯数值行无业务标识）
    """
    if not row_data:
        return False

    # 检查第一个非空单元格
    first_text = None
    for v in row_data:
        if v is not None and str(v).strip():
            first_text = str(v).strip()
            break
    if first_text:
        first_lower = first_text.lower()
        for kw in _SUMMARY_KEYWORDS:
            if kw.lower() in first_lower:
                return True

    # 纯数值行无业务标识（如 M52:N22 区域的总计行）
    non_empty = [v for v in row_data if v is not None and str(v).strip()]
    if not non_empty:
        return False
    text_cells = [v for v in non_empty if not _is_numeric_value(v)]
    if not text_cells:
        return True

    return False


def _range_overlaps_multicell_merge(min_row, min_col, max_row, max_col, merged_cells):
    """检查范围是否完全被跨行合并单元格覆盖。

    只有当范围的所有列都在某个跨行合并区域内时才返回 True。
    部分重叠（如 B3:C3 中 B 列合并但 C 列不合并）不拒绝。
    """
    for mr_min, mc_min, mr_max, mc_max in merged_cells:
        if mr_max <= mr_min:
            continue  # 非跨行合并
        # 检查行重叠
        if min_row > mr_max or max_row < mr_min:
            continue
        # 检查范围的所有列是否都在合并区域内
        if min_col >= mc_min and max_col <= mc_max:
            return True
    return False


# ─── 候选锚点构建 ─────────────────────────────────────────────────────────────

_MAX_TEXT_FACT_ANCHOR_COLS = 6  # text_fact 候选最大列数，防止 D23:N23 类宽范围


def _detect_header_row_and_data_start(sheet):
    """从 SheetContext 中识别表头行和业务数据起始行。

    Returns:
        (header_row_idx, data_start_row, rate_row_idx)
        header_row_idx: 0-indexed 表头行索引（含最多文本值的行）
        data_start_row: 1-indexed 业务数据起始行号
        rate_row_idx: 0-indexed 费率行索引，无则 None
    """
    if not sheet.rows or len(sheet.rows) < 2:
        return 0, 2, None

    # 表头行 = 前两行中含更多独立文本值的行
    # （Row 1 可能是通用列名如"列A"，Row 2 可能是真正字段名如"功能模块"）
    # 过滤掉纯占位符（如"列A"、"列1"等单字前缀+数字/字母的模式）
    _PLACEHOLDER_RE = re.compile(r'^[A-Za-z\u4e00-\u9fff]{1,2}[\dA-Za-z]+$')

    def _meaningful_diversity(row):
        texts = set()
        for v in row:
            if v is None:
                continue
            s = str(v).strip()
            if not s or _is_numeric_value(v):
                continue
            if _PLACEHOLDER_RE.match(s):
                continue  # 跳过占位符
            texts.add(s)
        return len(texts)

    row1_div = _meaningful_diversity(sheet.rows[0])
    row2_div = _meaningful_diversity(sheet.rows[1]) if len(sheet.rows) >= 2 else 0
    header_row_idx = 0 if row1_div >= row2_div else 1

    header_row = sheet.rows[header_row_idx]
    text_col_indices = [
        c for c, v in enumerate(header_row)
        if v is not None and str(v).strip() and not _is_numeric_value(v)
    ]

    if not text_col_indices:
        return 0, 2, None

    # 检查下一行是否为费率行：
    # 条件 — 与表头行在文本列上有相同文本值，且有额外数值列
    rate_row_idx = None
    next_row_idx = header_row_idx + 1
    if next_row_idx < len(sheet.rows):
        row2 = sheet.rows[next_row_idx]
        same_text_count = sum(
            1 for c in text_col_indices
            if c < len(row2) and row2[c] is not None
            and str(row2[c]).strip() == str(header_row[c]).strip()
        )
        numeric_count = sum(
            1 for c, v in enumerate(row2)
            if v is not None and _is_numeric_value(v)
        )
        # 费率行条件：文本列大部分重复表头 + 有数值列
        if same_text_count >= len(text_col_indices) * 0.5 and numeric_count >= 2:
            rate_row_idx = next_row_idx
            data_start_row = next_row_idx + 2  # 跳过表头 + 费率行
        else:
            data_start_row = next_row_idx + 1  # 无费率行，数据从表头下一行开始

    return header_row_idx, data_start_row, rate_row_idx


def _build_record_candidates_for_sheet(sheet, header_row_idx, data_start_row):
    """为单个工作表生成 record_with_schema_context 候选。

    直接扫描业务行，不依赖 block.header_context_range。
    """
    candidates = []
    header_row = sheet.rows[header_row_idx]
    header_excel_row = header_row_idx + 1

    # 找出表头中的文本列（用于 header_context_range 和列范围）
    text_col_indices = [
        c for c, v in enumerate(header_row)
        if v is not None and str(v).strip() and not _is_numeric_value(v)
    ]
    if not text_col_indices:
        return candidates

    h_min_col = min(text_col_indices) + 1  # 1-indexed
    h_max_col = max(text_col_indices) + 1
    header_context_range = f"{_col_letter(h_min_col)}{header_excel_row}:{_col_letter(h_max_col)}{header_excel_row}"

    # 扫描业务行
    for row_idx in range(data_start_row - 1, len(sheet.rows)):
        excel_row = row_idx + 1
        row_data = sheet.rows[row_idx]

        # 跳过空行
        non_empty = [v for v in row_data if v is not None and str(v).strip()]
        if not non_empty:
            continue

        # 跳过汇总行
        if _is_summary_row(row_data):
            continue

        # 跳过费率行和重复表头行：
        # 统计该行在表头文本列中有多少列的值与表头完全相同
        same_as_header = sum(
            1 for c in text_col_indices
            if c < len(row_data) and row_data[c] is not None
            and str(row_data[c]).strip() == str(header_row[c]).strip()
        )
        # 检查是否为费率行：表头文本列中出现数值，且该行有与表头匹配的文本
        # 费率行特征：部分文本列与表头相同 + 其他文本列有数值（如 1600, 2200）
        numeric_in_text_cols = sum(
            1 for c in text_col_indices
            if c < len(row_data) and row_data[c] is not None
            and _is_numeric_value(row_data[c])
        )
        if same_as_header >= 2 and numeric_in_text_cols >= 2:
            continue  # 费率行（表头子集 + 文本列中的数值）

        if same_as_header >= len(text_col_indices) * 0.7:
            continue  # 重复表头行（≥70% 文本列与表头相同）

        # 在表头文本列范围内，找连续文本列分组
        text_col_groups = []
        group_start = None
        for c in range(h_min_col, h_max_col + 1):
            v = row_data[c - 1] if c - 1 < len(row_data) else None
            is_text = v is not None and str(v).strip() and not _is_numeric_value(v)
            if is_text:
                if group_start is None:
                    group_start = c
            else:
                if group_start is not None:
                    text_col_groups.append((group_start, c - 1))
                    group_start = None
        if group_start is not None:
            text_col_groups.append((group_start, h_max_col))

        if not text_col_groups:
            continue  # 无文本列的行（纯数值行）不生成候选

        # 为每个连续文本列分组生成子范围候选
        sub_ranges = []
        for g_start, g_end in text_col_groups:
            sub_ranges.append((g_start, g_end))
            # 扩展：包含紧随其后的第一个数值列
            for c in range(g_end + 1, h_max_col + 1):
                v = row_data[c - 1] if c - 1 < len(row_data) else None
                if v is not None and _is_numeric_value(v):
                    sub_ranges.append((g_start, c))
                    break
                else:
                    break

        # 去重并生成候选
        seen = set()
        for sr_start, sr_end in sub_ranges:
            sr_key = (sr_start, sr_end)
            if sr_key in seen:
                continue
            seen.add(sr_key)

            if _range_overlaps_multicell_merge(
                excel_row, sr_start, excel_row, sr_end, sheet.merged_cells
            ):
                continue

            sub_anchor = f"{_col_letter(sr_start)}{excel_row}:{_col_letter(sr_end)}{excel_row}"
            fact_fields = []
            focus_parts = []
            for c in range(sr_start, sr_end + 1):
                h_val = header_row[c - 1] if c - 1 < len(header_row) else None
                fname = str(h_val).strip() if h_val else ""
                if not fname or _is_numeric_value(h_val):
                    continue
                fact_fields.append(fname)
                v = row_data[c - 1] if c - 1 < len(row_data) else None
                if v is not None and str(v).strip():
                    focus_parts.append(str(v).strip())

            if fact_fields and focus_parts:
                candidates.append({
                    "evidence_mode": "record_with_schema_context",
                    "sheet_name": sheet.sheet_name,
                    "anchor_range": sub_anchor,
                    "header_context_range": header_context_range,
                    "fact_fields": fact_fields,
                    "query_focus": " | ".join(focus_parts[:4]),
                })

    return candidates


def _build_candidate_anchors(sheets):
    """为 LLM 构建合法候选证据清单（统一格式）。

    Returns:
        list[dict]: 每项含 evidence_mode, sheet_name, anchor_range,
                    header_context_range, fact_fields, query_focus
    """
    candidates = []

    for sheet in sheets:
        # ── 通用 record_with_schema_context：自动检测表头行 ──
        header_row_idx, data_start_row, _ = _detect_header_row_and_data_start(sheet)
        record_cands = _build_record_candidates_for_sheet(sheet, header_row_idx, data_start_row)
        candidates.extend(record_cands)

        # ── 基于块的候选（field_value_pair, text_fact, 及已有 header_context_range 的块） ──
        for block in sheet.table_blocks:
            if block.block_index == 0:
                continue  # 跳过标准块

            if block.header_context_range:
                # ── record_with_schema_context：带表头上下文的业务行记录 ──
                h_bounds = _parse_range_str(block.header_context_range)
                if not h_bounds:
                    continue
                h_min_col, h_min_row, h_max_col, _ = h_bounds

                # 每个 allowed_anchor_range 生成独立候选，不合并
                for anchor in block.allowed_anchor_ranges:
                    bounds = _parse_range_str(anchor)
                    if not bounds:
                        continue
                    a_min_col, a_min_row, a_max_col, a_max_row = bounds
                    if a_max_row != a_min_row:
                        continue

                    row_data = sheet.rows[a_min_row - 1] if a_min_row - 1 < len(sheet.rows) else []

                    # 跳过汇总/合计行
                    if _is_summary_row(row_data):
                        continue

                    # 扫描业务行，找出 anchor 范围内连续文本列的分组
                    # 每个分组生成一个子范围候选
                    text_col_groups = []  # list of (start_col, end_col)
                    group_start = None
                    for c in range(a_min_col, a_max_col + 1):
                        v = row_data[c - 1] if c - 1 < len(row_data) else None
                        is_text = v is not None and str(v).strip() and not _is_numeric_value(v)
                        if is_text:
                            if group_start is None:
                                group_start = c
                        else:
                            if group_start is not None:
                                text_col_groups.append((group_start, c - 1))
                                group_start = None
                    if group_start is not None:
                        text_col_groups.append((group_start, a_max_col))

                    # 如果没有找到文本分组，回退到完整 anchor 范围
                    if not text_col_groups:
                        text_col_groups = [(a_min_col, a_max_col)]

                    # 为每个连续文本列分组生成候选
                    # 同时生成包含紧邻数值列的扩展候选（身份+数值模式）
                    sub_ranges = []
                    for g_start, g_end in text_col_groups:
                        # 纯文本子范围
                        sub_ranges.append((g_start, g_end))
                        # 扩展：包含紧随其后的第一个数值列（身份+主值模式）
                        first_numeric_end = None
                        for c in range(g_end + 1, a_max_col + 1):
                            v = row_data[c - 1] if c - 1 < len(row_data) else None
                            if v is not None and _is_numeric_value(v):
                                first_numeric_end = c
                                break
                            else:
                                break
                        if first_numeric_end is not None:
                            sub_ranges.append((g_start, first_numeric_end))

                    # 完整 anchor 范围也作为候选（用于需要多列数值的查询）
                    full_key = (a_min_col, a_max_col)
                    if full_key not in {(s, e) for s, e in sub_ranges}:
                        sub_ranges.append(full_key)

                    # 去重并生成候选
                    seen = set()
                    for sr_start, sr_end in sub_ranges:
                        sr_key = (sr_start, sr_end)
                        if sr_key in seen:
                            continue
                        seen.add(sr_key)

                        # 跳过与跨行合并单元格重叠的子范围
                        if _range_overlaps_multicell_merge(a_min_row, sr_start, a_min_row, sr_end, sheet.merged_cells):
                            continue

                        sub_anchor = f"{_col_letter(sr_start)}{a_min_row}:{_col_letter(sr_end)}{a_min_row}"
                        fact_fields = []
                        focus_parts = []
                        for c in range(sr_start, sr_end + 1):
                            h_val = None
                            if h_min_col <= c <= h_max_col:
                                h_val = sheet.rows[h_min_row - 1][c - 1] if c - 1 < len(sheet.rows[h_min_row - 1]) else None
                            fname = str(h_val).strip() if h_val else ""
                            if not fname or _is_numeric_value(h_val):
                                continue
                            fact_fields.append(fname)
                            v = row_data[c - 1] if c - 1 < len(row_data) else None
                            if v is not None and str(v).strip():
                                focus_parts.append(str(v).strip())

                        if fact_fields and focus_parts:
                            candidates.append({
                                "evidence_mode": "record_with_schema_context",
                                "sheet_name": sheet.sheet_name,
                                "anchor_range": sub_anchor,
                                "header_context_range": block.header_context_range,
                                "fact_fields": fact_fields,
                                "query_focus": " | ".join(focus_parts[:4]),
                            })
            else:
                # ── field_value_pair：字段名+数值垂直对 ──
                for anchor in block.allowed_anchor_ranges:
                    bounds = _parse_range_str(anchor)
                    if not bounds:
                        continue
                    min_col, min_row, max_col, max_row = bounds
                    row_count = max_row - min_row + 1
                    if row_count != 2:
                        continue
                    # 跳过与跨行合并单元格重叠的范围
                    if _range_overlaps_multicell_merge(min_row, min_col, max_row, max_col, sheet.merged_cells):
                        continue
                    field_name = ""
                    value = ""
                    row_values = []
                    for r in range(min_row, max_row + 1):
                        for c in range(min_col, max_col + 1):
                            val = sheet.rows[r - 1][c - 1] if c - 1 < len(sheet.rows[r - 1]) else None
                            if val is None:
                                continue
                            val_str = str(val).strip()
                            if not val_str or val_str == "[公式未计算]":
                                continue
                            row_values.append(val_str)
                            if _is_numeric_value(val):
                                value = val_str
                            else:
                                field_name = val_str
                    # 跳过含汇总关键词的行
                    if any(kw in v for v in row_values for kw in _SUMMARY_KEYWORDS):
                        continue
                    if field_name and value:
                        candidates.append({
                            "evidence_mode": "field_value_pair",
                            "sheet_name": sheet.sheet_name,
                            "anchor_range": anchor,
                            "header_context_range": None,
                            "fact_fields": [field_name],
                            "query_focus": f"{field_name}：{value}",
                        })

        # ── text_fact：标准块中的多列文本事实 ──
        for block in sheet.table_blocks:
            if block.block_index != 0:
                continue
            for anchor in block.allowed_anchor_ranges:
                bounds = _parse_range_str(anchor)
                if not bounds:
                    continue
                min_col, min_row, max_col, max_row = bounds
                if max_row != min_row:
                    continue
                row_data = sheet.rows[min_row - 1] if min_row - 1 < len(sheet.rows) else []
                if _is_summary_row(row_data):
                    continue
                if _range_overlaps_multicell_merge(min_row, min_col, max_row, max_col, sheet.merged_cells):
                    continue
                text_cols = []
                header_names = []
                non_empty_count = 0
                for c in range(min_col, max_col + 1):
                    val = sheet.rows[min_row - 1][c - 1] if c - 1 < len(sheet.rows[min_row - 1]) else None
                    if val is not None and str(val).strip():
                        non_empty_count += 1
                    if val and not _is_numeric_value(val) and str(val).strip():
                        text_cols.append(str(val).strip())
                    # 读取表头名
                    hdr = sheet.rows[0][c - 1] if c - 1 < len(sheet.rows[0]) else None
                    if hdr and str(hdr).strip():
                        header_names.append(str(hdr).strip())
                # 跳过过宽的范围（防止 D23:N23 类候选）
                anchor_col_span = max_col - min_col + 1
                if anchor_col_span > _MAX_TEXT_FACT_ANCHOR_COLS:
                    continue
                if len(text_cols) >= 1 and non_empty_count >= 2:
                    candidates.append({
                        "evidence_mode": "text_fact",
                        "sheet_name": sheet.sheet_name,
                        "anchor_range": anchor,
                        "header_context_range": None,
                        "fact_fields": header_names if header_names else text_cols[:3],
                        "query_focus": " | ".join(text_cols[:3]),
                    })

    return candidates


# ─── Prompt 构建 ──────────────────────────────────────────────────────────────

def _build_prompt(sheets, num_questions, topic_hint="", candidate_anchors=None):
    """构建发给 LLM 的完整 prompt。"""
    template = PROMPT_TEMPLATE_PATH.read_text(encoding="utf-8").strip()

    topic_hint_section = ""
    if topic_hint:
        topic_hint_section = f"- 主题方向：{topic_hint}"

    # 渲染所有表格块
    block_texts = []
    for sheet in sheets:
        for block in sheet.table_blocks:
            formula_warn = ""
            if block.has_formula_warnings:
                formula_warn = "\n\n⚠️ 本块包含公式单元格（无缓存计算值），请避免引用这些单元格作为核心证据。"
            header_ctx_info = ""
            if block.header_context_range:
                header_ctx_info = f"\nheader_context_range: {block.header_context_range}"
            block_text = (
                f"### 工作表: {sheet.sheet_name} — 表格块 {block.block_index + 1} "
                f"(行 {block.row_range[0]}-{block.row_range[1]})\n\n"
                f"{block.markdown}\n\n"
                f"allowed_anchor_ranges: {json.dumps(block.allowed_anchor_ranges, ensure_ascii=False)}"
                f"{header_ctx_info}"
                f"{formula_warn}"
            )
            block_texts.append(block_text)

    table_blocks_text = "\n\n---\n\n".join(block_texts)

    # 构建候选证据清单
    candidate_text = ""
    if candidate_anchors:
        _MODE_LABELS = {
            "record_with_schema_context": "带表头上下文的记录证据（anchor_range 覆盖单行业务行，header_context_range 自动附带）",
            "field_value_pair": "字段-数值对证据（anchor_range 包含字段名和数值，如参数表、费率表）",
            "text_fact": "文本事实证据（anchor_range 覆盖连续文本列）",
        }
        parts = []
        for mode, label in _MODE_LABELS.items():
            items = [a for a in candidate_anchors if a.get("evidence_mode") == mode]
            if not items:
                continue
            lines = [f"**{label}**："]
            for a in items[:20]:
                item_parts = [
                    f"**sheet_name**: {a['sheet_name']}",
                    f"**anchor_range**: `{a['anchor_range']}`",
                ]
                if a.get("header_context_range"):
                    item_parts.append(f"**header_context_range**: `{a['header_context_range']}`")
                if a.get("fact_fields"):
                    item_parts.append(f"**fact_fields**: [{', '.join(a['fact_fields'])}]")
                item_parts.append(f"**query_focus**: {a['query_focus']}")
                line = "- " + " | ".join(item_parts)
                lines.append(line)
            parts.append("\n".join(lines))

        if parts:
            candidate_text = "\n\n---\n\n## 合法候选证据清单\n\n" + "\n\n".join(parts)

    prompt = template.replace("{table_blocks_text}", table_blocks_text)
    prompt = prompt.replace("{num_questions}", str(num_questions))
    prompt = prompt.replace("{topic_hint_section}", topic_hint_section)
    prompt = prompt.replace("{candidate_anchors_section}", candidate_text)
    return prompt


# ─── LLM 响应解析 ─────────────────────────────────────────────────────────────

def _parse_llm_response(text):
    """解析 LLM 返回的 JSON 数组。"""
    text = text.strip()

    # 去除 markdown 代码块
    if text.startswith("```"):
        lines = text.split("\n")
        json_lines = []
        in_block = False
        for line in lines:
            if line.strip().startswith("```") and not in_block:
                in_block = True
                continue
            elif line.strip() == "```" and in_block:
                break
            elif in_block:
                json_lines.append(line)
        text = "\n".join(json_lines).strip()

    # 尝试直接解析
    try:
        data = json.loads(text)
        if isinstance(data, list):
            return data
    except json.JSONDecodeError:
        pass

    # 尝试提取 JSON 数组
    match = re.search(r'\[[\s\S]*\]', text)
    if match:
        try:
            data = json.loads(match.group())
            if isinstance(data, list):
                return data
        except json.JSONDecodeError:
            pass

    return []


# ─── 锚定范围验证 ─────────────────────────────────────────────────────────────

def _validate_anchor_range(anchor_range, allowed_ranges, max_row, max_col, sheet_ctx=None):
    """验证锚定范围是否合法。

    规则：
    1. 范围必须是某个白名单范围的子集
    2. 单行范围：直接通过
    3. 跨行范围：只允许表头(row 1) + 紧邻的一行数据(row 2)
       - 不允许跨越多条业务数据行
       - 语义块的 field anchors（如 B1:B2）自动满足此规则

    Returns (is_valid, reason).
    """
    bounds = _parse_range_str(anchor_range)
    if bounds is None:
        return False, f"无法解析范围 '{anchor_range}'"

    min_col, min_row, max_col_r, max_row_r = bounds

    if min_row < 1 or min_col < 1:
        return False, f"范围 {anchor_range} 起始位置越界"
    if max_row_r > max_row or max_col_r > max_col:
        return False, f"范围 {anchor_range} 超出工作表边界 (max_row={max_row}, max_col={max_col})"

    rows = max_row_r - min_row + 1
    cols = max_col_r - min_col + 1
    if rows > _MAX_EVIDENCE_ROWS:
        return False, f"证据范围行数 ({rows}) 超过上限 {_MAX_EVIDENCE_ROWS}"
    if cols > _MAX_EVIDENCE_COLS:
        return False, f"证据范围列数 ({cols}) 超过上限 {_MAX_EVIDENCE_COLS}"

    # 跨行范围收紧：
    # - 2 行：允许（子集检查即可，如字段名+数值、表头+业务行）
    # - 3+ 行：必须在白名单中明确出现（来自语义块）
    if rows >= 3 and anchor_range not in allowed_ranges:
        return False, (
            f"跨行范围 {anchor_range} 不在白名单中，"
            f"3行以上跨行锚点必须来自语义块"
        )

    # 检查白名单：范围必须是某个白名单范围的子集
    is_subset = False
    for allowed in allowed_ranges:
        allowed_bounds = _parse_range_str(allowed)
        if allowed_bounds is None:
            continue
        a_min_col, a_min_row, a_max_col, a_max_row = allowed_bounds
        if (min_col >= a_min_col and max_col_r <= a_max_col and
                min_row >= a_min_row and max_row_r <= a_max_row):
            is_subset = True
            break
    if not is_subset:
        return False, f"范围 '{anchor_range}' 不在白名单范围内"

    # 孤立数值检测：如果范围只覆盖一个或多个纯数值单元格，拒绝
    # 双源模型锚点也不例外——D4:D4 等孤立价格数值必须继续拒绝
    if sheet_ctx is not None:
        all_numeric = True
        has_any_value = False
        for r in range(min_row, max_row_r + 1):
            for c in range(min_col, max_col_r + 1):
                val = sheet_ctx.rows[r - 1][c - 1] if c - 1 < len(sheet_ctx.rows[r - 1]) else None
                if val is not None and str(val).strip():
                    has_any_value = True
                    if not _is_numeric_value(val):
                        all_numeric = False
                        break
            if not all_numeric:
                break
        if has_any_value and all_numeric:
            return False, (
                f"范围 {anchor_range} 只包含孤立数值，"
                f"数值类证据必须同时包含字段名称和对应数值"
            )

    return True, ""


# ─── 金标准渲染 ───────────────────────────────────────────────────────────────

def _render_reference_answer(anchor_range, sheet_ctx):
    """从 SheetContext 的 rows 数据中读取指定范围，渲染为 Markdown。

    Returns (rendered_text, has_formula_issue).
    """
    bounds = _parse_range_str(anchor_range)
    if bounds is None:
        return "", True

    min_col, min_row, max_col, max_row = bounds

    cell_values = []
    has_formula_issue = False
    for r in range(min_row, max_row + 1):
        row_vals = []
        for c in range(min_col, max_col + 1):
            val = sheet_ctx.rows[r - 1][c - 1] if c - 1 < len(sheet_ctx.rows[r - 1]) else None
            # 检查是否为公式单元格无缓存值
            if (r, c) in sheet_ctx.formula_cells_without_cache:
                has_formula_issue = True
                row_vals.append("[公式未计算]")
            elif isinstance(val, str) and val == "[公式未计算]":
                has_formula_issue = True
                row_vals.append(val)
            elif val is None:
                row_vals.append(None)
            else:
                row_vals.append(val)
        cell_values.append(row_vals)

    # 检查非空
    non_empty = sum(1 for row in cell_values for v in row if v is not None and str(v).strip())
    if non_empty == 0:
        return "", True

    rendered = _render_cell_values(cell_values)
    return rendered, has_formula_issue


def _render_cell_values(cell_values):
    """将二维单元格值数组渲染为纯文本格式（不含 Markdown 分隔符）。

    单行：值1 | 值2 | 值3
    多行（首行是表头）：字段1：值1；字段2：值2；...
    """
    if not cell_values:
        return ""

    # 单行：值用 | 分隔
    if len(cell_values) == 1:
        parts = [str(v) for v in cell_values[0] if v is not None and str(v).strip()]
        return " | ".join(parts) if parts else ""

    # 多行：首行是表头，后续行是数据 → 键值对格式
    header = cell_values[0]
    lines = []
    for row in cell_values[1:]:
        pairs = []
        for i, val in enumerate(row):
            if val is None or not str(val).strip():
                continue
            h = header[i] if i < len(header) and header[i] is not None else f"列{i+1}"
            pairs.append(f"{str(h).strip()}：{str(val).strip()}")
        if pairs:
            lines.append("；".join(pairs))

    return "\n".join(lines) if lines else ""


def _render_dual_source_reference(anchor_range, header_context_range, sheet_ctx):
    """双源渲染：表头上下文 + 业务行数据。

    按列对齐：只渲染 anchor 范围内每列对应的字段名和值。
    例如 anchor=D4:D4, header_ctx=B2:M2 → 只渲染 D 列：未税价（元）：73900

    Returns (rendered_text, has_formula_issue).
    """
    # 解析表头上下文范围（如 B2:M2）
    h_bounds = _parse_range_str(header_context_range)
    if h_bounds is None:
        return "", True
    h_min_col, h_min_row, h_max_col, h_max_row = h_bounds

    # 解析业务行范围（如 D4:D4）
    b_bounds = _parse_range_str(anchor_range)
    if b_bounds is None:
        return "", True
    b_min_col, b_min_row, b_max_col, b_max_row = b_bounds

    has_formula_issue = False
    lines = []

    # 按列对齐：遍历 anchor 的每一列，找对应的表头字段名
    for c in range(b_min_col, b_max_col + 1):
        # 读取表头字段名（从 header_context_range 中对应列）
        if h_min_col <= c <= h_max_col:
            fname = sheet_ctx.rows[h_min_row - 1][c - 1] if c - 1 < len(sheet_ctx.rows[h_min_row - 1]) else None
            fname = str(fname).strip() if fname else ""
        else:
            fname = ""

        # 读取业务行值
        val = sheet_ctx.rows[b_min_row - 1][c - 1] if c - 1 < len(sheet_ctx.rows[b_min_row - 1]) else None
        if (b_min_row, c) in sheet_ctx.formula_cells_without_cache:
            has_formula_issue = True
            fval = "[公式未计算]"
        elif isinstance(val, str) and val == "[公式未计算]":
            has_formula_issue = True
            fval = val
        elif val is None:
            fval = ""
        else:
            fval = str(val)

        if fname and fval:
            lines.append(f"{fname}：{fval}")

    rendered = "\n".join(lines)
    if not rendered.strip():
        return "", True

    return rendered, has_formula_issue


# ─── 题目验证与渲染 ───────────────────────────────────────────────────────────

def _build_evidence_schema_display(anchor_range, header_context_range, sheet_ctx):
    """构建供人阅读的字段=值 显示（不参与 Judge 匹配）。

    示例: "模块=CICD；一级需求=持续优化；需求描述=工具链可用性优化；类型=Improvement"
    """
    bounds = _parse_range_str(anchor_range)
    if bounds is None:
        return ""

    min_col, min_row, max_col, max_row = bounds

    # 获取表头行
    if header_context_range:
        h_bounds = _parse_range_str(header_context_range)
        if h_bounds:
            _, h_row, _, _ = h_bounds
            header_row_data = sheet_ctx.rows[h_row - 1] if h_row - 1 < len(sheet_ctx.rows) else []
        else:
            header_row_data = sheet_ctx.rows[0] if sheet_ctx.rows else []
    else:
        header_row_data = sheet_ctx.rows[0] if sheet_ctx.rows else []

    # 获取业务行数据
    data_row = sheet_ctx.rows[min_row - 1] if min_row - 1 < len(sheet_ctx.rows) else []

    pairs = []
    for c in range(min_col, max_col + 1):
        h_val = header_row_data[c - 1] if c - 1 < len(header_row_data) else None
        header_name = str(h_val).strip() if h_val else ""
        d_val = data_row[c - 1] if c - 1 < len(data_row) else None
        data_str = str(d_val).strip() if d_val is not None else ""
        if header_name and data_str:
            pairs.append(f"{header_name}={data_str}")

    return "；".join(pairs)


def _validate_and_render_question(raw_q, sheets_by_name, file_name):
    """验证 LLM 返回的单条题目，渲染金标准证据。

    支持双源模型：
    - 业务行锚点（如 B4:D4）：唯一业务事实来源
    - 表头上下文（如 B2:D2）：字段语义来源，从语义块元数据获取

    Returns (validated_dict_or_None, rejection_reason).
    """
    q_text = (raw_q.get("question") or "").strip()
    sheet_name = (raw_q.get("sheet_name") or "").strip()
    anchor_range = (raw_q.get("anchor_range") or "").strip()

    if not q_text:
        return None, "query 为空"
    if not sheet_name:
        return None, "sheet_name 为空"
    if not anchor_range:
        return None, "anchor_range 为空"

    # 查找工作表
    sheet = sheets_by_name.get(sheet_name)
    if sheet is None:
        return None, f"工作表 '{sheet_name}' 不存在，可用: {list(sheets_by_name.keys())}"

    # 验证范围
    valid, reason = _validate_anchor_range(
        anchor_range, sheet.allowed_anchor_ranges, sheet.max_row, sheet.max_col,
        sheet_ctx=sheet,
    )
    if not valid:
        return None, reason

    # 查找该锚点对应的语义块，获取 header_context_range（支持子集匹配）
    header_context_range = None
    anchor_bounds = _parse_range_str(anchor_range)
    for block in sheet.table_blocks:
        if not block.header_context_range:
            continue
        for allowed in block.allowed_anchor_ranges:
            if anchor_range == allowed:
                header_context_range = block.header_context_range
                break
            allowed_bounds = _parse_range_str(allowed)
            if anchor_bounds and allowed_bounds:
                a_min_col, a_min_row, a_max_col, a_max_row = anchor_bounds
                s_min_col, s_min_row, s_max_col, s_max_row = allowed_bounds
                if (a_min_col >= s_min_col and a_max_col <= s_max_col and
                        a_min_row >= s_min_row and a_max_row <= s_max_row):
                    header_context_range = block.header_context_range
                    break
        if header_context_range:
            break

    # 渲染金标准
    if header_context_range:
        # 双源渲染：表头上下文 + 业务行
        rendered, has_formula_issue = _render_dual_source_reference(
            anchor_range, header_context_range, sheet
        )
    else:
        # 单源渲染：仅业务行
        rendered, has_formula_issue = _render_reference_answer(anchor_range, sheet)

    if not rendered:
        return None, "证据范围为空"
    if has_formula_issue:
        return None, "证据范围包含公式单元格（无缓存计算值）"

    # 组装验证后的题目
    validated = {
        "question": q_text,
        "reference_answer": rendered,
        "source_excerpt": rendered,
        "sheet_name": sheet_name,
        "anchor_range": anchor_range,
        "evidence_sheet": sheet_name,       # 向后兼容
        "evidence_range": anchor_range,     # 向后兼容
        "source_format": _detect_format(file_name),
        "source_file_name": file_name,
        "difficulty": raw_q.get("difficulty", "事实"),
        "topic": raw_q.get("topic", ""),
    }

    # 添加表头上下文（如有）
    if header_context_range:
        validated["header_context_range"] = header_context_range

    # 添加供人阅读的 evidence_schema_display
    validated["evidence_schema_display"] = _build_evidence_schema_display(
        anchor_range, header_context_range, sheet,
    )

    return validated, ""


def _detect_format(file_name):
    """从文件名推断格式。"""
    ext = Path(file_name).suffix.lower()
    if ext == ".xlsx":
        return "xlsx"
    elif ext == ".xls":
        return "xls"
    elif ext == ".csv":
        return "csv"
    return "unknown"


# ─── LLM 调用 ─────────────────────────────────────────────────────────────────

def _call_llm_text(prompt, api_key, base_url, model, timeout=120):
    """标准 OpenAI 兼容 chat completion 请求（纯文本，无文件附件）。

    Args:
        timeout: 超时秒数，或 (connect_timeout, read_timeout) 元组。
    """
    url = base_url.rstrip("/") + "/chat/completions"
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }
    payload = {
        "model": model,
        "messages": [{"role": "user", "content": prompt}],
        "temperature": 0,
    }

    try:
        resp = requests.post(url, json=payload, headers=headers, timeout=timeout)
    except requests.exceptions.Timeout:
        raise RuntimeError(f"请求超时 ({timeout}s): {url}")
    except requests.exceptions.ConnectionError as e:
        raise RuntimeError(f"连接失败: {url}\n{e}")

    if resp.status_code != 200:
        raise RuntimeError(
            f"HTTP {resp.status_code} | URL: {url}\nResponse: {resp.text[:1000]}"
        )

    try:
        data = resp.json()
    except Exception:
        raise RuntimeError(f"JSON 解析失败 | Response: {resp.text[:1000]}")

    try:
        return data["choices"][0]["message"]["content"]
    except (KeyError, IndexError):
        raise RuntimeError(f"响应结构异常 | Response: {json.dumps(data, ensure_ascii=False)[:1000]}")


# ─── 主入口 ───────────────────────────────────────────────────────────────────

def _categorize_rejection(reason):
    """将拒绝原因归类为统计类别。"""
    if not reason:
        return "unknown"
    r = reason.lower()
    if "json" in r or "解析" in r or "格式" in r:
        return "json_format"
    if "白名单" in r or "不在" in r:
        return "whitelist"
    if "聚合" in r:
        return "aggregate"
    if "孤立数值" in r:
        return "isolated_numeric"
    if "跨行" in r:
        return "cross_row"
    if "空" in r or "为空" in r:
        return "empty"
    if "公式" in r:
        return "formula"
    if "重复" in r:
        return "duplicate"
    return "other"


def _validate_single_question(q, sheets_by_name, file_name,
                               semantic_field_names, semantic_anchors,
                               candidate_anchors=None):
    """验证单条题目，返回 (validated_dict_or_None, rejection_category, reason)。"""
    # 候选集合校验：anchor 必须匹配某个候选或为其子范围
    if candidate_anchors is not None:
        q_anchor = (q.get("anchor_range") or "").strip()
        q_sheet = (q.get("sheet_name") or "").strip()
        if q_anchor and q_sheet:
            ab = _parse_range_str(q_anchor)
            matched = False
            for cand in candidate_anchors:
                if cand["sheet_name"] != q_sheet:
                    continue
                if cand["anchor_range"] == q_anchor:
                    matched = True
                    break
                cb = _parse_range_str(cand["anchor_range"])
                if ab and cb:
                    a_mc, a_mr, a_xc, a_xr = ab
                    c_mc, c_mr, c_xc, c_xr = cb
                    if (a_mc >= c_mc and a_xc <= c_xc and
                            a_mr >= c_mr and a_xr <= c_xr):
                        matched = True
                        break
            if not matched:
                return None, "candidate_mismatch", (
                    f"anchor '{q_anchor}' 不在合法候选集合中"
                )

    validated, reason = _validate_and_render_question(q, sheets_by_name, file_name)
    if not validated:
        return None, _categorize_rejection(reason), reason

    consistent, consistency_reason = _validate_question_anchor_consistency(
        validated, semantic_field_names, semantic_anchors, sheets_by_name,
    )
    if not consistent:
        return None, _categorize_rejection(consistency_reason), consistency_reason

    validated["question_mode"] = MODE_RETRIEVAL
    return validated, None, None


# ─── Phase 1: Schema 分析 ────────────────────────────────────────────────────

def _build_schema_analysis_context(sheets):
    """构建紧凑的结构摘要供 schema 分析 LLM 调用。

    不发送原始二进制文件，只发送：表名、尺寸、合并单元格、表头候选行、
    抽样业务行、列名（带列号）、数值模式、总计/小计候选、公式警告。
    """
    parts = []
    for sheet in sheets:
        section = []
        section.append(f"=== Sheet: {sheet.sheet_name} ({sheet.max_row} rows x {sheet.max_col} cols) ===")

        # 合并单元格
        if sheet.merged_cells:
            merge_descs = []
            for mr, mc, mxr, mxc in sheet.merged_cells:
                tl = sheet.rows[mr - 1][mc - 1] if mr - 1 < len(sheet.rows) and mc - 1 < len(sheet.rows[mr - 1]) else None
                val = str(tl).strip()[:30] if tl else ""
                merge_descs.append(f"{_col_letter(mc)}{mr}:{_col_letter(mxc)}{mxr}" + (f" ({val})" if val else ""))
            section.append(f"Merged cells: {', '.join(merge_descs)}")

        # 检测表头行和数据起始行
        header_idx, data_start, rate_idx = _detect_header_row_and_data_start(sheet)
        section.append(f"Detected header_row={header_idx + 1}, data_start_row={data_start}")

        # 表头行（带列号标注）
        header_row = sheet.rows[header_idx] if header_idx < len(sheet.rows) else []
        col_labels = []
        for c in range(sheet.max_col):
            v = header_row[c] if c < len(header_row) else None
            label = str(v).strip()[:30] if v else ""
            col_labels.append(f"{_col_letter(c + 1)}={label}" if label else f"{_col_letter(c + 1)}=(empty)")
        section.append(f"Columns ({sheet.max_col}): {', '.join(col_labels)}")

        # 第 2 行（可能是费率行/副表头）
        if len(sheet.rows) > 1:
            row2 = sheet.rows[1]
            cells2 = []
            for c in range(sheet.max_col):
                v = row2[c] if c < len(row2) else None
                s = str(v).strip()[:25] if v is not None else ""
                cells2.append(s)
            section.append(f"Row 2 (possible rate/header): {' | '.join(cells2)}")

        # 抽样业务行（前 5 行数据，带行号）
        sample_count = 0
        for r in range(data_start - 1, min(len(sheet.rows), data_start + 30)):
            if sample_count >= 5:
                break
            row_data = sheet.rows[r]
            if _is_summary_row(row_data):
                section.append(f"Row {r + 1}: [SUMMARY ROW - exclude]")
                continue
            non_empty = sum(1 for v in row_data if v is not None and str(v).strip())
            if non_empty == 0:
                continue
            cells = []
            for c in range(sheet.max_col):
                v = row_data[c] if c < len(row_data) else None
                s = str(v).strip()[:25] if v is not None else ""
                cells.append(s)
            section.append(f"Row {r + 1}: {' | '.join(cells)}")
            sample_count += 1

        # 检测汇总行
        summary_rows = []
        for r in range(len(sheet.rows)):
            if _is_summary_row(sheet.rows[r]):
                summary_rows.append(r + 1)
        if summary_rows:
            section.append(f"Summary rows (MUST exclude): {summary_rows}")

        # 公式警告
        if sheet.formula_cells_without_cache:
            section.append(f"Formula warnings: {len(sheet.formula_cells_without_cache)} cells without cached values")

        # 数值模式统计（前 data_start+30 行内）
        numeric_cols = set()
        for r in range(data_start - 1, min(len(sheet.rows), data_start + 30)):
            for c in range(sheet.max_col):
                v = sheet.rows[r][c] if c < len(sheet.rows[r]) else None
                if v is not None and _is_numeric_value(v):
                    numeric_cols.add(c)
        if numeric_cols:
            num_labels = [_col_letter(c + 1) for c in sorted(numeric_cols)]
            section.append(f"Columns with numeric values: {', '.join(num_labels)}")

        parts.append("\n".join(section))

    return "\n\n".join(parts)


def _parse_schema_analysis_response(text):
    """解析 LLM 返回的 schema 分析 JSON。

    处理 markdown 代码块、部分 JSON 等。返回 dict。
    """
    text = text.strip()
    if text.startswith("```"):
        lines = text.split("\n")
        json_lines = []
        in_block = False
        for line in lines:
            if line.strip().startswith("```") and not in_block:
                in_block = True
                continue
            elif line.strip() == "```" and in_block:
                break
            elif in_block:
                json_lines.append(line)
        text = "\n".join(json_lines).strip()

    try:
        data = json.loads(text)
    except json.JSONDecodeError:
        # 尝试提取第一个 { 到最后一个 }
        start = text.find("{")
        end = text.rfind("}")
        if start >= 0 and end > start:
            try:
                data = json.loads(text[start:end + 1])
            except json.JSONDecodeError:
                raise ValueError("无法解析 schema 分析 LLM 返回的 JSON")
        else:
            raise ValueError("schema 分析 LLM 返回中未找到 JSON")

    # 验证和默认值
    if "fields" not in data or not isinstance(data["fields"], list):
        raise ValueError("schema 分析结果缺少 fields 数组")

    for f in data["fields"]:
        if "source_label" not in f or "col_index" not in f:
            raise ValueError(f"field 缺少 source_label 或 col_index: {f}")
        f.setdefault("inferred_role", "categorical")
        f.setdefault("confidence", 0.5)
        # 数值字段和分类字段默认 needs_confirmation=true
        if f["inferred_role"] in ("metric", "cost", "categorical"):
            f.setdefault("needs_confirmation", True)
        else:
            f.setdefault("needs_confirmation", False)

    data.setdefault("table_purpose", "")
    data.setdefault("header_row", 1)
    data.setdefault("data_start_row", 2)
    data.setdefault("excluded_rows", [])
    data.setdefault("reasoning", "")

    return data


def analyze_table_schema(sheets, api_key, base_url, model, timeout=60,
                         file_bytes=None, force_reanalyze=False,
                         allow_test_model=False):
    """Phase 1: 通过 LLM 分析表格结构。

    Args:
        sheets: SheetContext 列表
        api_key, base_url, model: LLM 配置
        timeout: LLM 超时秒数
        file_bytes: 文件原始字节（用于缓存）
        force_reanalyze: 强制重新分析（忽略缓存）
        allow_test_model: 允许测试模型名（仅测试使用，不写入生产缓存）

    Returns:
        dict: schema 分析结果，含 table_purpose, fields, excluded_rows, safe_question_fields 等
    """
    # 缓存检查
    if file_bytes:
        file_hash = _file_content_hash(file_bytes)
        if force_reanalyze:
            _delete_schema_cache(file_hash)
        else:
            cached = _load_schema_cache(file_hash, expected_model=model,
                                         allow_test_model=allow_test_model)
            if cached is not None:
                print(f"  [CACHE] Schema 缓存命中: {file_hash[:12]}... (model={cached.get('analysis_model')})")
                return cached

    # 拒绝 fake 模型进入生产分析（除非测试显式允许）
    if _is_fake_model(model) and not allow_test_model:
        raise ValueError(f"拒绝使用测试模型 '{model}' 进行 schema 分析")

    # 构建结构摘要
    context_text = _build_schema_analysis_context(sheets)

    # 加载 prompt 模板
    template = _SCHEMA_ANALYSIS_PROMPT_PATH.read_text(encoding="utf-8").strip()
    prompt = template.replace("{structure_summary}", context_text)

    # Phase 1 专用超时：连接 15s，读取 180s（结构分析 prompt 较长）
    _phase1_timeout = (15, 180)
    _max_attempts = 2

    import time as _time

    print(f"  [Schema] 分析: prompt={len(prompt)} chars, timeout=({_phase1_timeout[0]}s connect, {_phase1_timeout[1]}s read)")

    # LLM 调用（带重试）
    response_text = None
    _total_t0 = _time.time()
    _attempt_durations = []
    for _attempt in range(1, _max_attempts + 1):
        _t0 = _time.time()
        try:
            response_text = _call_llm_text(prompt, api_key, base_url, model, timeout=_phase1_timeout)
            _t1 = _time.time()
            _attempt_durations.append(round(_t1 - _t0, 1))
            print(f"  [Schema] 第 {_attempt} 次请求成功: 耗时={_t1 - _t0:.1f}s")
            break
        except (RuntimeError, TimeoutError) as e:
            _t1 = _time.time()
            _attempt_durations.append(round(_t1 - _t0, 1))
            _err_str = str(e)
            _is_retryable = "超时" in _err_str or "连接" in _err_str or "timeout" in _err_str.lower()
            if _attempt < _max_attempts and _is_retryable:
                print(f"  [Schema] 第 {_attempt} 次请求失败: {_err_str}，正在重试（{_attempt + 1}/{_max_attempts}）")
                continue
            else:
                raise

    _total_t1 = _time.time()
    _total_duration = round(_total_t1 - _total_t0, 1)
    print(f"  [Schema] 分析完成: 尝试={len(_attempt_durations)} 次, 总耗时={_total_duration}s")

    # 解析响应
    schema_data = _parse_schema_analysis_response(response_text)

    # ── 本地严格校验 LLM 返回 ──
    sheet = sheets[0] if sheets else None
    if sheet is None:
        raise ValueError("无工作表数据")

    # 校验字段：col_index 和 source_label 必须在真实表头中存在
    valid_fields = []
    actual_headers = {}
    if sheet.rows:
        for c in range(sheet.max_col):
            v = sheet.rows[0][c] if c < len(sheet.rows[0]) else None
            if v is not None and str(v).strip():
                actual_headers[c + 1] = str(v).strip()  # 1-indexed

    for f in schema_data["fields"]:
        col_idx = f.get("col_index", 0)
        label = f.get("source_label", "")
        # 校验列号在范围内
        if col_idx < 1 or col_idx > sheet.max_col:
            print(f"  [WARN] 丢弃非法字段: col_index={col_idx} 超出范围 (1-{sheet.max_col})")
            continue
        # 校验列号对应的表头与 source_label 匹配（允许部分匹配）
        actual = actual_headers.get(col_idx, "")
        if actual and label and actual != label:
            # 降级为 categorical
            if f.get("inferred_role") in ("context", "record_identifier"):
                print(f"  [WARN] 字段 '{label}' 在列 {col_idx} 实际为 '{actual}'，降为 categorical")
                f["inferred_role"] = "ambiguous"
                f["confidence"] = min(f.get("confidence", 0.5), 0.5)
        valid_fields.append(f)

    if not valid_fields:
        raise ValueError("Phase 1 未返回任何有效字段")

    schema_data["fields"] = valid_fields

    # 校验 excluded_rows
    valid_excluded = []
    for r in schema_data.get("excluded_rows", []):
        if isinstance(r, int) and 1 <= r <= sheet.max_row:
            valid_excluded.append(r)
    schema_data["excluded_rows"] = valid_excluded

    # 按角色分组字段（新角色体系）
    role_groups = {"record_identifier": [], "context": [], "metric": [], "cost": [], "categorical": [], "excluded": []}
    for f in schema_data["fields"]:
        role = f.get("inferred_role", "categorical")
        # 兼容旧角色名映射
        if role == "group":
            role = "context"
        elif role == "record":
            role = "record_identifier"
        elif role == "ambiguous":
            role = "categorical"
        if role not in role_groups:
            role = "categorical"
        role_groups[role].append(f)

    schema_data["record_identifier_fields"] = role_groups["record_identifier"]
    schema_data["context_fields"] = role_groups["context"]
    schema_data["metric_fields"] = role_groups["metric"]
    schema_data["cost_fields"] = role_groups["cost"]
    schema_data["categorical_fields"] = role_groups["categorical"]
    schema_data["excluded_fields"] = role_groups["excluded"]

    # 兼容旧接口
    schema_data["group_fields"] = role_groups["context"]
    schema_data["record_fields"] = role_groups["record_identifier"]
    schema_data["ambiguous_fields"] = role_groups["categorical"]

    # 计算 record_locator_fields 和 question_target_fields
    record_locators = [f["source_label"] for f in role_groups["record_identifier"]]
    question_targets = [f["source_label"] for f in role_groups["metric"] + role_groups["cost"] + role_groups["categorical"]]
    # 如果没有 record_identifier 字段，用 context 字段兜底
    if not record_locators:
        record_locators = [f["source_label"] for f in role_groups["context"]]
    # 如果没有目标字段，把 record 中非主标识的字段也作为目标
    if not question_targets and len(record_locators) > 1:
        question_targets = record_locators[1:]

    schema_data["record_locator_fields"] = record_locators
    schema_data["question_target_fields"] = question_targets

    # 兼容旧接口
    schema_data["safe_question_fields"] = record_locators + question_targets

    # 提取或生成 question_plan（新结构）
    qp = schema_data.get("question_plan", {})
    if not qp.get("target_field_priority"):
        # LLM 未返回新结构的 question_plan，自动生成
        target_field_priority = []
        target_field_quotas = {}
        patterns = set()
        priority = 1

        for f in role_groups["metric"]:
            target_field_priority.append({
                "field": f["source_label"],
                "role": "metric",
                "priority": priority,
                "reason": "数值度量字段",
            })
            patterns.add("record + metric")
            priority += 1

        for f in role_groups["cost"]:
            target_field_priority.append({
                "field": f["source_label"],
                "role": "cost",
                "priority": priority,
                "reason": "费用/金额字段",
            })
            patterns.add("record + cost")
            priority += 1

        for f in role_groups["categorical"]:
            target_field_priority.append({
                "field": f["source_label"],
                "role": "categorical",
                "priority": priority,
                "reason": "分类/状态字段",
            })
            patterns.add("record + categorical")
            priority += 1

        # 均分目标题数到各字段
        if target_field_priority:
            per_field = max(1, 5 // len(target_field_priority))
            for tfp in target_field_priority:
                target_field_quotas[tfp["field"]] = per_field

        qp = {
            "recommended_question_patterns": sorted(patterns),
            "target_field_priority": target_field_priority,
            "target_field_quotas": target_field_quotas,
            "forbidden_patterns": [
                "aggregation (sum, count, average, total across rows)",
                "cross-row comparison (higher than, lower than, difference between)",
                "totals (using summary/total rows as data)",
                "null values (questions about empty cells)",
                "formula without cached value",
                "isolated numeric (numeric without field label context)",
            ],
            "rationale": (
                f"表格含 {len(record_locators)} 个记录标识字段、"
                f"{len(role_groups['metric'])} 个数值字段、"
                f"{len(role_groups['cost'])} 个费用字段、"
                f"{len(role_groups['categorical'])} 个分类字段，"
                f"适合生成单记录检索查询"
            ),
        }
    schema_data["question_plan"] = qp

    # 元数据
    schema_data["sheet_name"] = sheets[0].sheet_name if sheets else ""
    schema_data["analysis_model"] = model
    schema_data["analysis_timestamp"] = datetime.now().isoformat()
    schema_data["schema_source"] = "test_mock" if _is_fake_model(model) else "llm"
    schema_data["llm_call_duration_sec"] = _total_duration
    schema_data["llm_call_attempts"] = len(_attempt_durations)
    schema_data["llm_attempt_durations"] = _attempt_durations

    print(f"  [OK] Phase 1 校验通过: {len(valid_fields)} 字段, "
          f"记录定位={len(record_locators)}, 目标字段={len(question_targets)}, "
          f"推荐题型={len(qp.get('recommended_question_patterns', []))} 种")

    # 缓存保存
    if file_bytes:
        _save_schema_cache(_file_content_hash(file_bytes), schema_data)

    return schema_data


def _slugify_field(name):
    """将字段名转为安全的 slug（用于 candidate_id）。"""
    return re.sub(r'[^a-zA-Z0-9\u4e00-\u9fff]', '_', str(name)).strip('_').lower()


def _build_candidate_anchors_from_schema(sheets, confirmed_schema):
    """基于 schema 的 question_plan 构建候选目录（candidate catalog）。

    每个候选 = 一条业务记录 + 一个目标字段。
    输出新格式：含 candidate_id, record_locator, target_field, allowed_evidence_fields。
    """
    candidates = []
    excluded_rows = set(confirmed_schema.get("excluded_rows", []))
    qp = confirmed_schema.get("question_plan", {})
    target_priority = qp.get("target_field_priority", [])
    record_fields = confirmed_schema.get("record_locator_fields", [])
    context_fields = confirmed_schema.get("context_fields", [])

    if not record_fields or not target_priority:
        return candidates

    # 确保 context_fields 是字符串列表（兼容字段对象列表）
    if context_fields and isinstance(context_fields[0], dict):
        context_fields = [f.get("source_label", "") for f in context_fields]
    context_fields = [f for f in context_fields if f]

    # 构建 target_field role 映射
    target_role_map = {}
    for tfp in target_priority:
        target_role_map[tfp.get("field", "")] = tfp.get("role", "metric")

    for sheet in sheets:
        header_row = sheet.rows[0] if sheet.rows else []

        # 构建字段名→列索引映射
        field_to_col = {}
        for c, h in enumerate(header_row):
            if h is not None and str(h).strip():
                field_to_col[str(h).strip()] = c  # 0-indexed

        # record 字段列索引
        record_col_indices = []
        for fname in record_fields:
            if fname in field_to_col:
                record_col_indices.append(field_to_col[fname])
        if not record_col_indices:
            continue

        # context 字段列索引
        context_col_indices = {}
        for fname in context_fields:
            if fname in field_to_col:
                context_col_indices[fname] = field_to_col[fname]

        # 检测表头行和数据起始行
        _, data_start, _ = _detect_header_row_and_data_start(sheet)

        # 扫描业务行
        for row_idx in range(data_start - 1, len(sheet.rows)):
            excel_row = row_idx + 1
            if excel_row in excluded_rows:
                continue
            row_data = sheet.rows[row_idx]
            if not any(v is not None and str(v).strip() for v in row_data):
                continue
            if _is_summary_row(row_data):
                continue

            # 构建 record_locator（记录标识字段的实际值）
            record_locator = {}
            for fname in record_fields:
                if fname in field_to_col:
                    c = field_to_col[fname]
                    v = row_data[c] if c < len(row_data) else None
                    if v is not None and str(v).strip():
                        record_locator[fname] = str(v).strip()

            if not record_locator:
                continue

            # 构建 available_context_fields
            available_context = {}
            for fname, c in context_col_indices.items():
                v = row_data[c] if c < len(row_data) else None
                if v is not None and str(v).strip():
                    available_context[fname] = str(v).strip()

            # 为每个 target_field 生成一个候选
            for tfp in target_priority:
                target_field = tfp.get("field", "")
                target_role = tfp.get("role", "metric")
                if target_field not in field_to_col:
                    continue
                target_col = field_to_col[target_field]  # 0-indexed

                # 读取目标字段的实际值
                target_val = row_data[target_col] if target_col < len(row_data) else None
                if target_val is None or not str(target_val).strip():
                    continue
                if str(target_val).strip() == "[公式未计算]":
                    continue

                # anchor 覆盖范围 = record列 到 target列（连续）
                all_cols = sorted(set(record_col_indices + [target_col]))
                anchor_start = min(all_cols) + 1  # 1-indexed
                anchor_end = max(all_cols) + 1

                # 跳过合并单元格
                if _range_overlaps_multicell_merge(excel_row, anchor_start, excel_row, anchor_end, sheet.merged_cells):
                    continue

                # 跳过孤立数值（anchor 只覆盖数值列）
                anchor_has_text = False
                for c in range(anchor_start - 1, anchor_end):
                    v = row_data[c] if c < len(row_data) else None
                    if v is not None and str(v).strip() and not _is_numeric_value(v):
                        anchor_has_text = True
                        break
                if not anchor_has_text:
                    continue

                anchor_range = f"{_col_letter(anchor_start)}{excel_row}:{_col_letter(anchor_end)}{excel_row}"

                # 构建 header_context_range
                h_start = min(all_cols) + 1
                h_end = max(all_cols) + 1
                header_context_range = f"{_col_letter(h_start)}1:{_col_letter(h_end)}1"

                # 构建 allowed_evidence_fields
                allowed_fields = list(record_locator.keys()) + [target_field]

                # 生成 candidate_id
                sheet_slug = _slugify_field(sheet.sheet_name)
                field_slug = _slugify_field(target_field)
                candidate_id = f"{sheet_slug}_row_{excel_row}_{field_slug}"

                candidates.append({
                    "candidate_id": candidate_id,
                    "evidence_mode": "record_with_schema_context",
                    "sheet_name": sheet.sheet_name,
                    "anchor_range": anchor_range,
                    "header_context_range": header_context_range,
                    "record_locator": record_locator,
                    "target_field": {
                        "label": target_field,
                        "role": target_role,
                        "value": target_val,
                    },
                    "available_context_fields": available_context,
                    "allowed_evidence_fields": allowed_fields,
                })

    return candidates


def _build_phase2_prompt(sheets, confirmed_schema, num_questions, topic_hint, candidate_anchors):
    """构建 Phase 2 prompt：注入 schema、question_plan、候选目录（candidate catalog）。"""
    template = PROMPT_TEMPLATE_PATH.read_text(encoding="utf-8").strip()

    topic_hint_section = ""
    if topic_hint:
        topic_hint_section = f"- 主题方向：{topic_hint}"

    # 提取 schema 信息
    qp = confirmed_schema.get("question_plan", {})
    record_fields = confirmed_schema.get("record_locator_fields", [])
    target_fields = confirmed_schema.get("question_target_fields", [])
    excluded_rows = confirmed_schema.get("excluded_rows", [])
    context_fields = confirmed_schema.get("context_fields", [])

    # 确保 context_fields 是字符串列表（兼容字段对象列表）
    if context_fields and isinstance(context_fields[0], dict):
        context_fields = [f.get("source_label", "") for f in context_fields]
    context_fields = [f for f in context_fields if f]

    # 构建 schema + question_plan 摘要
    schema_section = "## Schema 与出题计划\n\n"
    schema_section += f"- **记录定位字段**: {', '.join(record_fields)}\n"
    schema_section += f"- **上下文字段**: {', '.join(context_fields)}\n"
    schema_section += f"- **可出题目标字段**: {', '.join(target_fields)}\n"
    if excluded_rows:
        schema_section += f"- **已排除行**: {', '.join(str(r) for r in excluded_rows)}\n"

    # 注入 question_plan（新结构）
    if qp:
        patterns = qp.get("recommended_question_patterns", [])
        if patterns:
            schema_section += f"\n**推荐题型**: {', '.join(patterns)}\n"

        priority_list = qp.get("target_field_priority", [])
        if priority_list:
            schema_section += "\n### 目标字段优先级\n\n"
            schema_section += "| 字段 | 角色 | 优先级 | 理由 |\n|---|---|---|---|\n"
            for tfp in priority_list:
                schema_section += (
                    f"| {tfp.get('field', '')} | {tfp.get('role', '')} | "
                    f"{tfp.get('priority', '')} | {tfp.get('reason', '')} |\n"
                )

        quotas = qp.get("target_field_quotas", {})
        if quotas:
            schema_section += "\n**建议题数分配**:\n"
            for field_name, count in quotas.items():
                schema_section += f"- {field_name}: {count} 题\n"

        forbidden = qp.get("forbidden_patterns", [])
        if forbidden:
            schema_section += "\n**禁止题型**:\n"
            for fb in forbidden:
                schema_section += f"- {fb}\n"

        rationale = qp.get("rationale", "")
        if rationale:
            schema_section += f"\n**分析理由**: {rationale}\n"

    schema_section += (
        "\n**核心约束**: 每题 = 一条业务记录的 record_locator + 一个 target_field 值。"
        "你只能从候选目录中选择 candidate_id，不得自造 anchor_range、字段或数值。"
        "reference_answer 由本地渲染，你不输出。\n"
    )

    # 收集所有涉及列（用于渲染表格）
    all_fields = set(record_fields + context_fields + target_fields)

    # 渲染表格块（只包含涉及字段列）
    block_texts = []
    for sheet in sheets:
        header_row = sheet.rows[0] if sheet.rows else []
        col_indices = []
        for c, h in enumerate(header_row):
            if h is not None and str(h).strip() in all_fields:
                col_indices.append(c)
        if not col_indices:
            continue

        lines = []
        header_cells = [str(header_row[c]).strip() if c < len(header_row) and header_row[c] else "" for c in col_indices]
        lines.append("| 行号 | " + " | ".join(header_cells) + " |")
        lines.append("|---:|" + "|".join(["---"] * len(header_cells)) + "|")

        for r in range(len(sheet.rows)):
            excel_row = r + 1
            if excel_row in excluded_rows:
                continue
            if _is_summary_row(sheet.rows[r]):
                continue
            row_data = sheet.rows[r]
            cells = []
            for c in col_indices:
                v = row_data[c] if c < len(row_data) else None
                cells.append(str(v).strip() if v is not None else "")
            if all(not c for c in cells):
                continue
            lines.append(f"| {excel_row} | " + " | ".join(cells) + " |")

        block_text = f"### 工作表: {sheet.sheet_name}\n\n" + "\n".join(lines)
        block_texts.append(block_text)

    table_blocks_text = "\n\n---\n\n".join(block_texts)

    # 构建候选目录（candidate catalog）
    candidate_text = ""
    if candidate_anchors:
        # 按 target_field.role 分组
        role_groups = {}
        for a in candidate_anchors:
            role = a.get("target_field", {}).get("role", "metric")
            if role not in role_groups:
                role_groups[role] = []
            role_groups[role].append(a)

        _ROLE_LABELS = {
            "metric": "数值度量题候选",
            "cost": "费用/金额题候选",
            "categorical": "分类/状态题候选",
        }

        parts = []
        for role in ("metric", "cost", "categorical"):
            items = role_groups.get(role, [])
            if not items:
                continue
            label = _ROLE_LABELS.get(role, role)
            lines = [f"**{label}**（{len(items)} 个候选）："]
            for a in items[:20]:
                rl = a.get("record_locator", {})
                rl_str = ", ".join(f"{k}={v}" for k, v in list(rl.items())[:2])
                tf = a.get("target_field", {})
                item_parts = [
                    f"**candidate_id**: `{a['candidate_id']}`",
                    f"**record_locator**: {{{rl_str}}}",
                    f"**target_field**: {tf.get('label', '')}（{tf.get('role', '')}）= {tf.get('value', '')}",
                ]
                line = "- " + " | ".join(item_parts)
                lines.append(line)
            parts.append("\n".join(lines))

        if parts:
            candidate_text = "\n\n---\n\n## 合法候选目录（candidate catalog）\n\n" + "\n\n".join(parts)

    prompt = template.replace("{table_blocks_text}", table_blocks_text)
    prompt = prompt.replace("{num_questions}", str(num_questions))
    prompt = prompt.replace("{topic_hint_section}", topic_hint_section)
    prompt = prompt.replace("{candidate_anchors_section}", candidate_text)
    prompt = schema_section + "\n---\n\n" + prompt
    return prompt


# ─── Phase 2 新验证与渲染 ─────────────────────────────────────────────────────

def _render_phase2_reference_answer(candidate, sheet_ctx):
    """从候选目录项本地渲染 reference_answer。

    只渲染 allowed_evidence_fields 中的字段为键值对。
    格式：需求描述：工具链可用性优化；QTY. in total：17

    MUST NOT 使用 LLM 输出的任何值。只读取 sheet_ctx.rows。
    """
    anchor_range = candidate.get("anchor_range", "")
    header_context_range = candidate.get("header_context_range", "")
    allowed_fields = set(candidate.get("allowed_evidence_fields", []))

    if not anchor_range:
        return "", True

    # 解析 anchor range
    bounds = _parse_range_str(anchor_range)
    if bounds is None:
        return "", True
    b_min_col, b_min_row, b_max_col, b_max_row = bounds

    # 获取表头行（从 header_context_range 或 row 1）
    if header_context_range:
        h_bounds = _parse_range_str(header_context_range)
        if h_bounds:
            _, h_row, _, _ = h_bounds
            header_row_data = sheet_ctx.rows[h_row - 1] if h_row - 1 < len(sheet_ctx.rows) else []
        else:
            header_row_data = sheet_ctx.rows[0] if sheet_ctx.rows else []
    else:
        header_row_data = sheet_ctx.rows[0] if sheet_ctx.rows else []

    # 获取业务行数据
    data_row = sheet_ctx.rows[b_min_row - 1] if b_min_row - 1 < len(sheet_ctx.rows) else []

    has_formula_issue = False
    parts = []

    for c in range(b_min_col, b_max_col + 1):
        # 读取表头字段名
        h_val = header_row_data[c - 1] if c - 1 < len(header_row_data) else None
        field_name = str(h_val).strip() if h_val else ""

        # 只渲染 allowed_evidence_fields 中的字段
        if field_name not in allowed_fields:
            continue

        # 读取业务行值
        val = data_row[c - 1] if c - 1 < len(data_row) else None
        if (b_min_row, c) in sheet_ctx.formula_cells_without_cache:
            has_formula_issue = True
            continue
        if isinstance(val, str) and val == "[公式未计算]":
            has_formula_issue = True
            continue
        if val is None or not str(val).strip():
            continue

        parts.append(f"{field_name}：{str(val).strip()}")

    if not parts:
        return "", has_formula_issue

    return "；".join(parts), has_formula_issue


def _validate_phase2_question(q, candidate_catalog, sheets_by_name):
    """验证 Phase 2 LLM 输出，按 candidate_id 查找并本地渲染 reference_answer。

    Returns (validated_dict_or_None, rejection_category, reason).
    """
    candidate_id = (q.get("candidate_id") or "").strip()
    target_field_label = (q.get("target_field_label") or "").strip()
    question_text = (q.get("question") or "").strip()

    if not candidate_id:
        return None, "missing_field", "candidate_id 为空"
    if not target_field_label:
        return None, "missing_field", "target_field_label 为空"
    if not question_text:
        return None, "missing_field", "question 为空"

    # 查找 candidate_id
    candidate = None
    for c in candidate_catalog:
        if c["candidate_id"] == candidate_id:
            candidate = c
            break
    if candidate is None:
        return None, "candidate_mismatch", f"candidate_id '{candidate_id}' 不在候选目录中"

    # 校验 target_field_label
    expected_label = candidate.get("target_field", {}).get("label", "")
    if target_field_label != expected_label:
        return None, "field_mismatch", (
            f"target_field_label '{target_field_label}' 与候选目录不匹配，"
            f"期望 '{expected_label}'"
        )

    # 查找工作表
    sheet_name = candidate.get("sheet_name", "")
    sheet = sheets_by_name.get(sheet_name)
    if sheet is None:
        return None, "sheet_missing", f"工作表 '{sheet_name}' 不存在"

    # 本地渲染 reference_answer
    rendered, has_formula_issue = _render_phase2_reference_answer(candidate, sheet)
    if not rendered:
        return None, "empty_evidence", "证据范围为空"
    if has_formula_issue:
        return None, "formula", "证据范围包含公式单元格（无缓存计算值）"

    # 汇总行保护
    anchor_range = candidate.get("anchor_range", "")
    bounds = _parse_range_str(anchor_range)
    if bounds:
        _, b_min_row, _, _ = bounds
        data_row = sheet.rows[b_min_row - 1] if b_min_row - 1 < len(sheet.rows) else []
        if _is_summary_row(data_row):
            return None, "summary_row", "候选指向汇总行"

    # 组装验证后的题目
    validated = {
        "question": question_text,
        "reference_answer": rendered,
        "source_excerpt": rendered,
        "candidate_id": candidate_id,
        "target_field_label": target_field_label,
        "sheet_name": sheet_name,
        "anchor_range": anchor_range,
        "evidence_sheet": sheet_name,
        "evidence_range": anchor_range,
        "source_format": _detect_format(""),
        "difficulty": q.get("difficulty", "事实"),
        "topic": q.get("topic", ""),
        "record_locator": candidate.get("record_locator", {}),
        "question_mode": MODE_RETRIEVAL,
    }

    # 添加 evidence_schema_display
    hcr = candidate.get("header_context_range")
    if hcr:
        validated["header_context_range"] = hcr
    validated["evidence_schema_display"] = _build_evidence_schema_display(
        anchor_range, hcr, sheet,
    )

    return validated, None, None


def _dedup_by_record_and_field(questions):
    """按 (record_locator, target_field_label) 去重。"""
    seen = set()
    unique = []
    for q in questions:
        rl = q.get("record_locator", {})
        key = (json.dumps(rl, sort_keys=True, ensure_ascii=False),
               q.get("target_field_label", ""))
        if key not in seen:
            seen.add(key)
            unique.append(q)
    return unique


def generate_questions_from_schema(sheets, confirmed_schema, api_key, base_url, model,
                                   num_questions=5, difficulty="混合", topic_hint="",
                                   timeout=120, progress_callback=None, mode="retrieval",
                                   file_name=""):
    """Phase 2: 基于确认的 schema 生成检索题。

    使用候选目录（candidate catalog），LLM 通过 candidate_id 选择题目。
    reference_answer 完全本地渲染。

    Args:
        sheets: SheetContext 列表
        confirmed_schema: 用户确认的 schema dict
        api_key, base_url, model: LLM 配置
        num_questions: 目标题数
        difficulty: 难度偏好
        topic_hint: 主题方向
        timeout: LLM 超时秒数
        progress_callback: 进度回调
        mode: "retrieval" 或 "qa"
        file_name: 原始文件名

    Returns:
        tuple: (questions_list, stats_dict)
    """
    if progress_callback:
        progress_callback(0, 5, "解析表格文件")

    sheets_by_name = {s.sheet_name: s for s in sheets}

    if progress_callback:
        progress_callback(1, 5, "构建候选目录（基于确认 schema）")

    # 基于确认 schema 构建候选目录
    candidate_anchors = _build_candidate_anchors_from_schema(sheets, confirmed_schema)

    if not candidate_anchors:
        raise ValueError(
            "基于确认的 schema 未生成任何合法候选。"
            "请检查 safe_question_fields 是否包含可出题的文本字段。"
        )

    # 构建 candidate_id 索引
    candidate_by_id = {c["candidate_id"]: c for c in candidate_anchors}

    actual_target = min(num_questions, len(candidate_anchors))
    if actual_target < num_questions:
        print(f"  [INFO] 仅有 {len(candidate_anchors)} 个合法候选，最多生成 {actual_target} 题")

    # 构建 Phase 2 prompt
    prompt = _build_phase2_prompt(sheets, confirmed_schema, actual_target, topic_hint, candidate_anchors)
    effective_timeout = (15, timeout)

    print(f"  [PROMPT] Phase 2 Prompt: {len(prompt)} chars, 候选={len(candidate_anchors)}")

    if progress_callback:
        progress_callback(2, 5, "调用 LLM 生成检索查询")

    # LLM 调用
    import time as _time
    _t0 = _time.time()
    response_text = _call_llm_text(prompt, api_key, base_url, model, timeout=effective_timeout)
    _t1 = _time.time()
    raw_questions = _parse_llm_response(response_text)
    first_raw_count = len(raw_questions)
    print(f"  [LLM] 首次调用: prompt={len(prompt)} chars, 耗时={_t1 - _t0:.1f}s, 返回 {first_raw_count} 条")

    if progress_callback:
        progress_callback(3, 5, "验证并渲染金标准")

    # 验证（使用新的 _validate_phase2_question）
    rejection_stats = {}
    valid_questions = []
    used_candidate_ids = set()

    for q in raw_questions:
        validated, category, reason = _validate_phase2_question(
            q, candidate_anchors, sheets_by_name,
        )
        if validated:
            valid_questions.append(validated)
            used_candidate_ids.add(validated.get("candidate_id", ""))
        else:
            rejection_stats[category] = rejection_stats.get(category, 0) + 1

    # 先按 (record_locator, target_field_label) 去重，再按文本去重
    deduped = _dedup_by_record_and_field(valid_questions)
    unique_questions = deduplicate_questions(deduped)
    first_valid_count = len(unique_questions)

    # 补充调用
    supplement_count = 0
    supplement_valid = 0
    if len(unique_questions) < num_questions:
        remaining = num_questions - len(unique_questions)
        unused_candidates = [a for a in candidate_anchors if a["candidate_id"] not in used_candidate_ids]

        if unused_candidates and remaining > 0:
            if progress_callback:
                progress_callback(4, 5, f"补充调用（还需 {remaining} 条）")

            used_ids_str = ", ".join(f"`{cid}`" for cid in sorted(used_candidate_ids))
            unused_str = "\n".join(
                f"- `{a['candidate_id']}` → {a.get('record_locator', {})} | {a.get('target_field', {}).get('label', '')}"
                for a in unused_candidates[:remaining * 2]
            )
            supplement_prompt = (
                f"已生成的题目使用了以下 candidate_id，不得重复：\n{used_ids_str}\n\n"
                f"还需生成 {remaining} 条不重复的检索查询。"
                f"以下候选尚未使用，请从中选择：\n{unused_str}\n\n"
                f"请严格输出 JSON 数组，每个元素包含 candidate_id, question, target_field_label, difficulty, topic。"
            )

            try:
                _t0 = _time.time()
                supp_response = _call_llm_text(supplement_prompt, api_key, base_url, model, timeout=effective_timeout)
                _t1 = _time.time()
                supp_raw = _parse_llm_response(supp_response)
                supplement_count = len(supp_raw)
                print(f"  [LLM] 补充调用: 耗时={_t1 - _t0:.1f}s, 返回 {supplement_count} 条")

                for q in supp_raw:
                    if len(unique_questions) >= num_questions:
                        break
                    validated, category, reason = _validate_phase2_question(
                        q, candidate_anchors, sheets_by_name,
                    )
                    if validated and validated.get("candidate_id", "") not in used_candidate_ids:
                        unique_questions.append(validated)
                        used_candidate_ids.add(validated["candidate_id"])
                        supplement_valid += 1
                    elif not validated:
                        rejection_stats[category] = rejection_stats.get(category, 0) + 1
            except Exception as e:
                print(f"  [WARN] 补充调用失败: {e}")

    if len(unique_questions) > num_questions:
        unique_questions = unique_questions[:num_questions]

    if not unique_questions:
        raise ValueError("出题失败：所有查询均未通过校验")

    stats = {
        "target": num_questions,
        "first_raw_count": first_raw_count,
        "first_valid_count": first_valid_count,
        "rejection_stats": rejection_stats,
        "supplement_count": supplement_count,
        "supplement_valid": supplement_valid,
        "final_count": len(unique_questions),
        "sheet_count": len(sheets),
        "block_count": sum(len(s.table_blocks) for s in sheets),
        "formula_warnings": sum(len(s.formula_cells_without_cache) for s in sheets),
        "format_warnings": [],
        "candidate_count": len(candidate_anchors),
    }
    return unique_questions, stats


def generate_spreadsheet_questions(file_bytes, file_name, api_key, base_url, model,
                                   num_questions=5, difficulty="混合", topic_hint="",
                                   timeout=120, progress_callback=None, mode="retrieval"):
    """统一电子表格检索题生成。

    Args:
        file_bytes: 文件原始字节
        file_name: 文件名（用于推断格式）
        api_key, base_url, model: LLM 配置
        num_questions: 目标题数
        difficulty: 难度偏好
        topic_hint: 主题方向
        timeout: LLM 超时秒数
        progress_callback: 进度回调 (step, total, description)
        mode: 题目模式 ("retrieval" 或 "qa")

    Returns:
        tuple: (questions_list, stats_dict)
    """
    if progress_callback:
        progress_callback(0, 5, "解析表格文件")

    # 1. 检测格式并解析
    ext = Path(file_name).suffix.lower()
    if ext == ".xlsx":
        sheets = parse_xlsx_to_sheet_contexts(file_bytes)
    elif ext == ".xls":
        sheets = parse_xls_to_sheet_contexts(file_bytes)
    elif ext == ".csv":
        sheets = parse_csv_to_sheet_contexts(file_bytes, file_name)
    else:
        raise ValueError(f"不支持的表格格式: {ext}")

    if not sheets:
        raise ValueError(f"文件 {file_name} 中未提取到任何工作表数据")

    sheets_by_name = {s.sheet_name: s for s in sheets}
    total_blocks = sum(len(s.table_blocks) for s in sheets)
    total_formula_warnings = sum(len(s.formula_cells_without_cache) for s in sheets)
    format_warnings = []
    for s in sheets:
        format_warnings.extend(s.format_warnings)

    semantic_field_names = _extract_semantic_field_names(sheets)
    semantic_anchors = _extract_semantic_anchors(sheets)

    if progress_callback:
        progress_callback(1, 5, "构建候选锚点和 LLM Prompt")

    # 2. 构建候选锚点清单
    candidate_anchors = _build_candidate_anchors(sheets)

    if not candidate_anchors:
        raise ValueError(
            f"文件 {file_name} 中未生成任何合法候选锚点。"
            "可能原因：所有数据行均为汇总/合计行，或表格结构不支持生成检索题。"
        )

    # 3. 候选不足时调整目标
    actual_target = min(num_questions, len(candidate_anchors))
    if actual_target < num_questions:
        print(f"  [INFO] 仅有 {len(candidate_anchors)} 个合法候选，最多生成 {actual_target} 题（目标 {num_questions}）")

    # 4. 构建 prompt（含候选锚点）
    prompt = _build_prompt(sheets, actual_target, topic_hint, candidate_anchors)
    print(f"  [PROMPT] Prompt 字符数: {len(prompt)}，候选总数: {len(candidate_anchors)}")

    # Keep connection failures responsive, but allow the caller-configured
    # read timeout for slower model providers and queued requests.
    effective_timeout = (15, timeout)

    if progress_callback:
        progress_callback(2, 5, "调用 LLM 生成检索查询")

    # 5. 首次 LLM 调用
    import time as _time
    _t0 = _time.time()
    response_text = _call_llm_text(prompt, api_key, base_url, model, timeout=effective_timeout)
    _t1 = _time.time()
    raw_questions = _parse_llm_response(response_text)
    first_raw_count = len(raw_questions)
    print(f"  [LLM] 首次调用: prompt={len(prompt)} chars, 候选={len(candidate_anchors)}, 耗时={_t1-_t0:.1f}s, 返回 {first_raw_count} 条")

    if progress_callback:
        progress_callback(3, 5, "验证并渲染金标准")

    # 5. 首次验证
    rejection_stats = {}
    valid_questions = []
    used_anchors = set()

    for q in raw_questions:
        validated, category, reason = _validate_single_question(
            q, sheets_by_name, file_name, semantic_field_names, semantic_anchors,
            candidate_anchors=candidate_anchors,
        )
        if validated:
            valid_questions.append(validated)
            used_anchors.add(validated.get("anchor_range", ""))
        else:
            rejection_stats[category] = rejection_stats.get(category, 0) + 1

    # 6. 首次去重
    unique_questions = deduplicate_questions(valid_questions)
    first_valid_count = len(unique_questions)

    # 7. 补充调用（如果有效题数 < 目标数且仍有未使用的候选锚点）
    supplement_count = 0
    supplement_valid = 0
    if len(unique_questions) < num_questions:
        remaining = num_questions - len(unique_questions)
        # 找出未使用的候选锚点
        unused_candidates = []
        for a in candidate_anchors:
            if a["anchor_range"] not in used_anchors:
                unused_candidates.append(a)

        if unused_candidates and remaining > 0:
            if progress_callback:
                progress_callback(4, 5, f"补充调用（还需 {remaining} 条）")

            # 构建补充 prompt
            used_anchors_str = ", ".join(f"`{a}`" for a in sorted(used_anchors))
            unused_str = "\n".join(
                f"- `{a['anchor_range']}` → {a['query_focus']}"
                for a in unused_candidates[:remaining * 2]
            )
            supplement_prompt = (
                f"已生成的题目使用了以下锚点，不得重复：\n{used_anchors_str}\n\n"
                f"还需生成 {remaining} 条不重复的检索查询。"
                f"以下候选锚点尚未使用，请从中选择：\n{unused_str}\n\n"
                f"请严格输出 JSON 数组，每个元素包含 question, sheet_name, anchor_range, difficulty, topic。"
            )

            try:
                _t0 = _time.time()
                supp_response = _call_llm_text(supplement_prompt, api_key, base_url, model, timeout=effective_timeout)
                _t1 = _time.time()
                supp_raw = _parse_llm_response(supp_response)
                supplement_count = len(supp_raw)
                print(f"  [LLM] 补充调用: prompt={len(supplement_prompt)} chars, 耗时={_t1-_t0:.1f}s, 返回 {supplement_count} 条")

                for q in supp_raw:
                    if len(unique_questions) >= num_questions:
                        break
                    validated, category, reason = _validate_single_question(
                        q, sheets_by_name, file_name, semantic_field_names, semantic_anchors,
                        candidate_anchors=candidate_anchors,
                    )
                    if validated and validated.get("anchor_range", "") not in used_anchors:
                        unique_questions.append(validated)
                        used_anchors.add(validated["anchor_range"])
                        supplement_valid += 1
                    elif not validated:
                        rejection_stats[category] = rejection_stats.get(category, 0) + 1
            except Exception as e:
                print(f"  [WARN] 补充调用失败: {e}")

    # 8. 最终裁剪
    if len(unique_questions) > num_questions:
        unique_questions = unique_questions[:num_questions]

    if not unique_questions:
        raise ValueError("出题失败：所有查询均未通过锚定校验")

    stats = {
        "target": num_questions,
        "first_raw_count": first_raw_count,
        "first_valid_count": first_valid_count,
        "rejection_stats": rejection_stats,
        "supplement_count": supplement_count,
        "supplement_valid": supplement_valid,
        "final_count": len(unique_questions),
        "sheet_count": len(sheets),
        "block_count": total_blocks,
        "formula_warnings": total_formula_warnings,
        "format_warnings": format_warnings,
        "candidate_counts": {
            mode: sum(1 for a in candidate_anchors if a.get("evidence_mode") == mode)
            for mode in ("record_with_schema_context", "field_value_pair", "text_fact")
        },
    }
    return unique_questions, stats
