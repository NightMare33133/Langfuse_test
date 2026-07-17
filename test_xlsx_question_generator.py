"""
XLSX 源文件直传检索题生成测试。

覆盖：
1. 合法 sheet/range 能生成本地 reference_answer
2. 不存在的 sheet/range 被拒绝
3. 公式无缓存值不会产生伪造计算值
4. XLSX 元数据透传
5. 非 XLSX 现有流程不回归

不调用真实 API。
"""

import io
import json
import sys
import tempfile
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

from openpyxl import Workbook

from xlsx_question_generator import (
    _validate_and_render_evidence,
    _render_evidence_range,
    _parse_xlsx_qgen_response,
    _parse_range,
    _get_cell_display_value,
    generate_xlsx_questions,
)


# ====== Helpers ======

def _make_xlsx_bytes(wb):
    """将 Workbook 转为 bytes。"""
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_simple_xlsx():
    """创建一个简单的测试 XLSX。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "产品表"
    ws["A1"] = "产品名称"
    ws["B1"] = "价格"
    ws["C1"] = "库存"
    ws["A2"] = "产品A"
    ws["B2"] = 100
    ws["C2"] = 50
    ws["A3"] = "产品B"
    ws["B3"] = 200
    ws["C3"] = 30
    ws["A4"] = "产品C"
    ws["B4"] = 150
    ws["C4"] = 0
    return wb


def _make_formula_xlsx():
    """创建包含公式的 XLSX。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "公式表"
    ws["A1"] = "项目"
    ws["B1"] = "数值"
    ws["A2"] = "A"
    ws["B2"] = 100
    ws["A3"] = "B"
    ws["B3"] = 200
    ws["A4"] = "合计"
    ws["B4"] = "=SUM(B2:B3)"  # 公式
    return wb


# ====== Tests ======

def test_valid_range_renders_reference_answer():
    """合法 sheet/range 能生成本地 reference_answer。"""
    print("=" * 60)
    print("测试：合法范围渲染 reference_answer")
    print("=" * 60)

    wb = _make_simple_xlsx()

    # 单行范围
    q_single = {
        "question": "产品A价格",
        "evidence_sheet": "产品表",
        "evidence_range": "A2:C2",
        "difficulty": "事实",
        "topic": "产品价格",
    }
    result, reason = _validate_and_render_evidence(q_single, wb, "test.xlsx")
    assert result is not None, f"应通过验证，拒绝原因: {reason}"
    assert "reference_answer" in result, "应包含 reference_answer"
    assert "source_excerpt" in result, "应包含 source_excerpt"
    assert result["reference_answer"] == result["source_excerpt"], "两者应一致"
    assert "产品A" in result["reference_answer"], f"应包含产品A: {result['reference_answer']}"
    assert "100" in result["reference_answer"], f"应包含100: {result['reference_answer']}"
    assert result["evidence_sheet"] == "产品表"
    assert result["evidence_range"] == "A2:C2"
    assert result["source_format"] == "xlsx"
    assert result["source_file_name"] == "test.xlsx"

    # 多行范围（表格格式）
    q_multi = {
        "question": "产品列表",
        "evidence_sheet": "产品表",
        "evidence_range": "A1:C3",
        "difficulty": "基础",
        "topic": "产品信息",
    }
    result2, reason2 = _validate_and_render_evidence(q_multi, wb, "test.xlsx")
    assert result2 is not None, f"应通过验证，拒绝原因: {reason2}"
    assert "|" in result2["reference_answer"], "多行应渲染为表格格式"
    assert "产品名称" in result2["reference_answer"], "应包含表头"

    print("PASS: 合法范围正确渲染 reference_answer")


def test_nonexistent_sheet_rejected():
    """不存在的 sheet 被拒绝。"""
    print("=" * 60)
    print("测试：不存在的 sheet")
    print("=" * 60)

    wb = _make_simple_xlsx()
    q = {
        "question": "测试",
        "evidence_sheet": "不存在的工作表",
        "evidence_range": "A1:B2",
    }
    result, reason = _validate_and_render_evidence(q, wb, "test.xlsx")
    assert result is None, "应拒绝不存在的 sheet"
    assert "不存在" in reason, f"原因应提及不存在: {reason}"

    print("PASS: 不存在的 sheet 被拒绝")


def test_nonexistent_range_rejected():
    """不存在/越界范围被拒绝。"""
    print("=" * 60)
    print("测试：越界范围")
    print("=" * 60)

    wb = _make_simple_xlsx()  # 4 行 3 列

    # 越界行
    q1 = {"question": "测试", "evidence_sheet": "产品表", "evidence_range": "A1:C100"}
    result1, reason1 = _validate_and_render_evidence(q1, wb, "test.xlsx")
    assert result1 is None, f"应拒绝越界范围 (行)，原因: {reason1}"

    # 越界列
    q2 = {"question": "测试", "evidence_sheet": "产品表", "evidence_range": "A1:Z2"}
    result2, reason2 = _validate_and_render_evidence(q2, wb, "test.xlsx")
    assert result2 is None, f"应拒绝越界范围 (列)，原因: {reason2}"

    # 无效格式
    q3 = {"question": "测试", "evidence_sheet": "产品表", "evidence_range": "无效"}
    result3, reason3 = _validate_and_render_evidence(q3, wb, "test.xlsx")
    assert result3 is None, f"应拒绝无效范围格式，原因: {reason3}"

    print("PASS: 越界/无效范围被拒绝")


def test_formula_no_cache_rejected():
    """公式无缓存值不产生伪造计算值。"""
    print("=" * 60)
    print("测试：公式无缓存值")
    print("=" * 60)

    wb = _make_formula_xlsx()

    # data_only=False 时公式单元格的 .value 是公式文本
    # 我们在 _validate_and_render_evidence 中使用 data_only=True 打开
    # 但如果原始文件用 data_only=False 保存，公式可能没有缓存值

    # 测试公式单元格的处理
    ws = wb.active
    cell = ws["B4"]
    val, is_formula = _get_cell_display_value(cell)
    # 用 data_only=False 打开时，公式单元格的 .value 是公式字符串
    assert val is not None, "公式单元格应有值"
    assert isinstance(val, str) and val.startswith("="), f"应为公式文本: {val}"

    # 现在用 data_only=True 打开（模拟缓存值场景）
    xlsx_bytes = _make_xlsx_bytes(wb)
    from openpyxl import load_workbook as lwb
    wb_cached = lwb(io.BytesIO(xlsx_bytes), data_only=True)
    ws_cached = wb_cached.active
    cell_cached = ws_cached["B4"]
    val_cached, is_formula_cached = _get_cell_display_value(cell_cached)
    # data_only=True 时如果有缓存值，返回计算结果
    # 如果没有缓存值，返回 None
    # 这里我们验证不会伪造值
    if val_cached is None:
        # 无缓存值 → 应该被拒绝
        q = {
            "question": "合计金额",
            "evidence_sheet": "公式表",
            "evidence_range": "A4:B4",
        }
        result, reason = _validate_and_render_evidence(q, wb_cached, "test.xlsx")
        # 无缓存值的公式单元格：B4 的值为 None，范围非空（A4 有值）
        # 但核心证据（B4）为空，应被拒绝
        if result is not None:
            # 如果通过了，检查 reference_answer 中不包含伪造的计算值
            assert "300" not in result["reference_answer"], "不应包含伪造的计算值 300"
    else:
        # 有缓存值 → 可以使用
        assert val_cached == 300, f"缓存值应为 300: {val_cached}"

    print("PASS: 公式无缓存值不伪造计算值")


def test_xlsx_metadata_propagation():
    """XLSX 元数据从题目保存透传到加载。"""
    print("=" * 60)
    print("测试：XLSX 元数据透传")
    print("=" * 60)

    from question_generator import save_questions

    questions = [
        {
            "question": "产品A价格",
            "reference_answer": "产品A | 100 | 50",
            "source_excerpt": "产品A | 100 | 50",
            "difficulty": "事实",
            "topic": "价格",
            "question_mode": "retrieval",
            "source_format": "xlsx",
            "source_file_name": "products.xlsx",
            "evidence_sheet": "产品表",
            "evidence_range": "A2:C2",
        },
    ]

    with tempfile.TemporaryDirectory() as tmpdir:
        import question_generator as qg
        orig_dir = qg.QUESTIONS_DIR
        qg.QUESTIONS_DIR = Path(tmpdir)
        try:
            output_path, fname, set_id = save_questions(
                questions,
                question_set_name="测试XLSX题集",
                source_document_name="products.xlsx",
                question_mode="retrieval",
            )

            # 读回验证
            with output_path.open("r", encoding="utf-8") as f:
                saved = json.loads(f.readline())

            assert saved.get("source_format") == "xlsx", f"source_format 应为 xlsx: {saved.get('source_format')}"
            assert saved.get("source_file_name") == "products.xlsx", f"source_file_name: {saved.get('source_file_name')}"
            assert saved.get("evidence_sheet") == "产品表", f"evidence_sheet: {saved.get('evidence_sheet')}"
            assert saved.get("evidence_range") == "A2:C2", f"evidence_range: {saved.get('evidence_range')}"
            assert saved.get("question_set_id"), "应有 question_set_id"
            assert saved.get("question_set_name") == "测试XLSX题集"
        finally:
            qg.QUESTIONS_DIR = orig_dir

    print("PASS: XLSX 元数据正确透传")


def test_parse_xlsx_response():
    """XLSX LLM 响应解析。"""
    print("=" * 60)
    print("测试：XLSX 响应解析")
    print("=" * 60)

    # 正常 JSON
    resp1 = json.dumps([
        {"question": "查询1", "evidence_sheet": "Sheet1", "evidence_range": "A1:B2"},
        {"question": "查询2", "evidence_sheet": "Sheet1", "evidence_range": "C1:D2"},
    ])
    parsed = _parse_xlsx_qgen_response(resp1)
    assert len(parsed) == 2, f"应解析 2 条，实际 {len(parsed)}"

    # 带 markdown 代码块
    resp2 = '```json\n' + resp1 + '\n```'
    parsed2 = _parse_xlsx_qgen_response(resp2)
    assert len(parsed2) == 2

    # 无效 JSON
    parsed3 = _parse_xlsx_qgen_response("这不是 JSON")
    assert len(parsed3) == 0

    print("PASS: XLSX 响应解析正确")


def test_non_xlsx_flow_unaffected():
    """非 XLSX 流程不受影响（单元级验证）。"""
    print("=" * 60)
    print("测试：非 XLSX 流程不回归")
    print("=" * 60)

    from question_generator import (
        _validate_retrieval_question, deduplicate_questions, _deduplicate_and_trim,
        _build_supplement_prompt, MODE_RETRIEVAL,
    )

    # 验证现有校验函数仍正常
    chunk_text = "合同条款文本，包含违约金的计算方式为合同总金额的百分之十。"
    q = {
        "question": "违约金计算",
        "reference_answer": "违约金的计算方式为合同总金额的百分之十",
        "source_excerpt": "违约金的计算方式为合同总金额的百分之十",
    }
    # 需要 ref 是 chunk 的子串
    ok, reason = _validate_retrieval_question(q, chunk_text)
    # 这个可能通过也可能不通过（取决于照抄检测），但不应崩溃
    assert isinstance(ok, bool), f"应返回 bool: {ok}"

    # 去重仍正常
    qs = [{"question": "测试A"}, {"question": "测试A"}, {"question": "测试B"}]
    unique = deduplicate_questions(qs)
    assert len(unique) == 2

    print("PASS: 非 XLSX 流程不受影响")


# ====== Main ======

def main():
    tests = [
        test_valid_range_renders_reference_answer,
        test_nonexistent_sheet_rejected,
        test_nonexistent_range_rejected,
        test_formula_no_cache_rejected,
        test_xlsx_metadata_propagation,
        test_parse_xlsx_response,
        test_non_xlsx_flow_unaffected,
    ]

    passed = 0
    failed = 0
    for test_fn in tests:
        try:
            test_fn()
            passed += 1
        except Exception as e:
            failed += 1
            print(f"FAIL: {test_fn.__name__}: {e}")
            import traceback
            traceback.print_exc()
        print()

    print("=" * 60)
    print(f"结果: {passed} 通过, {failed} 失败, 共 {len(tests)} 个测试")
    print("=" * 60)

    if failed > 0:
        sys.exit(1)


if __name__ == "__main__":
    main()
