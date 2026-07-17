"""
XLSX 源文件直传检索题生成模块。

将原始 Excel 文件直接传给支持文件输入的 LLM，由 LLM 理解表格并定位证据位置。
本地只负责验证位置、保真渲染金标准证据和透传来源元数据。
"""

import base64
import io
import json
import re
from pathlib import Path

import requests
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

from question_generator import (
    deduplicate_questions, generate_question_set_id, save_questions,
    MODE_RETRIEVAL,
)

# ─── 常量 ────────────────────────────────────────────────────────────────────

PROMPT_TEMPLATE_PATH = Path(__file__).parent / "prompts" / "qgen_prompt_xlsx_retrieval.txt"
_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_MAX_EVIDENCE_ROWS = 20  # 单个证据范围最大行数
_MAX_EVIDENCE_COLS = 15  # 单个证据范围最大列数


# ─── LLM 调用 ────────────────────────────────────────────────────────────────

def _call_llm_with_file(prompt, file_bytes, file_name, api_key, base_url, model, timeout=120):
    """发送含文件附件的 multimodal 请求（OpenAI vision 格式）。"""
    url = base_url.rstrip("/") + "/chat/completions"
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }
    b64 = base64.b64encode(file_bytes).decode("utf-8")
    payload = {
        "model": model,
        "messages": [{
            "role": "user",
            "content": [
                {"type": "text", "text": prompt},
                {
                    "type": "image_url",
                    "image_url": {
                        "url": f"data:{_XLSX_MIME};base64,{b64}",
                    },
                },
            ],
        }],
        "temperature": 0,
    }

    try:
        resp = requests.post(url, json=payload, headers=headers, timeout=timeout)
    except requests.exceptions.Timeout:
        raise RuntimeError(f"请求超时 ({timeout}s): {url}")
    except requests.exceptions.ConnectionError as e:
        raise RuntimeError(f"连接失败: {url}\n{e}")

    if resp.status_code != 200:
        raise RuntimeError(
            f"HTTP {resp.status_code} | URL: {url}\nResponse: {resp.text[:1000]}"
        )

    try:
        data = resp.json()
    except Exception:
        raise RuntimeError(f"JSON 解析失败 | Response: {resp.text[:1000]}")

    try:
        return data["choices"][0]["message"]["content"]
    except (KeyError, IndexError):
        raise RuntimeError(f"响应结构异常 | Response: {json.dumps(data, ensure_ascii=False)[:1000]}")


def check_xlsx_llm_support(api_key, base_url, model, timeout=15):
    """检测模型/API 是否支持 XLSX 文件输入。

    发送一个最小 XLSX 文件附件请求，检查响应是否正常。
    Returns:
        bool: 是否支持
    """
    # 构建最小 XLSX
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Test"
    ws["A1"] = "test"
    buf = io.BytesIO()
    wb.save(buf)
    test_bytes = buf.getvalue()

    try:
        resp = _call_llm_with_file(
            "请回复 OK", test_bytes, "test.xlsx",
            api_key, base_url, model, timeout=timeout,
        )
        return bool(resp and len(resp.strip()) > 0)
    except Exception:
        return False


# ─── Prompt 构建 ──────────────────────────────────────────────────────────────

def _build_xlsx_prompt(wb, num_questions=5, difficulty="混合", topic_hint=""):
    """构建 XLSX 出题 prompt，包含工作表结构信息。"""
    template = PROMPT_TEMPLATE_PATH.read_text(encoding="utf-8").strip()

    topic_hint_section = ""
    if topic_hint:
        topic_hint_section = f"- 主题方向：{topic_hint}"

    prompt = template.replace("{num_questions}", str(num_questions))
    prompt = prompt.replace("{topic_hint_section}", topic_hint_section)
    return prompt


# ─── 响应解析 ─────────────────────────────────────────────────────────────────

def _parse_xlsx_qgen_response(text):
    """解析 LLM 返回的 JSON 数组。"""
    text = text.strip()

    # 去除 markdown 代码块
    if text.startswith("```"):
        lines = text.split("\n")
        json_lines = []
        in_block = False
        for line in lines:
            if line.strip().startswith("```") and not in_block:
                in_block = True
                continue
            elif line.strip() == "```" and in_block:
                break
            elif in_block:
                json_lines.append(line)
        text = "\n".join(json_lines).strip()

    # 尝试直接解析
    try:
        data = json.loads(text)
        if isinstance(data, list):
            return data
    except json.JSONDecodeError:
        pass

    # 尝试提取 JSON 数组
    match = re.search(r'\[[\s\S]*\]', text)
    if match:
        try:
            data = json.loads(match.group())
            if isinstance(data, list):
                return data
        except json.JSONDecodeError:
            pass

    return []


# ─── 证据验证与渲染 ──────────────────────────────────────────────────────────

def _parse_range(range_str):
    """解析 Excel 范围字符串，返回 (min_col, min_row, max_col, max_row)。

    openpyxl 的 range_boundaries 返回 (min_col, min_row, max_col, max_row)，1-indexed。
    """
    try:
        return range_boundaries(range_str)
    except Exception:
        return None


def _get_cell_display_value(cell):
    """获取单元格的显示值。

    公式单元格：若有缓存计算值则使用，否则返回公式文本并标记。
    """
    val = cell.value
    if val is None:
        return None, False

    # 检查是否为公式
    if isinstance(val, str) and val.startswith("="):
        # data_only=False 时公式单元格的 .value 是公式文本
        # 如果有 data_only=True 时的缓存值，openpyxl 会返回计算结果
        # 但我们现在用 data_only=False，所以公式单元格只有公式文本
        return val, True  # (value, is_formula)

    return val, False


def _validate_and_render_evidence(question, wb, file_name):
    """验证 LLM 返回的 sheet/range 并渲染金标准证据。

    Returns:
        tuple: (validated_question_or_None, rejection_reason)
    """
    sheet_name = (question.get("evidence_sheet") or "").strip()
    range_str = (question.get("evidence_range") or "").strip()
    q_text = (question.get("question") or "").strip()

    if not q_text:
        return None, "query 为空"
    if not sheet_name:
        return None, "evidence_sheet 为空"
    if not range_str:
        return None, "evidence_range 为空"

    # 检查工作表存在
    if sheet_name not in wb.sheetnames:
        return None, f"工作表 '{sheet_name}' 不存在，可用: {wb.sheetnames}"

    ws = wb[sheet_name]

    # 解析范围
    bounds = _parse_range(range_str)
    if bounds is None:
        return None, f"无法解析范围 '{range_str}'"

    min_col, min_row, max_col, max_row = bounds

    # 检查越界
    if min_row < 1 or min_col < 1:
        return None, f"范围 {range_str} 起始位置越界"
    if max_row > ws.max_row or max_col > ws.max_column:
        return None, f"范围 {range_str} 超出工作表边界 (max_row={ws.max_row}, max_col={ws.max_column})"

    # 检查范围大小
    rows = max_row - min_row + 1
    cols = max_col - min_col + 1
    if rows > _MAX_EVIDENCE_ROWS:
        return None, f"证据范围行数 ({rows}) 超过上限 {_MAX_EVIDENCE_ROWS}"
    if cols > _MAX_EVIDENCE_COLS:
        return None, f"证据范围列数 ({cols}) 超过上限 {_MAX_EVIDENCE_COLS}"

    # 读取单元格值
    has_formula_without_cache = False
    cell_values = []
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                            min_col=min_col, max_col=max_col):
        row_vals = []
        for cell in row:
            val, is_formula = _get_cell_display_value(cell)
            if is_formula:
                has_formula_without_cache = True
            row_vals.append(val)
        cell_values.append(row_vals)

    # 检查公式无缓存值
    if has_formula_without_cache:
        # 检查是否有任何公式单元格缺乏缓存值
        # data_only=False 时公式单元格的 .value 是公式字符串
        # 如果用户用 data_only=True 打开则会有缓存值
        # 这里我们拒绝纯公式文本作为证据
        formula_count = sum(
            1 for row in cell_values
            for v in row
            if isinstance(v, str) and v.startswith("=")
        )
        if formula_count > 0:
            return None, f"证据范围包含 {formula_count} 个公式单元格（无缓存计算值），拒绝"

    # 检查非空
    non_empty = sum(1 for row in cell_values for v in row if v is not None and str(v).strip())
    if non_empty == 0:
        return None, "证据范围为空"

    # 渲染为 Markdown
    rendered = _render_evidence_range(cell_values)

    # 组装验证后的题目
    validated = {
        "question": q_text,
        "reference_answer": rendered,
        "source_excerpt": rendered,
        "evidence_sheet": sheet_name,
        "evidence_range": range_str,
        "source_format": "xlsx",
        "source_file_name": file_name,
        "difficulty": question.get("difficulty", "事实"),
        "topic": question.get("topic", ""),
    }

    return validated, ""


def _render_evidence_range(cell_values):
    """将单元格值列表渲染为 Markdown 表格/键值格式。

    Args:
        cell_values: [[val, val, ...], [val, val, ...], ...]
    """
    if not cell_values:
        return ""

    # 单行：键值对格式
    if len(cell_values) == 1:
        row = cell_values[0]
        parts = []
        for v in row:
            if v is not None:
                parts.append(str(v))
        return " | ".join(parts) if parts else ""

    # 多行：表格格式
    lines = []
    # 首行作为表头
    header = cell_values[0]
    header_cells = [str(v) if v is not None else "" for v in header]
    lines.append("| " + " | ".join(header_cells) + " |")
    lines.append("| " + " | ".join(["---"] * len(header_cells)) + " |")

    for row in cell_values[1:]:
        row_cells = [str(v) if v is not None else "" for v in row]
        # 补齐列数
        while len(row_cells) < len(header_cells):
            row_cells.append("")
        lines.append("| " + " | ".join(row_cells[:len(header_cells)]) + " |")

    return "\n".join(lines)


# ─── 主入口 ───────────────────────────────────────────────────────────────────

def generate_xlsx_questions(file_bytes, file_name, api_key, base_url, model,
                            num_questions=5, difficulty="混合", topic_hint="",
                            timeout=120, progress_callback=None):
    """XLSX 直传检索题生成。

    Args:
        file_bytes: XLSX 文件原始字节
        file_name: 文件名
        api_key, base_url, model: LLM 配置
        num_questions: 目标题数
        difficulty: 难度偏好
        topic_hint: 主题方向
        timeout: LLM 超时秒数
        progress_callback: 进度回调

    Returns:
        tuple: (questions_list, stats_dict)
    """
    if progress_callback:
        progress_callback(0, 3, "解析 Excel 文件")

    # 打开 workbook 用于后续验证
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)

    if progress_callback:
        progress_callback(1, 3, "调用 LLM 生成检索查询")

    # 构建 prompt
    prompt = _build_xlsx_prompt(wb, num_questions, difficulty, topic_hint)

    # 调用 LLM
    response_text = _call_llm_with_file(
        prompt, file_bytes, file_name,
        api_key, base_url, model, timeout=timeout,
    )

    # 解析响应
    raw_questions = _parse_xlsx_qgen_response(response_text)
    raw_count = len(raw_questions)

    if progress_callback:
        progress_callback(2, 3, "验证证据并渲染金标准")

    # 验证并渲染
    valid_questions = []
    validation_eliminated = 0
    for q in raw_questions:
        validated, reason = _validate_and_render_evidence(q, wb, file_name)
        if validated:
            validated["question_mode"] = MODE_RETRIEVAL
            valid_questions.append(validated)
        else:
            validation_eliminated += 1
            print(f"  ⚠️ XLSX 证据校验不通过（{reason}），已过滤")

    wb.close()

    # 去重
    unique_questions = deduplicate_questions(valid_questions)
    dedup_eliminated = len(valid_questions) - len(unique_questions)

    # 裁剪到目标数
    if len(unique_questions) > num_questions:
        unique_questions = unique_questions[:num_questions]

    if not unique_questions:
        raise ValueError("XLSX 出题失败：所有查询均未通过证据校验")

    stats = {
        "raw_count": raw_count,
        "validation_eliminated": validation_eliminated,
        "dedup_eliminated": dedup_eliminated,
        "final_count": len(unique_questions),
        "target": num_questions,
    }
    return unique_questions, stats
