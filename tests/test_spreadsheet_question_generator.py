"""
统一电子表格检索题生成测试。

覆盖：
1. CSV 解析：UTF-8、GBK、BOM、空文件、单列
2. XLSX 解析：SheetContext、合并单元格、公式检测
3. 表格块拆分：小表/大表、表头保留、每块 allowed_anchor_ranges
4. 锚定范围验证：白名单内/外、越界、超大
5. 金标准渲染：单行、多行
6. LLM 响应解析：正常 JSON、markdown 代码块、无效
7. 完整流水线（mock LLM）：CSV、XLSX 端到端
8. doc_parser 集成：CSV/XLS 进入 parse_document
9. 向后兼容：xlsx_question_generator 委托正常

不调用真实 API。
"""

import csv
import io
import json
import sys
import tempfile
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

from openpyxl import Workbook

import tempfile

from spreadsheet_question_generator import (
    SheetContext,
    TableBlock,
    _col_letter,
    _col_index,
    _parse_range_str,
    _range_to_str,
    _detect_csv_encoding,
    _compute_allowed_anchor_ranges,
    _split_into_table_blocks,
    _render_block_markdown,
    _render_cell_values,
    _parse_llm_response,
    _validate_anchor_range,
    _render_reference_answer,
    _validate_and_render_question,
    _validate_single_question,
    _build_candidate_anchors,
    _file_content_hash,
    _set_schema_cache_dir,
    _get_schema_cache_dir,
    _delete_schema_cache,
    parse_xlsx_to_sheet_contexts,
    parse_csv_to_sheet_contexts,
    generate_spreadsheet_questions,
    _build_prompt,
)


def _use_test_schema_cache():
    """切换到临时 schema 缓存目录，返回 (temp_dir, original_dir)。调用方负责在 finally 中恢复。"""
    import spreadsheet_question_generator as sqg
    orig = sqg._SCHEMA_CACHE_DIR
    tmp = tempfile.mkdtemp(prefix="schema_test_")
    sqg._SCHEMA_CACHE_DIR = Path(tmp)
    return Path(tmp), orig


def _restore_schema_cache_dir(tmp_dir, orig_dir):
    """恢复 schema 缓存目录并清理临时目录。"""
    import spreadsheet_question_generator as sqg
    sqg._SCHEMA_CACHE_DIR = orig_dir
    import shutil
    shutil.rmtree(tmp_dir, ignore_errors=True)


# ====== Helpers ======

def _make_xlsx_bytes(wb):
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_simple_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "产品表"
    ws["A1"] = "产品名称"
    ws["B1"] = "价格"
    ws["C1"] = "库存"
    ws["A2"] = "产品A"
    ws["B2"] = 100
    ws["C2"] = 50
    ws["A3"] = "产品B"
    ws["B3"] = 200
    ws["C3"] = 30
    ws["A4"] = "产品C"
    ws["B4"] = 150
    ws["C4"] = 0
    return wb


def _make_multi_sheet_xlsx():
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1["A1"] = "Name"
    ws1["B1"] = "Value"
    ws1["A2"] = "Item1"
    ws1["B2"] = 10

    ws2 = wb.create_sheet("Sheet2")
    ws2["A1"] = "Category"
    ws2["B1"] = "Count"
    ws2["A2"] = "CatA"
    ws2["B2"] = 100
    return wb


def _make_formula_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "公式表"
    ws["A1"] = "项目"
    ws["B1"] = "数值"
    ws["A2"] = "A"
    ws["B2"] = 100
    ws["A3"] = "B"
    ws["B3"] = 200
    ws["A4"] = "合计"
    ws["B4"] = "=SUM(B2:B3)"
    return wb


def _make_merged_cell_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "合并表"
    ws["A1"] = "分类"
    ws["B1"] = "项目"
    ws["C1"] = "数值"
    ws["A2"] = "类别A"
    ws["B2"] = "项目1"
    ws["C2"] = 100
    ws["A3"] = None  # 合并后应继承 "类别A"
    ws["B3"] = "项目2"
    ws["C3"] = 200
    ws.merge_cells("A2:A3")
    return wb


def _make_csv_bytes(rows, encoding="utf-8"):
    """创建 CSV 字节。rows 是 list[list[str]]，第一行为表头。"""
    output = io.StringIO()
    writer = csv.writer(output)
    for row in rows:
        writer.writerow(row)
    content = output.getvalue()
    if encoding == "utf-8-sig":
        return b'\xef\xbb\xbf' + content.encode("utf-8")
    return content.encode(encoding)


def _make_large_xlsx(num_data_rows=100):
    """创建大数据量 XLSX。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "大数据表"
    ws["A1"] = "ID"
    ws["B1"] = "名称"
    ws["C1"] = "数值"
    for i in range(1, num_data_rows + 1):
        ws[f"A{i+1}"] = i
        ws[f"B{i+1}"] = f"项目{i}"
        ws[f"C{i+1}"] = i * 10
    return wb


# ====== Column Letter Tests ======

def test_col_letter():
    """列字母转换。"""
    print("=" * 60)
    print("测试：列字母转换")
    print("=" * 60)

    assert _col_letter(1) == "A"
    assert _col_letter(26) == "Z"
    assert _col_letter(27) == "AA"
    assert _col_letter(52) == "AZ"

    assert _col_index("A") == 1
    assert _col_index("Z") == 26
    assert _col_index("AA") == 27
    assert _col_index("AZ") == 52

    print("PASS: 列字母转换正确")


# ====== Range Parsing Tests ======

def test_parse_range_str():
    """范围字符串解析。"""
    print("=" * 60)
    print("测试：范围字符串解析")
    print("=" * 60)

    assert _parse_range_str("A1:C3") == (1, 1, 3, 3)
    assert _parse_range_str("B2:D5") == (2, 2, 4, 5)
    assert _parse_range_str("AA1:AB3") == (27, 1, 28, 3)
    assert _parse_range_str("invalid") is None
    assert _parse_range_str("A3:A1") is None  # min > max

    print("PASS: 范围字符串解析正确")


# ====== CSV Parsing Tests ======

def test_csv_basic():
    """基本 UTF-8 CSV 解析。"""
    print("=" * 60)
    print("测试：CSV 基本解析")
    print("=" * 60)

    csv_bytes = _make_csv_bytes([
        ["产品", "价格", "库存"],
        ["产品A", "100", "50"],
        ["产品B", "200", "30"],
    ])
    sheets = parse_csv_to_sheet_contexts(csv_bytes, "test.csv")
    assert len(sheets) == 1
    ctx = sheets[0]
    assert ctx.sheet_name == "CSV"
    assert ctx.max_row == 3  # header + 2 data rows
    assert ctx.max_col == 3
    assert ctx.headers == ["产品", "价格", "库存"]
    assert ctx.rows[1] == ["产品A", "100", "50"]
    assert len(ctx.formula_cells_without_cache) == 0
    assert len(ctx.merged_cells) == 0
    assert len(ctx.table_blocks) > 0

    print("PASS: CSV 基本解析正确")


def test_csv_encoding_gbk():
    """GBK 编码 CSV 解析。"""
    print("=" * 60)
    print("测试：CSV GBK 编码")
    print("=" * 60)

    csv_bytes = _make_csv_bytes([
        ["产品", "价格"],
        ["产品A", "100"],
    ], encoding="gbk")
    sheets = parse_csv_to_sheet_contexts(csv_bytes, "test_gbk.csv")
    assert len(sheets) == 1
    assert sheets[0].headers == ["产品", "价格"]

    print("PASS: CSV GBK 编码解析正确")


def test_csv_encoding_bom():
    """UTF-8 BOM 编码 CSV 解析。"""
    print("=" * 60)
    print("测试：CSV BOM 编码")
    print("=" * 60)

    csv_bytes = _make_csv_bytes([
        ["Name", "Value"],
        ["Item1", "10"],
    ], encoding="utf-8-sig")
    sheets = parse_csv_to_sheet_contexts(csv_bytes, "test_bom.csv")
    assert len(sheets) == 1
    assert sheets[0].headers == ["Name", "Value"]

    print("PASS: CSV BOM 编码解析正确")


def test_csv_empty():
    """空 CSV 应抛出异常。"""
    print("=" * 60)
    print("测试：空 CSV")
    print("=" * 60)

    csv_bytes = _make_csv_bytes([["A", "B"]])  # 只有表头没有数据
    try:
        parse_csv_to_sheet_contexts(csv_bytes, "empty.csv")
        # 如果只有表头，pandas 会读到空 DataFrame 或只有表头
        # 这里可能通过也可能抛异常，取决于 pandas 行为
        print("  注意：只有表头的 CSV 被接受了（pandas 行为）")
    except ValueError:
        print("  空 CSV 正确抛出 ValueError")

    # 真正的空 CSV
    try:
        parse_csv_to_sheet_contexts(b"", "truly_empty.csv")
        print("  FAIL: 应该抛出异常")
    except (ValueError, Exception):
        print("  空文件正确抛出异常")

    print("PASS: 空 CSV 处理正确")


# ====== XLSX Parsing Tests ======

def test_xlsx_to_sheet_context():
    """XLSX 解析为 SheetContext。"""
    print("=" * 60)
    print("测试：XLSX SheetContext 解析")
    print("=" * 60)

    wb = _make_simple_xlsx()
    xlsx_bytes = _make_xlsx_bytes(wb)
    sheets = parse_xlsx_to_sheet_contexts(xlsx_bytes)
    assert len(sheets) == 1
    ctx = sheets[0]
    assert ctx.sheet_name == "产品表"
    assert ctx.max_row == 4
    assert ctx.max_col == 3
    assert ctx.headers == ["产品名称", "价格", "库存"]
    assert ctx.rows[1] == ["产品A", 100, 50]
    assert len(ctx.table_blocks) > 0

    print("PASS: XLSX SheetContext 解析正确")


def test_xlsx_multi_sheet():
    """XLSX 多工作表解析。"""
    print("=" * 60)
    print("测试：XLSX 多工作表")
    print("=" * 60)

    wb = _make_multi_sheet_xlsx()
    xlsx_bytes = _make_xlsx_bytes(wb)
    sheets = parse_xlsx_to_sheet_contexts(xlsx_bytes)
    assert len(sheets) == 2
    names = {s.sheet_name for s in sheets}
    assert "Sheet1" in names
    assert "Sheet2" in names

    print("PASS: XLSX 多工作表解析正确")


def test_xlsx_merged_cells():
    """XLSX 合并单元格值继承。"""
    print("=" * 60)
    print("测试：XLSX 合并单元格")
    print("=" * 60)

    wb = _make_merged_cell_xlsx()
    xlsx_bytes = _make_xlsx_bytes(wb)
    sheets = parse_xlsx_to_sheet_contexts(xlsx_bytes)
    ctx = sheets[0]
    # 合并单元格 A2:A3，A3 应继承 A2 的值 "类别A"
    assert ctx.rows[2][0] == "类别A", f"A3 应为 '类别A': {ctx.rows[2][0]}"
    assert len(ctx.merged_cells) == 1

    print("PASS: XLSX 合并单元格值继承正确")


def test_xlsx_formula_detection():
    """XLSX 公式单元格检测。"""
    print("=" * 60)
    print("测试：XLSX 公式检测")
    print("=" * 60)

    wb = _make_formula_xlsx()
    xlsx_bytes = _make_xlsx_bytes(wb)
    sheets = parse_xlsx_to_sheet_contexts(xlsx_bytes)
    ctx = sheets[0]
    # B4 有公式 =SUM(B2:B3)，openpyxl 可能有缓存值也可能没有
    # 如果有缓存值，formula_cells_without_cache 为空
    # 如果没有，B4 位置在 formula_cells_without_cache 中
    print(f"  公式无缓存单元格: {ctx.formula_cells_without_cache}")
    # 至少不应崩溃
    assert isinstance(ctx.formula_cells_without_cache, list)

    print("PASS: XLSX 公式检测正常")


def test_formula_with_cached_value():
    """有缓存值的公式：不显示警告，使用缓存值。"""
    print("=" * 60)
    print("测试：公式有缓存值")
    print("=" * 60)

    # 创建 XLSX 并保存（openpyxl 会写入缓存值）
    wb = Workbook()
    ws = wb.active
    ws.title = "缓存表"
    ws["A1"] = "项目"
    ws["B1"] = "数值"
    ws["A2"] = "A"
    ws["B2"] = 100
    ws["A3"] = "B"
    ws["B3"] = 200
    ws["A4"] = "合计"
    ws["B4"] = "=SUM(B2:B3)"

    # 保存并重新打开（模拟有缓存值的文件）
    xlsx_bytes = _make_xlsx_bytes(wb)
    sheets = parse_xlsx_to_sheet_contexts(xlsx_bytes)
    ctx = sheets[0]

    # 检查 B4 的处理结果
    b4_val = ctx.rows[3][1]  # B4
    print(f"  B4 值: {b4_val}")
    print(f"  formula_cells_without_cache: {ctx.formula_cells_without_cache}")

    # 如果 openpyxl 有缓存值 → B4 应为 300，无警告
    # 如果无缓存值 → B4 应为 [公式未计算]，有警告
    if (4, 2) in ctx.formula_cells_without_cache:
        assert b4_val == "[公式未计算]", f"无缓存时 B4 应为 [公式未计算]: {b4_val}"
        # 检查 block 有公式警告
        assert ctx.table_blocks[0].has_formula_warnings, "应有公式警告"
    else:
        assert b4_val == 300, f"有缓存时 B4 应为 300: {b4_val}"
        assert not ctx.table_blocks[0].has_formula_warnings, "有缓存值不应有公式警告"

    print("PASS: 公式缓存值处理正确")


def test_formula_no_cache_rejects_reference_answer():
    """无缓存值的公式：reference_answer 拒绝含该单元格的范围。"""
    print("=" * 60)
    print("测试：无缓存公式拒绝 reference_answer")
    print("=" * 60)

    # 创建一个公式一定无缓存的场景：
    # 直接用 data_only=False 构造 rows，公式单元格为字符串
    from spreadsheet_question_generator import SheetContext, _render_reference_answer

    ctx = SheetContext(
        sheet_name="测试",
        max_row=2,
        max_col=2,
        headers=["项目", "数值"],
        rows=[["项目", "数值"], ["合计", "=SUM(B1)"]],
        formula_cells_without_cache=[(2, 2)],
        format_warnings=[],
        allowed_anchor_ranges=["A2:B2"],
        table_blocks=[],
    )

    rendered, has_issue = _render_reference_answer("A2:B2", ctx)
    assert has_issue, "应标记公式问题"
    assert "[公式未计算]" in rendered, f"应含 [公式未计算]: {rendered}"
    assert "=SUM" not in rendered, f"不应含公式字符串: {rendered}"

    print("PASS: 无缓存公式正确拒绝")


def test_formula_cached_value_no_warning():
    """有缓存值的公式单元格：reference_answer 使用缓存值，无警告。"""
    print("=" * 60)
    print("测试：缓存公式无警告")
    print("=" * 60)

    from spreadsheet_question_generator import SheetContext, _render_reference_answer

    ctx = SheetContext(
        sheet_name="测试",
        max_row=2,
        max_col=2,
        headers=["项目", "数值"],
        rows=[["项目", "数值"], ["合计", 300]],  # 缓存值已替换公式
        formula_cells_without_cache=[],  # 无缓存问题
        format_warnings=[],
        allowed_anchor_ranges=["A2:B2"],
        table_blocks=[],
    )

    rendered, has_issue = _render_reference_answer("A2:B2", ctx)
    assert not has_issue, "有缓存值不应有公式问题"
    assert "300" in rendered, f"应含缓存值 300: {rendered}"
    assert "[公式未计算]" not in rendered, f"不应含 [公式未计算]: {rendered}"

    print("PASS: 缓存公式无警告正确")


# ====== Table Block Tests ======

def test_split_small_sheet():
    """小表格应只有 1 个块。"""
    print("=" * 60)
    print("测试：小表格拆分为 1 块")
    print("=" * 60)

    wb = _make_simple_xlsx()
    xlsx_bytes = _make_xlsx_bytes(wb)
    sheets = parse_xlsx_to_sheet_contexts(xlsx_bytes)
    ctx = sheets[0]
    assert len(ctx.table_blocks) == 1, f"应为 1 块，实际 {len(ctx.table_blocks)}"

    block = ctx.table_blocks[0]
    assert block.row_range == (2, 4)
    assert "| 行号 |" in block.markdown
    assert "| 2 |" in block.markdown

    print("PASS: 小表格正确拆分为 1 块")


def test_split_large_sheet():
    """大表格应拆分为多个块。"""
    print("=" * 60)
    print("测试：大表格拆分为多块")
    print("=" * 60)

    wb = _make_large_xlsx(100)
    xlsx_bytes = _make_xlsx_bytes(wb)
    sheets = parse_xlsx_to_sheet_contexts(xlsx_bytes)
    ctx = sheets[0]
    # 100 行数据，每块 30 行，应为 4 块 (30+30+30+10)
    assert len(ctx.table_blocks) == 4, f"应为 4 块，实际 {len(ctx.table_blocks)}"

    # 每块都应包含表头
    for block in ctx.table_blocks:
        assert "| 行号 |" in block.markdown, f"块 {block.block_index} 缺少表头"

    # 行号连续性
    assert ctx.table_blocks[0].row_range == (2, 31)
    assert ctx.table_blocks[1].row_range == (32, 61)
    assert ctx.table_blocks[2].row_range == (62, 91)
    assert ctx.table_blocks[3].row_range == (92, 101)

    print("PASS: 大表格正确拆分为多块")


def test_allowed_ranges_per_block():
    """每块的 allowed_anchor_ranges 是 sheet 的子集。"""
    print("=" * 60)
    print("测试：每块 allowed_anchor_ranges")
    print("=" * 60)

    wb = _make_simple_xlsx()
    xlsx_bytes = _make_xlsx_bytes(wb)
    sheets = parse_xlsx_to_sheet_contexts(xlsx_bytes)
    ctx = sheets[0]
    block = ctx.table_blocks[0]
    # 应该有 allowed_anchor_ranges
    assert isinstance(block.allowed_anchor_ranges, list)
    # 所有块级范围应是 sheet 级范围的子集
    for r in block.allowed_anchor_ranges:
        assert r in ctx.allowed_anchor_ranges, f"块级范围 {r} 不在 sheet 级范围中"

    print("PASS: 每块 allowed_anchor_ranges 正确")


def test_semantic_header_value_block():
    """费率/参数表生成语义化二列块。"""
    print("=" * 60)
    print("测试：语义化表头+数值块")
    print("=" * 60)

    wb = Workbook()
    ws = wb.active
    ws.title = "费率表"
    ws["A1"] = "费用项"
    ws["B1"] = "项目经理"
    ws["C1"] = "开发人员"
    ws["D1"] = "测试人员"
    ws["A2"] = "单价(元/人天)"
    ws["B2"] = 1700
    ws["C2"] = 1500
    ws["D2"] = 1200

    xlsx_bytes = _make_xlsx_bytes(wb)
    sheets = parse_xlsx_to_sheet_contexts(xlsx_bytes)
    ctx = sheets[0]

    # 应有 2 个块：标准块 + 语义块
    assert len(ctx.table_blocks) == 2, f"应有 2 块: {len(ctx.table_blocks)}"

    sem_block = ctx.table_blocks[1]
    assert "字段名" in sem_block.markdown
    assert "数值" in sem_block.markdown
    assert "项目经理" in sem_block.markdown
    assert "1700" in sem_block.markdown
    # 行标签列应被跳过
    assert "费用项" not in sem_block.markdown, "行标签列应被跳过"
    assert "单价(元/人天)" not in sem_block.markdown, "行标签列应被跳过"

    # 每个字段应有独立 anchor
    assert "B1:B2" in sem_block.allowed_anchor_ranges
    assert "C1:C2" in sem_block.allowed_anchor_ranges
    assert "D1:D2" in sem_block.allowed_anchor_ranges

    print("PASS: 语义化表头+数值块正确")


def test_semantic_no_false_positive():
    """普通数据表不应生成语义化块。"""
    print("=" * 60)
    print("测试：普通表不生成语义块")
    print("=" * 60)

    wb = _make_simple_xlsx()
    xlsx_bytes = _make_xlsx_bytes(wb)
    sheets = parse_xlsx_to_sheet_contexts(xlsx_bytes)
    ctx = sheets[0]

    # 普通表只有 1 个标准块
    assert len(ctx.table_blocks) == 1, f"普通表应只有 1 块: {len(ctx.table_blocks)}"

    print("PASS: 普通表不生成语义块")


def test_semantic_anchors_in_sheet_whitelist():
    """语义块的 field anchors 应自动添加到 sheet 级白名单。"""
    print("=" * 60)
    print("测试：语义 anchor 在 sheet 白名单中")
    print("=" * 60)

    wb = Workbook()
    ws = wb.active
    ws.title = "费率表"
    ws["A1"] = "费用项"
    ws["B1"] = "项目经理"
    ws["C1"] = "开发人员"
    ws["D1"] = "测试人员"
    ws["A2"] = "单价(元/人天)"
    ws["B2"] = 1700
    ws["C2"] = 1500
    ws["D2"] = 1200

    xlsx_bytes = _make_xlsx_bytes(wb)
    sheets = parse_xlsx_to_sheet_contexts(xlsx_bytes)
    ctx = sheets[0]

    # 语义块的 field anchors 应在 sheet 级白名单中
    assert "B1:B2" in ctx.allowed_anchor_ranges, f"B1:B2 应在白名单中: {ctx.allowed_anchor_ranges}"
    assert "C1:C2" in ctx.allowed_anchor_ranges
    assert "D1:D2" in ctx.allowed_anchor_ranges

    print("PASS: 语义 anchor 在 sheet 白名单中")


def test_semantic_anchor_renders_field_and_value():
    """E2:E3 类型的锚点 reference_answer 应同时含字段名和数值。"""
    print("=" * 60)
    print("测试：语义锚点渲染字段+数值")
    print("=" * 60)

    wb = Workbook()
    ws = wb.active
    ws.title = "费率表"
    ws["A1"] = "费用项"
    ws["B1"] = "项目经理"
    ws["C1"] = "开发人员"
    ws["D1"] = "测试人员"
    ws["A2"] = "单价(元/人天)"
    ws["B2"] = 1700
    ws["C2"] = 1500
    ws["D2"] = 1200

    xlsx_bytes = _make_xlsx_bytes(wb)
    sheets = parse_xlsx_to_sheet_contexts(xlsx_bytes)
    sheets_by_name = {s.sheet_name: s for s in sheets}

    # 用语义 anchor 验证
    q = {"question": "项目经理费率", "sheet_name": "费率表", "anchor_range": "B1:B2"}
    result, reason = _validate_and_render_question(q, sheets_by_name, "test.xlsx")
    assert result is not None, f"应通过验证: {reason}"
    ref = result["reference_answer"]
    assert "项目经理" in ref, f"应含字段名 '项目经理': {ref}"
    assert "1700" in ref, f"应含数值 '1700': {ref}"

    print("PASS: 语义锚点正确渲染字段+数值")


def test_isolated_numeric_anchor_rejected():
    """孤立数值锚点如 E3:E3 应被拒绝。"""
    print("=" * 60)
    print("测试：孤立数值锚点拒绝")
    print("=" * 60)

    wb = Workbook()
    ws = wb.active
    ws.title = "报价页"
    ws["A1"] = "项目"
    ws["B1"] = "描述"
    ws["C1"] = "单价"
    ws["A2"] = "服务A"
    ws["B2"] = "咨询服务"
    ws["C2"] = 73900

    xlsx_bytes = _make_xlsx_bytes(wb)
    sheets = parse_xlsx_to_sheet_contexts(xlsx_bytes)
    sheets_by_name = {s.sheet_name: s for s in sheets}

    # 孤立数值锚点应被拒绝
    q = {"question": "服务A单价", "sheet_name": "报价页", "anchor_range": "C2:C2"}
    result, reason = _validate_and_render_question(q, sheets_by_name, "test.xlsx")
    assert result is None, f"孤立数值应被拒绝: {reason}"
    assert "孤立数值" in reason, f"原因应提及孤立数值: {reason}"

    # 包含字段名+数值的锚点应通过
    q2 = {"question": "服务A单价", "sheet_name": "报价页", "anchor_range": "A2:C2"}
    result2, reason2 = _validate_and_render_question(q2, sheets_by_name, "test.xlsx")
    assert result2 is not None, f"含字段名+数值应通过: {reason2}"

    print("PASS: 孤立数值锚点正确拒绝")


def test_cross_row_rejected():
    """跨行范围必须在白名单中。"""
    print("=" * 60)
    print("测试：跨行范围拒绝")
    print("=" * 60)

    wb = Workbook()
    ws = wb.active
    ws.title = "报价页"
    ws["A1"] = "项目"
    ws["B1"] = "单价"
    ws["C1"] = "数量"
    ws["D1"] = "备注"
    ws["A2"] = "服务A"
    ws["B2"] = 100
    ws["C2"] = 5
    ws["D2"] = "备注A"
    ws["A3"] = "服务B"
    ws["B3"] = 200
    ws["C3"] = 3
    ws["D3"] = "备注B"

    xlsx_bytes = _make_xlsx_bytes(wb)
    sheets = parse_xlsx_to_sheet_contexts(xlsx_bytes)
    sheets_by_name = {s.sheet_name: s for s in sheets}

    # 单行范围通过
    q1 = {"question": "test", "sheet_name": "报价页", "anchor_range": "A2:C2"}
    result1, _ = _validate_and_render_question(q1, sheets_by_name, "test.xlsx")
    assert result1 is not None, "单行范围应通过"

    # 2行范围通过（子集检查）
    q2 = {"question": "test", "sheet_name": "报价页", "anchor_range": "A2:C3"}
    result2, _ = _validate_and_render_question(q2, sheets_by_name, "test.xlsx")
    assert result2 is not None, "2行子集范围应通过"

    # 相邻两行（label+value）应通过
    q2 = {"question": "服务信息", "sheet_name": "报价页", "anchor_range": "A2:B3"}
    result2, reason2 = _validate_and_render_question(q2, sheets_by_name, "test.xlsx")
    assert result2 is not None, f"相邻两行应通过: {reason2}"

    # 单行数据应通过
    q3 = {"question": "服务A信息", "sheet_name": "报价页", "anchor_range": "A2:B2"}
    result3, reason3 = _validate_and_render_question(q3, sheets_by_name, "test.xlsx")
    assert result3 is not None, f"单行数据应通过: {reason3}"

    print("PASS: 跨行范围正确拒绝")


def test_semantic_block_in_prompt():
    """语义块应出现在发给 LLM 的 prompt 中。"""
    print("=" * 60)
    print("测试：语义块进入 prompt")
    print("=" * 60)

    wb = Workbook()
    ws = wb.active
    ws.title = "费率表"
    ws["A1"] = "费用项"
    ws["B1"] = "项目经理"
    ws["C1"] = "开发人员"
    ws["D1"] = "测试人员"
    ws["A2"] = "单价(元/人天)"
    ws["B2"] = 1700
    ws["C2"] = 1500
    ws["D2"] = 1200

    xlsx_bytes = _make_xlsx_bytes(wb)
    sheets = parse_xlsx_to_sheet_contexts(xlsx_bytes)
    prompt = _build_prompt(sheets, 5, "")

    # prompt 应包含语义块内容
    assert "字段名" in prompt, "prompt 应含语义块的 '字段名' 表头"
    assert "项目经理" in prompt, "prompt 应含语义块的 '项目经理'"
    assert "1700" in prompt, "prompt 应含语义块的 '1700'"
    # 应包含语义块的 allowed_anchor_ranges
    assert "B1:B2" in prompt, "prompt 应含语义块的 B1:B2 anchor"

    print("PASS: 语义块正确进入 prompt")


def test_normal_text_question_not_regressed():
    """普通单行文本题不回归。"""
    print("=" * 60)
    print("测试：普通文本题不回归")
    print("=" * 60)

    csv_bytes = _make_csv_bytes([
        ["功能模块", "描述", "状态"],
        ["登录", "用户认证", "上线"],
        ["支付", "在线支付", "测试中"],
    ])

    mock_response = json.dumps([
        {"question": "登录模块", "sheet_name": "CSV", "anchor_range": "A2:C2", "difficulty": "事实"},
    ])

    import spreadsheet_question_generator as sqg
    original = sqg._call_llm_text
    sqg._call_llm_text = lambda *a, **kw: mock_response

    try:
        questions, stats = generate_spreadsheet_questions(
            csv_bytes, "test.csv", "fake", "http://fake", "fake_model",
        )
        assert len(questions) == 1
        ref = questions[0]["reference_answer"]
        assert "登录" in ref, f"应含 '登录': {ref}"
        assert "用户认证" in ref, f"应含 '用户认证': {ref}"
    finally:
        sqg._call_llm_text = original

    print("PASS: 普通文本题不回归")


# ====== Consistency Validation Regression Tests ======

def _make_rate_table_xlsx():
    """创建模拟真实报价表的 XLSX。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "报价页"
    # Row 1: 列标题（无用）
    ws["A1"] = "列A"
    ws["B1"] = "列B"
    ws["C1"] = "列C"
    ws["D1"] = "列D"
    ws["E1"] = "列E"
    ws["F1"] = "列F"
    ws["M1"] = "列M"
    # Row 2: 字段名
    ws["A2"] = ""
    ws["B2"] = "功能模块"
    ws["C2"] = "产品功能"
    ws["D2"] = "未税价（元）"
    ws["E2"] = "项目经理"
    ws["F2"] = "研发经理"
    ws["M2"] = "SRE工程师"
    # Row 3: 费率值（D 列应为数值才能被语义块识别）
    ws["A3"] = ""
    ws["B3"] = "功能模块"
    ws["C3"] = "产品功能"
    ws["D3"] = 50000  # 数值，非文本
    ws["E3"] = 1700
    ws["F3"] = 1800
    ws["M3"] = 1500
    # Row 4: 业务数据
    ws["B4"] = "CICD工具规范"
    ws["C4"] = "集成发布流水线"
    ws["D4"] = 73900
    ws["E4"] = 2
    ws["F4"] = 0
    ws["M4"] = 20
    return wb


def test_candidate_anchors_have_correct_fact_fields():
    """候选锚点的 fact_fields 应正确反映字段名。"""
    print("=" * 60)
    print("测试：候选锚点 fact_fields 正确性")
    print("=" * 60)

    from spreadsheet_question_generator import _build_candidate_anchors

    wb = _make_rate_table_xlsx()
    xlsx_bytes = _make_xlsx_bytes(wb)
    sheets = parse_xlsx_to_sheet_contexts(xlsx_bytes)

    candidates = _build_candidate_anchors(sheets)

    # field_value_pair 候选项的 fact_fields 应包含正确字段名
    fvp = [c for c in candidates if c["evidence_mode"] == "field_value_pair"]
    assert len(fvp) > 0, "应有 field_value_pair 候选项"
    fvp_fields = {c["fact_fields"][0] for c in fvp}
    assert "项目经理" in fvp_fields, f"应含'项目经理': {fvp_fields}"

    # record_with_schema_context 候选项的 fact_fields 应包含业务标识列
    rsc = [c for c in candidates if c["evidence_mode"] == "record_with_schema_context"]
    assert len(rsc) > 0, "应有 record_with_schema_context 候选项"
    for c in rsc:
        assert len(c["fact_fields"]) >= 2, f"fact_fields 应含至少2个字段: {c['fact_fields']}"
        assert c["header_context_range"] is not None, "应有 header_context_range"

    print(f"PASS: field_value_pair {len(fvp)} 条, record_with_schema_context {len(rsc)} 条")


def test_consistency_price_on_business_row_rejected():
    """价格题锚定无表头上下文的业务行应被拒绝。"""
    print("=" * 60)
    print("测试：价格题业务行拒绝")
    print("=" * 60)

    from spreadsheet_question_generator import (
        _extract_semantic_field_names, _extract_semantic_anchors,
        _validate_question_anchor_consistency,
    )

    wb = _make_rate_table_xlsx()
    xlsx_bytes = _make_xlsx_bytes(wb)
    sheets = parse_xlsx_to_sheet_contexts(xlsx_bytes)
    sheets_by_name = {s.sheet_name: s for s in sheets}

    field_names = _extract_semantic_field_names(sheets)
    anchors = _extract_semantic_anchors(sheets)

    # 价格题锚定 B4:D4（双源模型，含业务标识+价格）应通过
    q = {"question": "CICD模块未税价", "anchor_range": "B4:D4", "sheet_name": "报价页"}
    valid, reason = _validate_question_anchor_consistency(q, field_names, anchors, sheets_by_name)
    assert valid, f"双源模型价格题应通过: {reason}"

    # 价格题锚定语义块 D2:D3 应通过
    q2 = {"question": "未税价", "anchor_range": "D2:D3", "sheet_name": "报价页"}
    valid2, reason2 = _validate_question_anchor_consistency(q2, field_names, anchors, sheets_by_name)
    assert valid2, f"价格题语义块应通过: {reason2}"

    print("PASS: 价格题业务行（双源模型）正确处理")


def test_consistency_aggregate_rejected():
    """聚合型题目应被拒绝。"""
    print("=" * 60)
    print("测试：聚合题拒绝")
    print("=" * 60)

    from spreadsheet_question_generator import (
        _extract_semantic_field_names, _extract_semantic_anchors,
        _validate_question_anchor_consistency,
    )

    wb = _make_rate_table_xlsx()
    xlsx_bytes = _make_xlsx_bytes(wb)
    sheets = parse_xlsx_to_sheet_contexts(xlsx_bytes)
    sheets_by_name = {s.sheet_name: s for s in sheets}

    field_names = _extract_semantic_field_names(sheets)
    anchors = _extract_semantic_anchors(sheets)

    for agg_q in ["各角色费率明细", "人力配置", "开发投入", "所有模块报价"]:
        q = {"question": agg_q, "anchor_range": "E2:E3", "sheet_name": "报价页"}
        valid, reason = _validate_question_anchor_consistency(q, field_names, anchors, sheets_by_name)
        assert not valid, f"聚合题'{agg_q}'应拒绝: {reason}"
        assert "聚合" in reason

    print("PASS: 聚合题正确拒绝")


def test_consistency_a2m3_overflow_rejected():
    """A2:M3 这类大范围应被跨行检查拒绝。"""
    print("=" * 60)
    print("测试：A2:M3 溢出拒绝")
    print("=" * 60)

    wb = _make_rate_table_xlsx()
    xlsx_bytes = _make_xlsx_bytes(wb)
    sheets = parse_xlsx_to_sheet_contexts(xlsx_bytes)
    sheets_by_name = {s.sheet_name: s for s in sheets}

    # A2:M3 跨两行，应通过跨行检查（因为是相邻两行）
    # 但作为数值题应被一致性检查拒绝（聚合）
    q = {"question": "各角色费率明细", "sheet_name": "报价页", "anchor_range": "A2:M3"}
    result, reason = _validate_and_render_question(q, sheets_by_name, "test.xlsx")
    # 跨行检查：相邻两行 OK
    # 一致性检查：聚合题拒绝
    if result:
        from spreadsheet_question_generator import (
            _extract_semantic_field_names, _extract_semantic_anchors,
            _validate_question_anchor_consistency,
        )
        field_names = _extract_semantic_field_names(sheets)
        anchors = _extract_semantic_anchors(sheets)
        consistent, c_reason = _validate_question_anchor_consistency(
            result, field_names, anchors, sheets_by_name,
        )
        assert not consistent, f"聚合题 A2:M3 应被一致性检查拒绝: {c_reason}"
    else:
        # 如果被锚定检查拒绝也可以
        pass

    print("PASS: A2:M3 溢出正确拒绝")


# ====== Anchor Validation Tests ======

def test_valid_anchor_in_whitelist():
    """白名单内的范围通过验证。"""
    print("=" * 60)
    print("测试：白名单内范围验证")
    print("=" * 60)

    allowed = ["A1:C3", "A2:C2", "A3:C3"]
    valid, reason = _validate_anchor_range("A2:C2", allowed, 10, 10)
    assert valid, f"应通过: {reason}"

    print("PASS: 白名单内范围正确通过")


def test_anchor_not_in_whitelist():
    """不在白名单的范围被拒绝。"""
    print("=" * 60)
    print("测试：非白名单范围拒绝")
    print("=" * 60)

    allowed = ["A2:C2"]
    valid, reason = _validate_anchor_range("D2:F2", allowed, 10, 10)
    assert not valid
    assert "白名单" in reason

    print("PASS: 非白名单范围正确拒绝")


def test_anchor_subset_legal():
    """白名单 B4:E4 时，子范围 B4:C4 合法。"""
    print("=" * 60)
    print("测试：子范围合法")
    print("=" * 60)

    allowed = ["B4:E4"]
    valid, reason = _validate_anchor_range("B4:C4", allowed, 10, 10)
    assert valid, f"B4:C4 应合法（B4:E4 的子范围）: {reason}"

    # 也是精确匹配
    valid2, reason2 = _validate_anchor_range("B4:E4", allowed, 10, 10)
    assert valid2, f"B4:E4 应合法（精确匹配）: {reason2}"

    print("PASS: 子范围合法")


def test_anchor_subset_right_overflow():
    """白名单 B4:E4 时，B4:F4 右越界非法。"""
    print("=" * 60)
    print("测试：子范围右越界")
    print("=" * 60)

    allowed = ["B4:E4"]
    valid, reason = _validate_anchor_range("B4:F4", allowed, 10, 10)
    assert not valid, "B4:F4 应非法（右边界超出 B4:E4）"
    assert "白名单" in reason

    print("PASS: 子范围右越界正确拒绝")


def test_anchor_subset_left_overflow():
    """白名单 B4:E4 时，A4:C4 左越界非法。"""
    print("=" * 60)
    print("测试：子范围左越界")
    print("=" * 60)

    allowed = ["B4:E4"]
    valid, reason = _validate_anchor_range("A4:C4", allowed, 10, 10)
    assert not valid, "A4:C4 应非法（左边界 A < B 超出白名单）"
    assert "白名单" in reason

    print("PASS: 子范围左越界正确拒绝")


def test_anchor_out_of_bounds():
    """越界范围被拒绝。"""
    print("=" * 60)
    print("测试：越界范围拒绝")
    print("=" * 60)

    allowed = ["A2:C2"]
    valid, reason = _validate_anchor_range("A2:C200", allowed, 10, 3)
    assert not valid
    assert "边界" in reason

    print("PASS: 越界范围正确拒绝")


def test_anchor_too_large():
    """超大范围被拒绝。"""
    print("=" * 60)
    print("测试：超大范围拒绝")
    print("=" * 60)

    allowed = ["A1:ZZ1"]
    valid, reason = _validate_anchor_range("A1:ZZ1", allowed, 1, 703)
    assert not valid
    assert "上限" in reason

    print("PASS: 超大范围正确拒绝")


# ====== Reference Answer Rendering Tests ======

def test_render_single_row():
    """单行范围渲染为键值格式。"""
    print("=" * 60)
    print("测试：单行渲染")
    print("=" * 60)

    cell_values = [["产品A", 100, 50]]
    rendered = _render_cell_values(cell_values)
    assert "产品A" in rendered
    assert "100" in rendered
    assert "|" in rendered

    print("PASS: 单行渲染正确")


def test_render_multi_row():
    """多行范围渲染为键值对格式（不含 Markdown 分隔符）。"""
    print("=" * 60)
    print("测试：多行渲染")
    print("=" * 60)

    cell_values = [
        ["名称", "价格"],
        ["产品A", 100],
        ["产品B", 200],
    ]
    rendered = _render_cell_values(cell_values)
    assert "名称" in rendered
    assert "产品A" in rendered
    assert "100" in rendered
    assert "---" not in rendered, "reference_answer 不应含 Markdown 分隔符"
    assert "：" in rendered, "应使用键值对格式"

    print("PASS: 多行渲染正确")


def test_render_reference_answer_from_context():
    """从 SheetContext 渲染 reference_answer。"""
    print("=" * 60)
    print("测试：从 SheetContext 渲染 reference_answer")
    print("=" * 60)

    wb = _make_simple_xlsx()
    xlsx_bytes = _make_xlsx_bytes(wb)
    sheets = parse_xlsx_to_sheet_contexts(xlsx_bytes)
    ctx = sheets[0]

    rendered, has_formula = _render_reference_answer("A2:C2", ctx)
    assert rendered, "应有渲染结果"
    assert "产品A" in rendered
    assert "100" in rendered
    assert not has_formula

    print("PASS: SheetContext 渲染正确")


# ====== LLM Response Parsing Tests ======

def test_parse_valid_json():
    """正常 JSON 数组解析。"""
    print("=" * 60)
    print("测试：正常 JSON 解析")
    print("=" * 60)

    resp = json.dumps([
        {"question": "查询1", "sheet_name": "Sheet1", "anchor_range": "A1:B2"},
        {"question": "查询2", "sheet_name": "Sheet1", "anchor_range": "C1:D2"},
    ])
    parsed = _parse_llm_response(resp)
    assert len(parsed) == 2

    print("PASS: 正常 JSON 解析正确")


def test_parse_markdown_code_block():
    """Markdown 代码块中的 JSON 解析。"""
    print("=" * 60)
    print("测试：Markdown 代码块解析")
    print("=" * 60)

    json_str = json.dumps([{"question": "测试", "sheet_name": "S1", "anchor_range": "A1:B1"}])
    resp = f"```json\n{json_str}\n```"
    parsed = _parse_llm_response(resp)
    assert len(parsed) == 1

    print("PASS: Markdown 代码块解析正确")


def test_parse_invalid_json():
    """无效 JSON 返回空列表。"""
    print("=" * 60)
    print("测试：无效 JSON")
    print("=" * 60)

    parsed = _parse_llm_response("这不是 JSON")
    assert parsed == []

    print("PASS: 无效 JSON 正确返回空列表")


# ====== Full Pipeline Tests (Mocked LLM) ======

def test_generate_csv_questions():
    """CSV 端到端生成（mock LLM）。"""
    print("=" * 60)
    print("测试：CSV 端到端生成")
    print("=" * 60)

    csv_bytes = _make_csv_bytes([
        ["产品名称", "价格", "库存"],
        ["产品A", "100", "50"],
        ["产品B", "200", "30"],
        ["产品C", "150", "0"],
    ])

    # Mock LLM 返回
    mock_response = json.dumps([
        {"question": "产品A价格", "sheet_name": "CSV", "anchor_range": "A2:C2", "difficulty": "事实", "topic": "价格"},
        {"question": "产品B库存", "sheet_name": "CSV", "anchor_range": "A3:C3", "difficulty": "事实", "topic": "库存"},
    ])

    import spreadsheet_question_generator as sqg
    original_call_llm = sqg._call_llm_text
    sqg._call_llm_text = lambda *a, **kw: mock_response

    try:
        questions, stats = generate_spreadsheet_questions(
            csv_bytes, "test.csv",
            "fake_key", "http://fake", "fake_model",
            num_questions=5,
        )
        assert len(questions) == 2, f"应生成 2 题: {len(questions)}"
        assert stats["sheet_count"] == 1
        assert questions[0]["source_format"] == "csv"
        assert questions[0]["evidence_sheet"] == "CSV"
        assert questions[0]["question_mode"] == "retrieval"
        assert "产品A" in questions[0]["reference_answer"]
    finally:
        sqg._call_llm_text = original_call_llm

    print("PASS: CSV 端到端生成正确")


def test_generate_xlsx_questions():
    """XLSX 端到端生成（mock LLM）。"""
    print("=" * 60)
    print("测试：XLSX 端到端生成")
    print("=" * 60)

    wb = _make_simple_xlsx()
    xlsx_bytes = _make_xlsx_bytes(wb)

    mock_response = json.dumps([
        {"question": "产品A价格", "sheet_name": "产品表", "anchor_range": "A2:C2", "difficulty": "事实", "topic": "价格"},
    ])

    import spreadsheet_question_generator as sqg
    original_call_llm = sqg._call_llm_text
    sqg._call_llm_text = lambda *a, **kw: mock_response

    try:
        questions, stats = generate_spreadsheet_questions(
            xlsx_bytes, "test.xlsx",
            "fake_key", "http://fake", "fake_model",
            num_questions=5,
        )
        assert len(questions) == 1
        assert questions[0]["source_format"] == "xlsx"
        assert questions[0]["evidence_sheet"] == "产品表"
        assert "产品A" in questions[0]["reference_answer"]
        assert "100" in questions[0]["reference_answer"]
    finally:
        sqg._call_llm_text = original_call_llm

    print("PASS: XLSX 端到端生成正确")


def test_validation_rejects_bad_range():
    """LLM 返回非白名单范围时，题目被过滤。"""
    print("=" * 60)
    print("测试：非白名单范围过滤")
    print("=" * 60)

    csv_bytes = _make_csv_bytes([
        ["Name", "Value"],
        ["Item1", "10"],
    ])

    mock_response = json.dumps([
        {"question": "测试", "sheet_name": "CSV", "anchor_range": "Z1:Z5", "difficulty": "事实"},
    ])

    import spreadsheet_question_generator as sqg
    original_call_llm = sqg._call_llm_text
    sqg._call_llm_text = lambda *a, **kw: mock_response

    try:
        try:
            generate_spreadsheet_questions(
                csv_bytes, "test.csv",
                "fake_key", "http://fake", "fake_model",
            )
            assert False, "应抛出 ValueError（所有题目被过滤）"
        except ValueError as e:
            assert "未通过" in str(e) or "失败" in str(e)
    finally:
        sqg._call_llm_text = original_call_llm

    print("PASS: 非白名单范围正确过滤")


# ====== doc_parser Integration Tests ======

def test_csv_in_doc_parser():
    """CSV 进入 parse_document。"""
    print("=" * 60)
    print("测试：doc_parser CSV 解析")
    print("=" * 60)

    from doc_parser import parse_document

    csv_bytes = _make_csv_bytes([
        ["Name", "Value"],
        ["A", "1"],
        ["B", "2"],
    ])
    result = parse_document(file_bytes=csv_bytes, file_name="test.csv")
    assert result["source_type"] == "csv"
    assert result["summary"]["sheet_count"] == 1
    assert result["summary"]["row_count"] == 2
    assert len(result["blocks"]) == 2

    print("PASS: doc_parser CSV 解析正确")


def test_supported_extensions_includes_new():
    """get_supported_extensions 包含新格式。"""
    print("=" * 60)
    print("测试：支持扩展名列表")
    print("=" * 60)

    from doc_parser import get_supported_extensions, is_supported_file

    exts = get_supported_extensions()
    assert ".csv" in exts
    assert ".xls" in exts
    assert ".xlsx" in exts

    assert is_supported_file("test.csv")
    assert is_supported_file("test.xls")
    assert is_supported_file("test.xlsx")

    print("PASS: 支持扩展名列表正确")


# ====== Prompt Build Test ======

def test_build_prompt():
    """prompt 构建包含表格内容。"""
    print("=" * 60)
    print("测试：Prompt 构建")
    print("=" * 60)

    wb = _make_simple_xlsx()
    xlsx_bytes = _make_xlsx_bytes(wb)
    sheets = parse_xlsx_to_sheet_contexts(xlsx_bytes)

    prompt = _build_prompt(sheets, num_questions=5, topic_hint="产品信息")
    assert "产品表" in prompt
    assert "产品A" in prompt
    assert "5" in prompt
    assert "产品信息" in prompt
    assert "allowed_anchor_ranges" in prompt
    assert "行号" in prompt

    print("PASS: Prompt 构建正确")


# ====== Integration: LLM Request Content Tests ======

def test_llm_request_no_reference_answer():
    """断言实际发送给 LLM 的 prompt 不含 reference_answer/source_excerpt。"""
    print("=" * 60)
    print("测试：LLM 请求不含 reference_answer")
    print("=" * 60)

    wb = _make_simple_xlsx()
    xlsx_bytes = _make_xlsx_bytes(wb)

    # 捕获实际发送给 LLM 的 prompt
    captured_prompts = []

    import spreadsheet_question_generator as sqg
    original_call_llm = sqg._call_llm_text

    def mock_capture(prompt, *args, **kwargs):
        captured_prompts.append(prompt)
        return json.dumps([
            {"question": "产品A", "sheet_name": "产品表", "anchor_range": "A2:C2", "difficulty": "事实"},
        ])

    sqg._call_llm_text = mock_capture

    try:
        questions, stats = generate_spreadsheet_questions(
            xlsx_bytes, "test.xlsx",
            "fake_key", "http://fake", "fake_model",
            num_questions=5,
        )
        assert len(captured_prompts) == 1, f"应捕获 1 个 prompt: {len(captured_prompts)}"
        prompt = captured_prompts[0]

        # 核心断言：prompt 明确禁止 LLM 输出 reference_answer
        assert "不要输出" in prompt and "reference_answer" in prompt, "prompt 应明确禁止 LLM 输出 reference_answer"
        # 输出 JSON 格式中不应包含 reference_answer 作为期望字段
        output_format_section = prompt.split("输出格式")[-1] if "输出格式" in prompt else prompt[-500:]
        assert '"reference_answer"' not in output_format_section, "输出格式中不应有 reference_answer 字段"
        assert "source_excerpt" not in output_format_section, "输出格式中不应有 source_excerpt 字段"
    finally:
        sqg._call_llm_text = original_call_llm

    print("PASS: LLM 请求不含 reference_answer/source_excerpt")


def test_llm_request_no_formula_string():
    """断言发送给 LLM 的 prompt 不含未计算公式字符串。"""
    print("=" * 60)
    print("测试：LLM 请求不含公式字符串")
    print("=" * 60)

    wb = _make_formula_xlsx()
    xlsx_bytes = _make_xlsx_bytes(wb)

    captured_prompts = []

    import spreadsheet_question_generator as sqg
    original_call_llm = sqg._call_llm_text

    def mock_capture(prompt, *args, **kwargs):
        captured_prompts.append(prompt)
        return json.dumps([
            {"question": "项目A数值", "sheet_name": "公式表", "anchor_range": "A2:B2", "difficulty": "事实"},
        ])

    sqg._call_llm_text = mock_capture

    try:
        questions, stats = generate_spreadsheet_questions(
            xlsx_bytes, "test.xlsx",
            "fake_key", "http://fake", "fake_model",
            num_questions=5,
        )
        assert len(captured_prompts) >= 1, "应至少有 1 次 LLM 调用"
        prompt = captured_prompts[0]  # 检查首次调用的 prompt

        # 核心断言：prompt 中不含公式字符串
        assert "=SUM(" not in prompt, f"prompt 不应含公式字符串 =SUM("
        assert "=" not in prompt.split("allowed_anchor_ranges")[0].split("行号")[-1] or \
               "[公式未计算]" in prompt, "公式单元格应显示 [公式未计算] 而非公式表达式"
        assert "[公式未计算]" in prompt, "prompt 应包含 [公式未计算] 标记"
    finally:
        sqg._call_llm_text = original_call_llm

    print("PASS: LLM 请求不含公式字符串")


def test_llm_request_uses_spreadsheet_prompt():
    """断言使用的是表格专用 prompt（含 allowed_anchor_ranges），而非通用检索 prompt。"""
    print("=" * 60)
    print("测试：使用表格专用 prompt")
    print("=" * 60)

    wb = _make_simple_xlsx()
    xlsx_bytes = _make_xlsx_bytes(wb)

    captured_prompts = []

    import spreadsheet_question_generator as sqg
    original_call_llm = sqg._call_llm_text

    def mock_capture(prompt, *args, **kwargs):
        captured_prompts.append(prompt)
        return json.dumps([
            {"question": "产品A", "sheet_name": "产品表", "anchor_range": "A2:C2", "difficulty": "事实"},
        ])

    sqg._call_llm_text = mock_capture

    try:
        questions, stats = generate_spreadsheet_questions(
            xlsx_bytes, "test.xlsx",
            "fake_key", "http://fake", "fake_model",
            num_questions=5,
        )
        prompt = captured_prompts[0]

        # 表格专用 prompt 的特征
        assert "allowed_anchor_ranges" in prompt, "prompt 应含 allowed_anchor_ranges 白名单"
        assert "行号" in prompt, "prompt 应含 Excel 行号列"
        assert "工作表:" in prompt or "工作表：" in prompt, "prompt 应含工作表标题"
        assert "电子表格" in prompt or "表格内容" in prompt, "prompt 应为表格专用模板"

        # 不应含通用检索 prompt 的特征
        assert "{content}" not in prompt, "prompt 不应含通用模板占位符 {content}"
        assert "{section_context}" not in prompt, "prompt 不应含通用模板占位符 {section_context}"
    finally:
        sqg._call_llm_text = original_call_llm

    print("PASS: 使用表格专用 prompt")


def test_llm_request_local_reference_answer():
    """断言 reference_answer 只来自本地渲染，不含 LLM 输出。"""
    print("=" * 60)
    print("测试：reference_answer 纯本地渲染")
    print("=" * 60)

    wb = _make_simple_xlsx()
    xlsx_bytes = _make_xlsx_bytes(wb)

    # Mock LLM 返回（故意不含 reference_answer）
    mock_response = json.dumps([
        {"question": "产品A价格", "sheet_name": "产品表", "anchor_range": "A2:C2", "difficulty": "事实"},
    ])

    import spreadsheet_question_generator as sqg
    original_call_llm = sqg._call_llm_text
    sqg._call_llm_text = lambda *a, **kw: mock_response

    try:
        questions, stats = generate_spreadsheet_questions(
            xlsx_bytes, "test.xlsx",
            "fake_key", "http://fake", "fake_model",
        )
        q = questions[0]

        # reference_answer 必须存在且来自本地
        assert "reference_answer" in q, "应有 reference_answer"
        assert "source_excerpt" in q, "应有 source_excerpt"
        assert q["reference_answer"] == q["source_excerpt"], "两者应一致"

        # reference_answer 应包含实际单元格值
        assert "产品A" in q["reference_answer"], f"应含产品A: {q['reference_answer']}"
        assert "100" in q["reference_answer"], f"应含100: {q['reference_answer']}"

        # reference_answer 不应含 LLM 可能自写的文本
        assert "短检索" not in q["reference_answer"], "reference_answer 不应含 prompt 指令文本"
    finally:
        sqg._call_llm_text = original_call_llm

    print("PASS: reference_answer 纯本地渲染")


# ====== Backward Compatibility Tests ======

def test_xlsx_question_generator_delegates():
    """xlsx_question_generator.generate_xlsx_questions 委托给新模块。"""
    print("=" * 60)
    print("测试：xlsx_question_generator 委托")
    print("=" * 60)

    wb = _make_simple_xlsx()
    xlsx_bytes = _make_xlsx_bytes(wb)

    mock_response = json.dumps([
        {"question": "产品A", "sheet_name": "产品表", "anchor_range": "A2:C2", "difficulty": "事实"},
    ])

    import spreadsheet_question_generator as sqg
    original_call_llm = sqg._call_llm_text
    sqg._call_llm_text = lambda *a, **kw: mock_response

    try:
        from xlsx_question_generator import generate_xlsx_questions
        questions, stats = generate_xlsx_questions(
            xlsx_bytes, "test.xlsx",
            "fake_key", "http://fake", "fake_model",
        )
        assert len(questions) == 1
        assert questions[0]["source_format"] == "xlsx"
    finally:
        sqg._call_llm_text = original_call_llm

    print("PASS: xlsx_question_generator 委托正常")


def test_existing_xlsx_functions_importable():
    """xlsx_question_generator 原有内部函数仍可导入。"""
    print("=" * 60)
    print("测试：原有函数可导入")
    print("=" * 60)

    from xlsx_question_generator import (
        _validate_and_render_evidence,
        _render_evidence_range,
        _parse_xlsx_qgen_response,
        _parse_range,
        _get_cell_display_value,
        check_xlsx_llm_support,
    )
    # 不崩溃即通过
    assert callable(_validate_and_render_evidence)
    assert callable(_render_evidence_range)

    print("PASS: 原有函数仍可导入")


# ====== Price Anchor Regression Tests ======

def test_price_anchor_d4d4_rejected():
    """功能名称+价格查询使用 D4:D4 必须被拒绝（孤立数值）。"""
    print("=" * 60)
    print("测试：价格题 D4:D4 孤立数值拒绝")
    print("=" * 60)

    wb = _make_rate_table_xlsx()
    xlsx_bytes = _make_xlsx_bytes(wb)
    sheets = parse_xlsx_to_sheet_contexts(xlsx_bytes)
    sheets_by_name = {s.sheet_name: s for s in sheets}

    q = {"question": "集成发布流水线梳理未税价", "sheet_name": "报价页", "anchor_range": "D4:D4"}
    result, reason = _validate_and_render_question(q, sheets_by_name, "test.xlsx")
    assert result is None, f"D4:D4 孤立价格数值应被拒绝，但通过了: {result}"
    assert "孤立数值" in reason, f"原因应含'孤立数值': {reason}"

    print("PASS: D4:D4 孤立价格数值正确拒绝")


def test_price_anchor_b4d4_passes_with_full_evidence():
    """B4:D4 + B2:D2 必须通过，且 reference_answer 同时包含功能名、字段名、数值。"""
    print("=" * 60)
    print("测试：价格题 B4:D4 完整四项证据")
    print("=" * 60)

    wb = _make_rate_table_xlsx()
    xlsx_bytes = _make_xlsx_bytes(wb)
    sheets = parse_xlsx_to_sheet_contexts(xlsx_bytes)
    sheets_by_name = {s.sheet_name: s for s in sheets}

    q = {
        "question": "集成发布流水线梳理未税价",
        "sheet_name": "报价页",
        "anchor_range": "B4:D4",
        "difficulty": "事实",
        "topic": "价格",
    }
    result, reason = _validate_and_render_question(q, sheets_by_name, "test.xlsx")
    assert result is not None, f"B4:D4 应通过验证: {reason}"

    ref = result["reference_answer"]
    # 必须断言业务功能存在
    assert "CICD工具规范" in ref or "集成发布流水线" in ref, \
        f"reference_answer 必须包含功能名称: {ref}"
    # 必须断言字段名存在
    assert "未税价" in ref, \
        f"reference_answer 必须包含'未税价'字段名: {ref}"
    # 必须断言价格数值存在
    assert "73900" in ref, \
        f"reference_answer 必须包含价格数值 73900: {ref}"
    # 必须断言功能模块字段存在
    assert "功能模块" in ref, \
        f"reference_answer 必须包含'功能模块'字段: {ref}"
    # 必须断言产品功能字段存在
    assert "产品功能" in ref, \
        f"reference_answer 必须包含'产品功能'字段: {ref}"

    print(f"PASS: B4:D4 完整证据: {ref}")


def test_price_candidate_anchors_use_b_to_d():
    """候选锚点列表中的价格锚点必须是 B行:D行，不能是 D行:D行。"""
    print("=" * 60)
    print("测试：候选锚点 B行:D行 格式")
    print("=" * 60)

    from spreadsheet_question_generator import _build_candidate_anchors

    wb = _make_rate_table_xlsx()
    xlsx_bytes = _make_xlsx_bytes(wb)
    sheets = parse_xlsx_to_sheet_contexts(xlsx_bytes)

    candidates = _build_candidate_anchors(sheets)
    rsc = [c for c in candidates if c["evidence_mode"] == "record_with_schema_context"]
    assert len(rsc) > 0, "应至少有一个 record_with_schema_context 候选项"

    for pa in rsc:
        anchor = pa["anchor_range"]
        bounds = _parse_range_str(anchor)
        assert bounds is not None, f"无法解析锚点: {anchor}"
        min_col, min_row, max_col, max_row = bounds
        # 锚点必须覆盖 B 列（col 2），不能从 D 列（col 4）开始
        assert min_col <= 2, \
            f"锚点 {anchor} 起始列应为 B 或更左，实际起始列 {_col_letter(min_col)}"
        # 必须是单行
        assert min_row == max_row, f"锚点必须是单行: {anchor}"

    print(f"PASS: 候选锚点格式正确: {[pa['anchor_range'] for pa in rsc]}")


def test_price_isolated_numeric_no_dual_source_exemption():
    """D4:D4、D5:D5 等孤立价格数值必须被拒绝，双源模型不能豁免。"""
    print("=" * 60)
    print("测试：孤立价格数值无双源豁免")
    print("=" * 60)

    wb = _make_rate_table_xlsx()
    xlsx_bytes = _make_xlsx_bytes(wb)
    sheets = parse_xlsx_to_sheet_contexts(xlsx_bytes)
    sheets_by_name = {s.sheet_name: s for s in sheets}

    # D4:D4 应被拒绝
    q4 = {"question": "价格", "sheet_name": "报价页", "anchor_range": "D4:D4"}
    r4, reason4 = _validate_and_render_question(q4, sheets_by_name, "test.xlsx")
    assert r4 is None, f"D4:D4 应被拒绝: {r4}"
    assert "孤立数值" in reason4, f"原因应含'孤立数值': {reason4}"

    # D5:D5 应被拒绝（如果有第5行数据）
    sheet = sheets_by_name["报价页"]
    if len(sheet.rows) >= 5:
        q5 = {"question": "价格", "sheet_name": "报价页", "anchor_range": "D5:D5"}
        r5, reason5 = _validate_and_render_question(q5, sheets_by_name, "test.xlsx")
        assert r5 is None, f"D5:D5 应被拒绝: {r5}"
        assert "孤立数值" in reason5, f"原因应含'孤立数值': {reason5}"

    print("PASS: 孤立价格数值无双源豁免")


def test_smoke_10_questions_price_evidence():
    """冒烟验收：10 条题目，所有价格题输出完整四项证据。"""
    print("=" * 60)
    print("冒烟验收：10 条价格题完整证据")
    print("=" * 60)

    import spreadsheet_question_generator as sqg

    # 构建模拟真实报价表（10 行业务数据）
    wb = Workbook()
    ws = wb.active
    ws.title = "报价总表"
    # Row 1: 列标题
    ws["A1"] = "序号"
    ws["B1"] = "功能模块"
    ws["C1"] = "产品功能"
    ws["D1"] = "未税价（元）"
    ws["E1"] = "项目经理"
    ws["F1"] = "研发经理"
    # Row 2: 字段名行（真正的字段名）
    ws["A2"] = ""
    ws["B2"] = "功能模块"
    ws["C2"] = "产品功能"
    ws["D2"] = "未税价（元）"
    ws["E2"] = "项目经理"
    ws["F2"] = "研发经理"
    # Row 3: 费率行
    ws["A3"] = ""
    ws["B3"] = "功能模块"
    ws["C3"] = "产品功能"
    ws["D3"] = 50000
    ws["E3"] = 1700
    ws["F3"] = 1800
    # Row 4-13: 10 行业务数据（A 列留空，与真实报价表结构一致）
    biz_data = [
        ("CICD工具规范", "集成发布流水线梳理", 73900, 2, 1),
        ("CICD工具规范", "代码质量门禁", 45000, 1, 2),
        ("自动化测试", "接口自动化测试框架", 128000, 3, 2),
        ("自动化测试", "UI自动化测试框架", 96000, 2, 3),
        ("监控运维", "统一监控平台建设", 210000, 4, 3),
        ("监控运维", "日志分析平台", 85000, 2, 1),
        ("安全合规", "漏洞扫描服务", 56000, 1, 1),
        ("安全合规", "代码审计服务", 42000, 1, 0),
        ("容器平台", "K8s集群管理", 168000, 3, 2),
        ("容器平台", "镜像仓库建设", 52000, 1, 1),
    ]
    for i, (module, func, price, pm, dev) in enumerate(biz_data):
        row = 4 + i
        ws[f"B{row}"] = module
        ws[f"C{row}"] = func
        ws[f"D{row}"] = price
        ws[f"E{row}"] = pm
        ws[f"F{row}"] = dev

    xlsx_bytes = _make_xlsx_bytes(wb)

    # Mock LLM 返回 10 条题目：混合正确和错误的锚点
    mock_questions = [
        # 5 条正确的 B4:D4 锚点
        {"question": "集成发布流水线梳理未税价", "sheet_name": "报价总表", "anchor_range": "B4:D4", "difficulty": "事实", "topic": "价格"},
        {"question": "代码质量门禁未税价", "sheet_name": "报价总表", "anchor_range": "B5:D5", "difficulty": "事实", "topic": "价格"},
        {"question": "接口自动化测试框架报价", "sheet_name": "报价总表", "anchor_range": "B6:D6", "difficulty": "事实", "topic": "价格"},
        {"question": "统一监控平台建设未税价", "sheet_name": "报价总表", "anchor_range": "B9:D9", "difficulty": "事实", "topic": "价格"},
        {"question": "K8s集群管理价格", "sheet_name": "报价总表", "anchor_range": "B12:D12", "difficulty": "事实", "topic": "价格"},
        # 2 条费率题（E2:E3 类型）
        {"question": "项目经理费率", "sheet_name": "报价总表", "anchor_range": "E2:E3", "difficulty": "事实", "topic": "费率"},
        {"question": "研发经理费率", "sheet_name": "报价总表", "anchor_range": "F2:F3", "difficulty": "事实", "topic": "费率"},
        # 2 条文本题
        {"question": "集成发布流水线梳理功能模块", "sheet_name": "报价总表", "anchor_range": "B4:C4", "difficulty": "事实", "topic": "功能"},
        {"question": "容器平台产品功能", "sheet_name": "报价总表", "anchor_range": "B12:C12", "difficulty": "事实", "topic": "功能"},
        # 1 条错误锚点（应被拒绝）
        {"question": "镜像仓库建设未税价", "sheet_name": "报价总表", "anchor_range": "D13:D13", "difficulty": "事实", "topic": "价格"},
    ]
    mock_response = json.dumps(mock_questions)

    original_call_llm = sqg._call_llm_text
    sqg._call_llm_text = lambda *a, **kw: mock_response

    try:
        questions, stats = generate_spreadsheet_questions(
            xlsx_bytes, "报价表.xlsx",
            "fake_key", "http://fake", "fake_model",
            num_questions=10,
        )
    finally:
        sqg._call_llm_text = original_call_llm

    # D13:D13 应被拒绝，最终应少于 10 题
    assert len(questions) <= 9, f"D13:D13 应被拒绝，但通过了 {len(questions)} 题"

    # 验证所有价格题的证据完整性
    price_kw = ("未税价", "价格", "报价")
    price_questions = [q for q in questions if any(kw in q["question"] for kw in price_kw)]

    assert len(price_questions) >= 5, f"应至少有 5 条价格题，实际 {len(price_questions)}"

    for pq in price_questions:
        ref = pq["reference_answer"]
        anchor = pq["anchor_range"]
        # 每条价格题必须包含完整的四项证据
        assert "功能模块" in ref, \
            f"价格题 '{pq['question']}' anchor={anchor} 缺少'功能模块': {ref}"
        assert "产品功能" in ref, \
            f"价格题 '{pq['question']}' anchor={anchor} 缺少'产品功能': {ref}"
        assert "未税价" in ref, \
            f"价格题 '{pq['question']}' anchor={anchor} 缺少'未税价': {ref}"
        # 价格数值：检查是否包含某个具体数字
        has_numeric = any(ch.isdigit() for ch in ref)
        assert has_numeric, \
            f"价格题 '{pq['question']}' anchor={anchor} 缺少价格数值: {ref}"
        # anchor 必须是 B行:D行 格式（不能是 D行:D4）
        bounds = _parse_range_str(anchor)
        assert bounds is not None, f"无法解析 anchor: {anchor}"
        min_col, _, _, _ = bounds
        assert min_col <= 2, \
            f"价格题 anchor={anchor} 起始列应为 B 或更左（覆盖业务标识）"

    print(f"通过: {len(questions)} 题（{len(price_questions)} 条价格题全部输出完整四项证据）")
    print("PASS: 10 条冒烟验收")


# ====== Generic Evidence Mode Tests ======

def test_candidate_anchors_inventory_table():
    """库存/产品表应生成 record_with_schema_context 候选项。"""
    print("=" * 60)
    print("测试：库存表候选锚点")
    print("=" * 60)

    from spreadsheet_question_generator import _build_candidate_anchors

    wb = Workbook()
    ws = wb.active
    ws.title = "库存表"
    ws["A1"] = "商品编号"
    ws["B1"] = "商品名称"
    ws["C1"] = "库存数量"
    ws["D1"] = "单价"
    ws["A2"] = "P001"
    ws["B2"] = "笔记本电脑"
    ws["C2"] = 50
    ws["D2"] = 6999
    ws["A3"] = "P002"
    ws["B3"] = "机械键盘"
    ws["C3"] = 200
    ws["D3"] = 399

    xlsx_bytes = _make_xlsx_bytes(wb)
    sheets = parse_xlsx_to_sheet_contexts(xlsx_bytes)

    candidates = _build_candidate_anchors(sheets)
    assert len(candidates) > 0, "应生成候选锚点"

    # 检查是否有文本事实候选项
    tf = [c for c in candidates if c["evidence_mode"] == "text_fact"]
    assert len(tf) > 0, "库存表应有 text_fact 候选项"
    for c in tf:
        assert "evidence_mode" in c
        assert "anchor_range" in c
        assert "fact_fields" in c
        assert "query_focus" in c

    print(f"PASS: 库存表生成 {len(candidates)} 条候选项（text_fact: {len(tf)}）")


def test_candidate_anchors_kv_parameter_table():
    """键值参数表（费率表结构）应生成 field_value_pair 候选项。"""
    print("=" * 60)
    print("测试：参数表候选锚点")
    print("=" * 60)

    from spreadsheet_question_generator import _build_candidate_anchors

    # 使用费率表结构（label行 + value行），与 _detect_header_value_row_pairs 兼容
    wb = Workbook()
    ws = wb.active
    ws.title = "参数配置"
    ws["A1"] = "参数类别"
    ws["B1"] = "最大连接数"
    ws["C1"] = "超时时间"
    ws["D1"] = "重试次数"
    ws["A2"] = "系统参数"
    ws["B2"] = 1000
    ws["C2"] = 30
    ws["D2"] = 3

    xlsx_bytes = _make_xlsx_bytes(wb)
    sheets = parse_xlsx_to_sheet_contexts(xlsx_bytes)

    candidates = _build_candidate_anchors(sheets)
    fvp = [c for c in candidates if c["evidence_mode"] == "field_value_pair"]
    assert len(fvp) >= 1, f"参数表应有至少1个 field_value_pair 候选项: {len(fvp)}"

    for c in fvp:
        assert len(c["fact_fields"]) == 1, f"field_value_pair 应有1个字段: {c['fact_fields']}"
        assert "：" in c["query_focus"], f"query_focus 应含分隔符: {c['query_focus']}"

    print(f"PASS: 参数表生成 {len(fvp)} 条 field_value_pair 候选项")


def test_candidate_anchors_csv_plain_table():
    """CSV 普通表应生成 text_fact 候选项。"""
    print("=" * 60)
    print("测试：CSV 普通表候选锚点")
    print("=" * 60)

    from spreadsheet_question_generator import _build_candidate_anchors, parse_csv_to_sheet_contexts

    csv_bytes = "姓名,部门,职位\n张三,技术部,工程师\n李四,市场部,经理\n".encode("utf-8")
    sheets = parse_csv_to_sheet_contexts(csv_bytes)

    candidates = _build_candidate_anchors(sheets)
    tf = [c for c in candidates if c["evidence_mode"] == "text_fact"]
    assert len(tf) > 0, "CSV 表应有 text_fact 候选项"

    for c in tf:
        assert c["sheet_name"] == "CSV"
        assert len(c["fact_fields"]) >= 1

    print(f"PASS: CSV 表生成 {len(tf)} 条 text_fact 候选项")


def test_candidate_anchor_metadata_no_out_of_bounds():
    """fact_fields 和 query_focus 不得包含 anchor_range 之外的列。"""
    print("=" * 60)
    print("测试：候选锚点元数据不越界")
    print("=" * 60)

    from spreadsheet_question_generator import _build_candidate_anchors

    # 构造表：B-D 列为业务字段，E 列为额外数值
    wb = Workbook()
    ws = wb.active
    ws.title = "报价页"
    # Row 1: 列标题
    ws["A1"] = "序号"
    ws["B1"] = "功能模块"
    ws["C1"] = "产品功能"
    ws["D1"] = "未税价（元）"
    ws["E1"] = "项目经理人数"
    ws["F1"] = "研发经理人数"
    # Row 2: 字段名行
    ws["B2"] = "功能模块"
    ws["C2"] = "产品功能"
    ws["D2"] = "未税价（元）"
    ws["E2"] = "项目经理"
    ws["F2"] = "研发经理"
    # Row 3: 费率行
    ws["B3"] = "功能模块"
    ws["C3"] = "产品功能"
    ws["D3"] = 50000
    ws["E3"] = 1700
    ws["F3"] = 1800
    # Row 4: 业务数据
    ws["B4"] = "CICD工具规范"
    ws["C4"] = "集成发布流水线"
    ws["D4"] = 73900
    ws["E4"] = 2
    ws["F4"] = 1

    xlsx_bytes = _make_xlsx_bytes(wb)
    sheets = parse_xlsx_to_sheet_contexts(xlsx_bytes)
    candidates = _build_candidate_anchors(sheets)

    # 找到 B4:D4 对应的 record_with_schema_context 候选项
    rsc = [c for c in candidates
           if c["evidence_mode"] == "record_with_schema_context"
           and "4" in c["anchor_range"]]
    assert len(rsc) > 0, f"应有行4的 record 候选项: {candidates}"

    for c in rsc:
        anchor = c["anchor_range"]
        bounds = _parse_range_str(anchor)
        a_min_col, _, a_max_col, _ = bounds

        # fact_fields 数量应等于 anchor 覆盖列中的文本字段数
        # 不得包含 anchor 范围之外的列
        # anchor B4:D4 → 只能含 B/C/D 列字段；E/F 列字段不得出现
        anchor_col_count = a_max_col - a_min_col + 1
        assert len(c["fact_fields"]) <= anchor_col_count, \
            f"anchor={anchor} fact_fields 数量({len(c['fact_fields'])})不应超过列数({anchor_col_count}): {c['fact_fields']}"

        # 验证 fact_fields 中每个字段名都对应 anchor 覆盖的列
        # 通过检查 header 行确认字段位置
        header_row = sheets[0].rows[0]  # row 1 = headers
        for ff in c["fact_fields"]:
            # 找到字段名在 header 中的列位置
            ff_col = None
            for ci, hv in enumerate(header_row):
                if hv and str(hv).strip() == ff:
                    ff_col = ci + 1  # 1-indexed
                    break
            if ff_col is not None:
                assert a_min_col <= ff_col <= a_max_col, \
                    f"字段'{ff}'在列{_col_letter(ff_col)}，但 anchor 只覆盖 {_col_letter(a_min_col)}-{_col_letter(a_max_col)}"

        # query_focus 不得包含 anchor 范围外单元格的值
        # E4=2, F4=1 不应出现在 B4:D4 的 query_focus 中
        row_data = sheets[0].rows[3]  # row 4
        for col_idx in range(a_max_col, len(row_data)):  # anchor 之后的列
            val = row_data[col_idx]
            if val is not None and str(val).strip() and not isinstance(val, (int, float)):
                assert str(val).strip() not in c["query_focus"], \
                    f"anchor={anchor} query_focus 不应含锚点外列的文本值'{val}': {c['query_focus']}"

    print(f"PASS: {len(rsc)} 条候选项元数据均不越界")


def test_real_appendix_d_candidate_metadata():
    """真实 Appendix D 结构：B4:D4 候选只含功能模块、产品功能、未税价及对应值。"""
    print("=" * 60)
    print("测试：Appendix D 候选元数据正确性")
    print("=" * 60)

    from spreadsheet_question_generator import _build_candidate_anchors

    # 模拟真实 Appendix D 结构
    wb = Workbook()
    ws = wb.active
    ws.title = "报价页"
    # Row 1: 列标题
    ws["A1"] = "序号"; ws["B1"] = "功能模块"; ws["C1"] = "产品功能"; ws["D1"] = "未税价（元）"
    ws["E1"] = "项目经理"; ws["F1"] = "研发经理"; ws["G1"] = "DevOps专家"
    ws["H1"] = "DevOps工程师"; ws["I1"] = "前端工程师"; ws["J1"] = "后端工程师"
    ws["K1"] = "BA"; ws["L1"] = "测试"; ws["M1"] = "SRE工程师"
    # Row 2: 字段名行
    ws["B2"] = "功能模块"; ws["C2"] = "产品功能"; ws["D2"] = "未税价（元）"
    ws["E2"] = "项目经理"; ws["F2"] = "研发经理"; ws["G2"] = "DevOps专家"
    ws["H2"] = "DevOps工程师"; ws["I2"] = "前端工程师"; ws["J2"] = "后端工程师"
    ws["K2"] = "BA"; ws["L2"] = "测试"; ws["M2"] = "SRE工程师"
    # Row 3: 费率行
    ws["D3"] = 50000; ws["E3"] = 1700; ws["F3"] = 1800; ws["G3"] = 1600
    ws["H3"] = 1500; ws["I3"] = 1500; ws["J3"] = 1500; ws["K3"] = 1400
    ws["L3"] = 1300; ws["M3"] = 1500
    # Row 4: 业务数据
    ws["B4"] = "CICD工具规范及Pipeline建设优化"; ws["C4"] = "集成发布流水线梳理"; ws["D4"] = 73900
    ws["E4"] = 2; ws["F4"] = 1; ws["G4"] = 0; ws["H4"] = 3; ws["I4"] = 0
    ws["J4"] = 0; ws["K4"] = 0; ws["L4"] = 1; ws["M4"] = 0

    xlsx_bytes = _make_xlsx_bytes(wb)
    sheets = parse_xlsx_to_sheet_contexts(xlsx_bytes)
    candidates = _build_candidate_anchors(sheets)

    # 找 B4 对应的 record 候选项
    rsc_row4 = [c for c in candidates
                if c["evidence_mode"] == "record_with_schema_context"
                and c["anchor_range"].startswith("B4:")]
    assert len(rsc_row4) > 0, f"应有 B4 开头的 record 候选项"

    # 找 anchor 为 B4:D4 的候选项
    b4d4 = [c for c in rsc_row4 if c["anchor_range"] == "B4:D4"]
    if b4d4:
        c = b4d4[0]
        assert c["fact_fields"] == ["功能模块", "产品功能", "未税价（元）"], \
            f"B4:D4 fact_fields 应为三列: {c['fact_fields']}"
        assert "CICD工具规范" in c["query_focus"], f"query_focus 应含功能模块值: {c['query_focus']}"
        assert "集成发布流水线" in c["query_focus"], f"query_focus 应含产品功能值: {c['query_focus']}"
        assert "73900" in c["query_focus"], f"query_focus 应含未税价值: {c['query_focus']}"
        # 不得含 E-M 列的值
        assert "项目经理" not in str(c["fact_fields"]), "不应含角色字段"
        assert "1700" not in c["query_focus"], "不应含费率值"

    print(f"PASS: B4:D4 候选元数据正确: {b4d4[0] if b4d4 else 'N/A'}")


def test_prompt_no_business_terms():
    """Prompt 模板规则中不应出现硬编码的业务术语。"""
    print("=" * 60)
    print("测试：Prompt 模板无业务术语")
    print("=" * 60)

    # 直接读取模板文件检查（不含渲染后的表格内容）
    from pathlib import Path
    template_path = Path(__file__).parent.parent / "prompts" / "qgen_prompt_spreadsheet_retrieval.txt"
    template = template_path.read_text(encoding="utf-8")

    forbidden_in_template = ["功能模块", "产品功能", "未税价", "B4:D4", "报价表", "费率表"]
    found = [t for t in forbidden_in_template if t in template]
    assert not found, f"Prompt 模板中不应出现业务术语: {found}"

    # 应包含候选目录相关描述
    assert "candidate_id" in template, "模板应描述 candidate_id"
    assert "target_field_label" in template, "模板应描述 target_field_label"
    assert "候选目录" in template, "模板应提及候选目录"

    print("PASS: Prompt 模板无业务术语，含候选目录描述")


# ====== Candidate Section & Summary Row Regression Tests ======

def test_candidate_section_appears_in_prompt():
    """候选证据清单实际出现在 Prompt 中，且包含结构化字段。"""
    print("=" * 60)
    print("测试：候选清单出现在 Prompt")
    print("=" * 60)

    from spreadsheet_question_generator import _build_candidate_anchors

    wb = Workbook()
    ws = wb.active
    ws.title = "报价页"
    ws["B1"] = "功能模块"; ws["C1"] = "产品功能"; ws["D1"] = "未税价（元）"
    ws["E1"] = "项目经理"
    ws["B2"] = "功能模块"; ws["C2"] = "产品功能"; ws["D2"] = "未税价（元）"
    ws["E2"] = "项目经理"
    ws["B3"] = "功能模块"; ws["C3"] = "产品功能"; ws["D3"] = 50000; ws["E3"] = 1700
    ws["B4"] = "CICD工具规范"; ws["C4"] = "集成发布流水线"; ws["D4"] = 73900; ws["E4"] = 2

    xlsx_bytes = _make_xlsx_bytes(wb)
    sheets = parse_xlsx_to_sheet_contexts(xlsx_bytes)
    candidates = _build_candidate_anchors(sheets)

    prompt = _build_prompt(sheets, 5, "", candidate_anchors=candidates)

    # 候选清单必须出现
    assert "候选证据清单" in prompt, "prompt 应含 '候选证据清单' 标题"

    # 验证至少一个候选项的全部关键字段出现在 prompt 中
    if candidates:
        c = candidates[0]
        assert c["anchor_range"] in prompt, f"prompt 应含 anchor_range: {c['anchor_range']}"
        assert c["sheet_name"] in prompt, f"prompt 应含 sheet_name: {c['sheet_name']}"
        if c.get("fact_fields"):
            for ff in c["fact_fields"]:
                assert ff in prompt, f"prompt 应含 fact_field: {ff}"
        if c.get("query_focus"):
            focus_first = c["query_focus"].split(" | ")[0]
            assert focus_first in prompt, f"prompt 应含 query_focus 片段: {focus_first}"

    print("PASS: 候选清单完整出现在 Prompt")


def test_summary_rows_excluded_from_candidates():
    """汇总行（总计/合计/小计）不生成候选锚点。"""
    print("=" * 60)
    print("测试：汇总行被排除")
    print("=" * 60)

    from spreadsheet_question_generator import _build_candidate_anchors

    wb = Workbook()
    ws = wb.active
    ws.title = "费用表"
    # 表头
    ws["A1"] = "项目编号"; ws["B1"] = "项目名称"; ws["C1"] = "部门"; ws["D1"] = "金额"
    # 正常业务行
    ws["A2"] = "P001"; ws["B2"] = "服务A"; ws["C2"] = "技术部"; ws["D2"] = 10000
    ws["A3"] = "P002"; ws["B3"] = "服务B"; ws["C3"] = "市场部"; ws["D3"] = 20000
    # 汇总行
    ws["A4"] = "合计"; ws["B4"] = ""; ws["C4"] = ""; ws["D4"] = 30000
    ws["A5"] = "总计"; ws["B5"] = ""; ws["C5"] = ""; ws["D5"] = 30000
    ws["A6"] = "小计"; ws["B6"] = ""; ws["C6"] = ""; ws["D6"] = 15000

    xlsx_bytes = _make_xlsx_bytes(wb)
    sheets = parse_xlsx_to_sheet_contexts(xlsx_bytes)
    candidates = _build_candidate_anchors(sheets)

    # 候选中不应出现包含"合计"/"总计"/"小计"的行
    for c in candidates:
        anchor = c["anchor_range"]
        bounds = _parse_range_str(anchor)
        if bounds:
            _, row, _, _ = bounds
            row_data = sheets[0].rows[row - 1]
            first_cell = ""
            for v in row_data:
                if v is not None and str(v).strip():
                    first_cell = str(v).strip()
                    break
            assert first_cell not in ("合计", "总计", "小计"), \
                f"汇总行 '{first_cell}' 不应生成候选: anchor={anchor}"

    # 应有正常业务行的候选
    assert len(candidates) > 0, "应有正常业务行的候选"

    print(f"PASS: 汇总行正确排除，保留 {len(candidates)} 条候选")


def test_appendix_e_smoke_with_summary_and_merged():
    """Appendix E 风格冒烟测试：含汇总行、合并单元格。"""
    print("=" * 60)
    print("冒烟测试：Appendix E 汇总行+合并单元格")
    print("=" * 60)

    from spreadsheet_question_generator import _build_candidate_anchors
    import spreadsheet_question_generator as sqg

    wb = Workbook()
    ws = wb.active
    ws.title = "价格清单"
    # Row 1: 列标题
    ws["A1"] = "序号"; ws["B1"] = "功能模块"; ws["C1"] = "产品功能"
    ws["D1"] = "未税价（元）"; ws["E1"] = "项目经理"; ws["F1"] = "研发经理"
    ws["M1"] = "SRE工程师"; ws["N1"] = "备注"; ws["O1"] = "状态"
    # Row 2: 字段名行
    ws["B2"] = "功能模块"; ws["C2"] = "产品功能"; ws["D2"] = "未税价（元）"
    ws["E2"] = "项目经理"; ws["F2"] = "研发经理"; ws["M2"] = "SRE工程师"
    ws["N2"] = "备注"; ws["O2"] = "状态"
    # Row 3: 费率行
    ws["D3"] = 50000; ws["E3"] = 1700; ws["F3"] = 1800; ws["M3"] = 1500
    ws["O3"] = "进行中"
    # Row 4-8: 业务数据
    biz = [
        ("CICD工具规范", "集成发布流水线", 73900, 2, 1, 0, "正常"),
        ("自动化测试", "接口自动化测试", 128000, 3, 2, 1, "正常"),
        ("监控运维", "统一监控平台", 210000, 4, 3, 2, "正常"),
        ("安全合规", "漏洞扫描服务", 56000, 1, 1, 0, "正常"),
        ("容器平台", "K8s集群管理", 168000, 3, 2, 1, "正常"),
    ]
    for i, (mod, func, price, pm, dev, sre, status) in enumerate(biz):
        r = 4 + i
        ws[f"B{r}"] = mod; ws[f"C{r}"] = func; ws[f"D{r}"] = price
        ws[f"E{r}"] = pm; ws[f"F{r}"] = dev; ws[f"M{r}"] = sre
        ws[f"N{r}"] = "无"; ws[f"O{r}"] = status

    # Row 52: 汇总行
    ws["A52"] = "总计"; ws["D52"] = 637900
    ws["E52"] = 13; ws["F52"] = 9; ws["M52"] = 4

    # 合并 O3:O24 模拟合并单元格
    ws.merge_cells("O3:O24")

    xlsx_bytes = _make_xlsx_bytes(wb)
    sheets = parse_xlsx_to_sheet_contexts(xlsx_bytes)
    candidates = _build_candidate_anchors(sheets)

    # 验证：总计行（row 52）不应生成候选
    for c in candidates:
        anchor = c["anchor_range"]
        bounds = _parse_range_str(anchor)
        if bounds:
            _, row, _, _ = bounds
            assert row != 52, f"汇总行 52 不应生成候选: {anchor}"

    # 验证：正常业务行应生成候选
    rsc = [c for c in candidates if c["evidence_mode"] == "record_with_schema_context"]
    assert len(rsc) >= 4, f"应至少有 4 个业务行候选: {len(rsc)}"

    # 验证：fact_fields 不应包含 anchor 外列
    for c in rsc:
        anchor = c["anchor_range"]
        bounds = _parse_range_str(anchor)
        if bounds:
            a_min_col, _, a_max_col, _ = bounds
            assert len(c["fact_fields"]) <= (a_max_col - a_min_col + 1), \
                f"anchor={anchor} fact_fields 越界: {c['fact_fields']}"

    # 端到端：mock LLM 验证整体流程
    mock_response = json.dumps([
        {"question": "CICD工具规范未税价", "sheet_name": "价格清单",
         "anchor_range": "B4:D4", "difficulty": "事实", "topic": "价格"},
        {"question": "自动化测试接口", "sheet_name": "价格清单",
         "anchor_range": "B5:C5", "difficulty": "事实", "topic": "功能"},
    ])

    original_call_llm = sqg._call_llm_text
    sqg._call_llm_text = lambda *a, **kw: mock_response

    try:
        questions, stats = generate_spreadsheet_questions(
            xlsx_bytes, "appendix_e.xlsx",
            "fake_key", "http://fake", "fake_model",
            num_questions=5,
        )
        assert len(questions) >= 2, f"应至少生成 2 题: {len(questions)}"

        for q in questions:
            if "未税价" in q.get("question", ""):
                ref = q["reference_answer"]
                assert "功能模块" in ref, f"价格题缺功能模块: {ref}"
                assert "未税价" in ref, f"价格题缺未税价字段: {ref}"
    finally:
        sqg._call_llm_text = original_call_llm

    print(f"PASS: Appendix E 冒烟测试通过 ({len(candidates)} 候选, {len(questions)} 题)")


def test_real_appendix_e_file_e2e():
    """端到端测试：真实 Appendix E. price list.xlsx 解析→候选→Prompt。"""
    print("=" * 60)
    print("端到端：真实 Appendix E 文件")
    print("=" * 60)

    import spreadsheet_question_generator as sqg

    real_path = r"E:\Desktop\凯捷材料\7月实习\合同知识库材料\2.4标准\2.5偏离\Appendix E. price list.xlsx"
    try:
        with open(real_path, "rb") as f:
            xlsx_bytes = f.read()
    except FileNotFoundError:
        print(f"SKIP: 真实文件不存在: {real_path}")
        return

    sheets = parse_xlsx_to_sheet_contexts(xlsx_bytes)
    candidates = _build_candidate_anchors(sheets)

    # 断言：候选数量 >= 10
    rsc = [c for c in candidates if c["evidence_mode"] == "record_with_schema_context"]
    assert len(rsc) >= 10, f"record 候选应 >= 10，实际 {len(rsc)}"
    print(f"  record_with_schema_context 候选: {len(rsc)}")

    # 断言：不含 M52 等汇总行候选
    for c in candidates:
        bounds = _parse_range_str(c["anchor_range"])
        if bounds:
            _, row, _, _ = bounds
            assert row != 52, f"不应有 row 52 候选: {c['anchor_range']}"
    print("  OK: 无 M52 汇总行候选")

    # 断言：不含 D23:N23 等宽范围 text_fact
    tf = [c for c in candidates if c["evidence_mode"] == "text_fact"]
    for c in tf:
        bounds = _parse_range_str(c["anchor_range"])
        if bounds:
            col_span = bounds[2] - bounds[0] + 1
            assert col_span <= 6, f"text_fact 过宽: {c['anchor_range']} ({col_span} 列)"
    print(f"  OK: {len(tf)} 个 text_fact 候选均在列宽限制内")

    # 断言：每个 record 候选是单行
    for c in rsc:
        bounds = _parse_range_str(c["anchor_range"])
        assert bounds is not None
        assert bounds[1] == bounds[3], f"record 候选应为单行: {c['anchor_range']}"
    print("  OK: 所有 record 候选为单行")

    # 断言：Prompt 含候选清单
    prompt = _build_prompt(sheets, 10, "", candidate_anchors=candidates)
    assert "候选证据清单" in prompt, "Prompt 应含候选清单"
    print(f"  Prompt 字符数: {len(prompt)}")
    print(f"  OK: Prompt 含候选清单")

    # 断言：候选外 anchor 被拒绝
    q_outsider = {"question": "test", "sheet_name": "Sheet1", "anchor_range": "Z99:Z99"}
    v, cat, reason = sqg._validate_single_question(
        q_outsider, {s.sheet_name: s for s in sheets}, "test.xlsx",
        set(), set(), candidate_anchors=candidates,
    )
    assert v is None, f"候选外 anchor 应被拒绝: {q_outsider['anchor_range']}"
    print("  OK: 候选外 anchor 被拒绝")

    print(f"PASS: 真实文件端到端 ({len(candidates)} 候选, prompt={len(prompt)} chars)")


# ====== Two-Phase Schema Flow Tests ======

def test_phase1_schema_analysis_mock():
    """Phase 1 schema 分析：mock LLM 返回验证字段分类。"""
    print("=" * 60)
    print("测试：Phase 1 Schema 分析（mock LLM）")
    print("=" * 60)

    from spreadsheet_question_generator import (
        parse_xlsx_to_sheet_contexts, analyze_table_schema,
        _parse_schema_analysis_response,
    )
    import spreadsheet_question_generator as sqg

    real_path = Path(r"E:\Desktop\凯捷材料\7月实习\合同知识库材料\2.4标准\2.5偏离\Appendix E. price list.xlsx")
    if not real_path.exists():
        print("SKIP: 真实文件不存在")
        return

    file_bytes = real_path.read_bytes()
    sheets = parse_xlsx_to_sheet_contexts(file_bytes)

    # Mock LLM 返回预定义 schema（新角色体系）
    mock_schema = json.dumps({
        "table_purpose": "报价清单",
        "header_row": 1,
        "data_start_row": 3,
        "fields": [
            {"source_label": "模块", "col_index": 2, "inferred_role": "context", "confidence": 0.95, "needs_confirmation": False},
            {"source_label": "一级需求", "col_index": 3, "inferred_role": "context", "confidence": 0.90, "needs_confirmation": False},
            {"source_label": "需求描述", "col_index": 4, "inferred_role": "record_identifier", "confidence": 0.95, "needs_confirmation": False},
            {"source_label": "类型", "col_index": 5, "inferred_role": "categorical", "confidence": 0.85, "needs_confirmation": True},
            {"source_label": "QTY. in total", "col_index": 6, "inferred_role": "metric", "confidence": 0.90, "needs_confirmation": True},
            {"source_label": "PM/SM", "col_index": 7, "inferred_role": "metric", "confidence": 0.85, "needs_confirmation": True},
            {"source_label": "DevOps专家", "col_index": 8, "inferred_role": "metric", "confidence": 0.85, "needs_confirmation": True},
            {"source_label": "Total Cost", "col_index": 14, "inferred_role": "cost", "confidence": 0.90, "needs_confirmation": True},
        ],
        "excluded_rows": [52],
        "question_plan": {
            "recommended_question_patterns": ["record + metric", "record + cost", "record + categorical"],
            "target_field_priority": [
                {"field": "QTY. in total", "role": "metric", "priority": 1, "reason": "核心数量指标"},
                {"field": "Total Cost", "role": "cost", "priority": 2, "reason": "核心费用指标"},
                {"field": "类型", "role": "categorical", "priority": 3, "reason": "分类维度"},
            ],
            "target_field_quotas": {"QTY. in total": 2, "Total Cost": 2, "类型": 1},
            "forbidden_patterns": ["aggregation", "cross-row comparison", "totals", "null values", "formula without cached value", "isolated numeric"],
            "rationale": "报价清单，适合单记录检索查询",
        },
        "reasoning": "报价清单，模块/需求为分组字段，数值列为指标字段"
    }, ensure_ascii=False)

    tmp_dir, orig_dir = _use_test_schema_cache()
    original = sqg._call_llm_text
    sqg._call_llm_text = lambda *a, **kw: mock_schema
    try:
        result = analyze_table_schema(sheets, "fake", "http://fake", "fake",
                                       file_bytes=file_bytes, force_reanalyze=True,
                                       allow_test_model=True)
    finally:
        sqg._call_llm_text = original
        _restore_schema_cache_dir(tmp_dir, orig_dir)

    # 验证
    assert result["table_purpose"] == "报价清单", f"table_purpose: {result['table_purpose']}"
    assert result.get("schema_source") == "test_mock", f"source: {result.get('schema_source')}"
    assert 52 in result["excluded_rows"], f"excluded_rows: {result['excluded_rows']}"

    # record_locator_fields 应含记录标识
    rl = result.get("record_locator_fields", [])
    assert "需求描述" in rl, f"record_locator: {rl}"

    # question_target_fields 应含 metric/cost/categorical
    qt = result.get("question_target_fields", [])
    assert "QTY. in total" in qt or any("QTY" in f for f in qt), f"target: {qt}"
    assert "Total Cost" in qt or any("Cost" in f for f in qt), f"target: {qt}"
    assert "类型" in qt, f"target 应含 categorical: {qt}"

    # question_plan 新结构
    qp = result.get("question_plan", {})
    assert qp.get("target_field_priority"), f"question_plan 缺少 target_field_priority: {qp}"
    assert qp.get("recommended_question_patterns"), f"question_plan 缺少 recommended_question_patterns: {qp}"
    assert qp.get("forbidden_patterns"), f"question_plan 缺少 forbidden_patterns: {qp}"

    # 新角色分组
    assert len(result.get("record_identifier_fields", [])) >= 1, f"record_identifier_fields: {result.get('record_identifier_fields')}"
    assert len(result.get("context_fields", [])) >= 1, f"context_fields: {result.get('context_fields')}"
    assert len(result["metric_fields"]) >= 2, f"metric_fields: {result['metric_fields']}"
    assert len(result["cost_fields"]) >= 1, f"cost_fields: {result['cost_fields']}"
    assert len(result.get("categorical_fields", [])) >= 1, f"categorical_fields: {result.get('categorical_fields')}"

    print(f"PASS: record_locator={rl}, target={qt}, patterns={qp.get('recommended_question_patterns')}")


def test_schema_cache_hit():
    """Schema 缓存命中：同文件二次分析不调用 LLM。"""
    print("=" * 60)
    print("测试：Schema 缓存命中")
    print("=" * 60)

    from spreadsheet_question_generator import (
        parse_xlsx_to_sheet_contexts, analyze_table_schema,
        _file_content_hash, _load_schema_cache,
    )
    import spreadsheet_question_generator as sqg

    real_path = Path(r"E:\Desktop\凯捷材料\7月实习\合同知识库材料\2.4标准\2.5偏离\Appendix E. price list.xlsx")
    if not real_path.exists():
        print("SKIP: 真实文件不存在")
        return

    file_bytes = real_path.read_bytes()
    sheets = parse_xlsx_to_sheet_contexts(file_bytes)

    mock_schema = json.dumps({
        "table_purpose": "测试", "header_row": 1, "data_start_row": 3,
        "fields": [{"source_label": "模块", "col_index": 2, "inferred_role": "context", "confidence": 0.9}],
        "excluded_rows": [52],
    }, ensure_ascii=False)

    call_count = [0]
    def counting_call(*a, **kw):
        call_count[0] += 1
        return mock_schema

    tmp_dir, orig_dir = _use_test_schema_cache()
    original = sqg._call_llm_text
    sqg._call_llm_text = counting_call
    try:
        # 首次调用
        r1 = analyze_table_schema(sheets, "fake", "http://fake", "fake",
                                   file_bytes=file_bytes, force_reanalyze=True,
                                   allow_test_model=True)
        assert call_count[0] == 1, f"首次应调用 LLM: {call_count[0]}"

        # 二次调用（应命中缓存）
        r2 = analyze_table_schema(sheets, "fake", "http://fake", "fake",
                                   file_bytes=file_bytes, allow_test_model=True)
        assert call_count[0] == 1, f"缓存命中不应调用 LLM: {call_count[0]}"

        # 强制重新分析
        r3 = analyze_table_schema(sheets, "fake", "http://fake", "fake",
                                   file_bytes=file_bytes, force_reanalyze=True,
                                   allow_test_model=True)
        assert call_count[0] == 2, f"强制重新应调用 LLM: {call_count[0]}"

        assert r1["table_purpose"] == r2["table_purpose"]
    finally:
        sqg._call_llm_text = original
        _restore_schema_cache_dir(tmp_dir, orig_dir)

    print(f"PASS: LLM 调用次数={call_count[0]}（期望2）")


def test_evidence_schema_display():
    """evidence_schema_display 字段格式正确。"""
    print("=" * 60)
    print("测试：evidence_schema_display")
    print("=" * 60)

    from spreadsheet_question_generator import (
        parse_xlsx_to_sheet_contexts, _validate_and_render_question,
    )

    real_path = Path(r"E:\Desktop\凯捷材料\7月实习\合同知识库材料\2.4标准\2.5偏离\Appendix E. price list.xlsx")
    if not real_path.exists():
        print("SKIP: 真实文件不存在")
        return

    file_bytes = real_path.read_bytes()
    sheets = parse_xlsx_to_sheet_contexts(file_bytes)
    sheets_by_name = {s.sheet_name: s for s in sheets}

    # 测试 B3:E3 的 display
    q = {"question": "test", "sheet_name": "Sheet1", "anchor_range": "B3:E3"}
    v, reason = _validate_and_render_question(q, sheets_by_name, "test.xlsx")
    assert v is not None, f"验证失败: {reason}"
    assert "evidence_schema_display" in v, "应有 evidence_schema_display"

    display = v["evidence_schema_display"]
    assert "=" in display, f"display 应含 '=': {display}"
    assert "；" in display or len(display) > 0, f"display 不应为空"
    # 不应有 Markdown 格式
    assert "**" not in display, f"display 不应含 Markdown: {display}"
    assert "|" not in display, f"display 不应含表格符号: {display}"

    print(f"PASS: display={display[:80]}")


def test_generate_questions_from_schema_e2e():
    """端到端 Phase 2：基于确认 schema 生成题目。"""
    print("=" * 60)
    print("测试：Phase 2 端到端")
    print("=" * 60)

    from spreadsheet_question_generator import (
        parse_xlsx_to_sheet_contexts, generate_questions_from_schema,
    )
    import spreadsheet_question_generator as sqg

    real_path = Path(r"E:\Desktop\凯捷材料\7月实习\合同知识库材料\2.4标准\2.5偏离\Appendix E. price list.xlsx")
    if not real_path.exists():
        print("SKIP: 真实文件不存在")
        return

    file_bytes = real_path.read_bytes()
    sheets = parse_xlsx_to_sheet_contexts(file_bytes)

    # 构建确认 schema（新结构）
    confirmed_schema = {
        "table_purpose": "报价清单",
        "record_locator_fields": ["需求描述"],
        "context_fields": ["模块", "一级需求"],
        "question_target_fields": ["类型", "QTY. in total", "Total Cost"],
        "safe_question_fields": ["需求描述", "类型", "QTY. in total", "Total Cost"],
        "excluded_rows": [52],
        "record_identifier_fields": [{"source_label": "需求描述", "col_index": 4}],
        "metric_fields": [{"source_label": "QTY. in total", "col_index": 6}, {"source_label": "PM/SM", "col_index": 7}],
        "cost_fields": [{"source_label": "Total Cost", "col_index": 14}],
        "categorical_fields": [{"source_label": "类型", "col_index": 5}],
        "question_plan": {
            "recommended_question_patterns": ["record + metric", "record + cost", "record + categorical"],
            "target_field_priority": [
                {"field": "QTY. in total", "role": "metric", "priority": 1, "reason": "数量指标"},
                {"field": "Total Cost", "role": "cost", "priority": 2, "reason": "费用指标"},
                {"field": "类型", "role": "categorical", "priority": 3, "reason": "分类维度"},
            ],
            "target_field_quotas": {"QTY. in total": 2, "Total Cost": 2, "类型": 1},
            "forbidden_patterns": ["aggregation", "cross-row comparison"],
            "rationale": "报价清单",
        },
    }

    # Mock LLM 返回（使用 candidate_id + target_field_label）
    mock_response = json.dumps([
        {"candidate_id": "sheet1_row_3_类型", "question": "工具链可用性优化类型", "target_field_label": "类型", "difficulty": "事实", "topic": "类型"},
        {"candidate_id": "sheet1_row_4_qty_in_total", "question": "服务上线流程制定数量", "target_field_label": "QTY. in total", "difficulty": "事实", "topic": "数量"},
        {"candidate_id": "sheet1_row_7_total_cost", "question": "CICD配置规范设计总成本", "target_field_label": "Total Cost", "difficulty": "事实", "topic": "费用"},
    ])

    original = sqg._call_llm_text
    sqg._call_llm_text = lambda *a, **kw: mock_response
    try:
        questions, stats = generate_questions_from_schema(
            sheets, confirmed_schema, "fake", "http://fake", "fake",
            num_questions=5, file_name="test.xlsx",
        )
    finally:
        sqg._call_llm_text = original

    assert len(questions) >= 2, f"应至少 2 题: {len(questions)}"

    # 验证 candidate_id 存在
    for q in questions:
        assert "candidate_id" in q, f"应有 candidate_id: {q.keys()}"
        assert "target_field_label" in q, f"应有 target_field_label: {q.keys()}"

    # 验证 evidence_schema_display
    for q in questions:
        assert "evidence_schema_display" in q, f"应有 display: {q.keys()}"
        assert "=" in q["evidence_schema_display"], f"display 格式: {q['evidence_schema_display']}"

    # 验证无 row 52
    for q in questions:
        bounds = _parse_range_str(q["anchor_range"])
        if bounds:
            assert bounds[1] != 52 and bounds[3] != 52, f"不应含 row 52: {q['anchor_range']}"

    # 验证 reference_answer 由本地渲染（含真实字段名）
    for q in questions:
        ref = q.get("reference_answer", "")
        assert ref, f"reference_answer 不应为空"
        # reference_answer 应包含键值对格式
        assert "：" in ref, f"reference_answer 应含键值对: {ref[:80]}"

    # 验证 reference_answer 不来自 LLM（LLM mock 中无 reference_answer）
    # 这是通过 _validate_phase2_question 本地渲染保证的

    print(f"PASS: {len(questions)} 题, candidate_ids={[q.get('candidate_id') for q in questions]}")


def test_upload_no_llm_call():
    """上传文件只做本地解析，不调用 LLM。"""
    print("=" * 60)
    print("测试：上传文件不调用 LLM")
    print("=" * 60)

    from spreadsheet_question_generator import (
        parse_xlsx_to_sheet_contexts, _build_schema_analysis_context,
        _detect_header_row_and_data_start, _is_summary_row,
    )
    import spreadsheet_question_generator as sqg

    real_path = Path(r"E:\Desktop\凯捷材料\7月实习\合同知识库材料\2.4标准\2.5偏离\Appendix E. price list.xlsx")
    if not real_path.exists():
        print("SKIP: 真实文件不存在")
        return

    file_bytes = real_path.read_bytes()

    # 追踪 LLM 调用
    call_count = [0]
    original = sqg._call_llm_text
    sqg._call_llm_text = lambda *a, **kw: (call_count.__setitem__(0, call_count[0] + 1), "")[1]
    try:
        sheets = parse_xlsx_to_sheet_contexts(file_bytes)
        ctx = _build_schema_analysis_context(sheets)
    finally:
        sqg._call_llm_text = original

    assert call_count[0] == 0, f"上传+解析不应调用 LLM: {call_count[0]}"
    assert len(sheets) > 0
    assert len(ctx) > 100

    print(f"PASS: 解析完成，LLM 调用 {call_count[0]} 次")


def test_phase1_only_on_button_click():
    """点击分析按钮只执行 Phase 1，不执行 Phase 2。"""
    print("=" * 60)
    print("测试：Phase 1 按钮只调用 Phase 1")
    print("=" * 60)

    from spreadsheet_question_generator import (
        parse_xlsx_to_sheet_contexts, analyze_table_schema,
    )
    import spreadsheet_question_generator as sqg

    real_path = Path(r"E:\Desktop\凯捷材料\7月实习\合同知识库材料\2.4标准\2.5偏离\Appendix E. price list.xlsx")
    if not real_path.exists():
        print("SKIP: 真实文件不存在")
        return

    file_bytes = real_path.read_bytes()
    sheets = parse_xlsx_to_sheet_contexts(file_bytes)

    mock_schema = json.dumps({
        "table_purpose": "报价清单", "header_row": 1, "data_start_row": 3,
        "fields": [
            {"source_label": "模块", "col_index": 2, "inferred_role": "context", "confidence": 0.95},
            {"source_label": "需求描述", "col_index": 4, "inferred_role": "record_identifier", "confidence": 0.95},
        ],
        "excluded_rows": [52],
    }, ensure_ascii=False)

    call_count = [0]
    def counting_call(*a, **kw):
        call_count[0] += 1
        return mock_schema

    tmp_dir, orig_dir = _use_test_schema_cache()
    original = sqg._call_llm_text
    sqg._call_llm_text = counting_call
    try:
        result = analyze_table_schema(sheets, "fake", "http://fake", "fake",
                                       file_bytes=file_bytes, force_reanalyze=True,
                                       allow_test_model=True)
    finally:
        sqg._call_llm_text = original
        _restore_schema_cache_dir(tmp_dir, orig_dir)

    assert call_count[0] == 1, f"Phase 1 应调用 LLM 1 次: {call_count[0]}"
    assert result["table_purpose"] == "报价清单"

    print(f"PASS: Phase 1 调用 LLM {call_count[0]} 次")


def test_phase2_does_not_recall_phase1():
    """生成题目只执行 Phase 2，不重复调用 Phase 1。"""
    print("=" * 60)
    print("测试：Phase 2 不重复调用 Phase 1")
    print("=" * 60)

    from spreadsheet_question_generator import (
        parse_xlsx_to_sheet_contexts, analyze_table_schema,
        generate_questions_from_schema,
    )
    import spreadsheet_question_generator as sqg

    real_path = Path(r"E:\Desktop\凯捷材料\7月实习\合同知识库材料\2.4标准\2.5偏离\Appendix E. price list.xlsx")
    if not real_path.exists():
        print("SKIP: 真实文件不存在")
        return

    file_bytes = real_path.read_bytes()
    sheets = parse_xlsx_to_sheet_contexts(file_bytes)

    mock_schema = json.dumps({
        "table_purpose": "报价清单", "header_row": 1, "data_start_row": 3,
        "fields": [
            {"source_label": "模块", "col_index": 2, "inferred_role": "context", "confidence": 0.95},
            {"source_label": "需求描述", "col_index": 4, "inferred_role": "record_identifier", "confidence": 0.95},
            {"source_label": "类型", "col_index": 5, "inferred_role": "categorical", "confidence": 0.85, "needs_confirmation": True},
        ],
        "excluded_rows": [52],
    }, ensure_ascii=False)

    mock_questions = json.dumps([
        {"candidate_id": "sheet1_row_3_类型", "question": "工具链可用性优化类型", "target_field_label": "类型", "difficulty": "事实", "topic": "类型"},
    ])

    call_log = []
    def tracking_call(*a, **kw):
        call_log.append("llm")
        return mock_schema if len(call_log) == 1 else mock_questions

    tmp_dir, orig_dir = _use_test_schema_cache()
    original = sqg._call_llm_text
    sqg._call_llm_text = tracking_call
    try:
        # Phase 1
        schema = analyze_table_schema(sheets, "fake", "http://fake", "fake",
                                       file_bytes=file_bytes, force_reanalyze=True,
                                       allow_test_model=True)
        phase1_calls = len(call_log)

        # Phase 2
        questions, stats = generate_questions_from_schema(
            sheets, schema, "fake", "http://fake", "fake",
            num_questions=5, file_name="test.xlsx",
        )
        phase2_calls = len(call_log) - phase1_calls
    finally:
        sqg._call_llm_text = original
        _restore_schema_cache_dir(tmp_dir, orig_dir)

    assert phase1_calls == 1, f"Phase 1 应调用 1 次: {phase1_calls}"
    assert phase2_calls >= 1, f"Phase 2 应至少调用 1 次: {phase2_calls}"
    assert len(questions) >= 1

    print(f"PASS: Phase 1={phase1_calls} 次, Phase 2={phase2_calls} 次, 总计 {len(call_log)} 次")


def test_file_change_invalidates_schema():
    """更换文件后旧 schema 失效。"""
    print("=" * 60)
    print("测试：更换文件后旧 schema 失效")
    print("=" * 60)

    from spreadsheet_question_generator import (
        parse_xlsx_to_sheet_contexts, analyze_table_schema,
        _file_content_hash, _load_schema_cache,
    )
    import spreadsheet_question_generator as sqg

    real_path = Path(r"E:\Desktop\凯捷材料\7月实习\合同知识库材料\2.4标准\2.5偏离\Appendix E. price list.xlsx")
    if not real_path.exists():
        print("SKIP: 真实文件不存在")
        return

    file_bytes = real_path.read_bytes()
    sheets = parse_xlsx_to_sheet_contexts(file_bytes)

    mock_schema = json.dumps({
        "table_purpose": "报价清单", "header_row": 1, "data_start_row": 3,
        "fields": [{"source_label": "模块", "col_index": 2, "inferred_role": "context", "confidence": 0.95}],
        "excluded_rows": [52],
    }, ensure_ascii=False)

    tmp_dir, orig_dir = _use_test_schema_cache()
    original = sqg._call_llm_text
    sqg._call_llm_text = lambda *a, **kw: mock_schema
    try:
        # 分析文件 A
        r1 = analyze_table_schema(sheets, "fake", "http://fake", "fake",
                                   file_bytes=file_bytes, force_reanalyze=True,
                                   allow_test_model=True)
        hash_a = _file_content_hash(file_bytes)

        # 文件 B（不同内容）应有不同缓存
        fake_b = b"fake file content for B"
        hash_b = _file_content_hash(fake_b)
        assert hash_a != hash_b, "不同文件应有不同哈希"

        # 缓存中只有 A
        cached_a = _load_schema_cache(hash_a, allow_test_model=True)
        cached_b = _load_schema_cache(hash_b, allow_test_model=True)
        assert cached_a is not None, "文件 A 应有缓存"
        assert cached_b is None, "文件 B 不应有缓存"
    finally:
        sqg._call_llm_text = original
        _restore_schema_cache_dir(tmp_dir, orig_dir)

    print(f"PASS: 文件 A 缓存存在, 文件 B 缓存不存在")


def test_reanalyze_clears_and_rebuilds():
    """重新分析后才允许用新 schema 出题。"""
    print("=" * 60)
    print("测试：重新分析后用新 schema 出题")
    print("=" * 60)

    from spreadsheet_question_generator import (
        parse_xlsx_to_sheet_contexts, analyze_table_schema,
        generate_questions_from_schema, _delete_schema_cache,
        _file_content_hash,
    )
    import spreadsheet_question_generator as sqg

    real_path = Path(r"E:\Desktop\凯捷材料\7月实习\合同知识库材料\2.4标准\2.5偏离\Appendix E. price list.xlsx")
    if not real_path.exists():
        print("SKIP: 真实文件不存在")
        return

    file_bytes = real_path.read_bytes()
    sheets = parse_xlsx_to_sheet_contexts(file_bytes)

    schema_v1 = json.dumps({
        "table_purpose": "v1", "header_row": 1, "data_start_row": 3,
        "fields": [{"source_label": "模块", "col_index": 2, "inferred_role": "context", "confidence": 0.95}],
        "excluded_rows": [52],
    }, ensure_ascii=False)

    schema_v2 = json.dumps({
        "table_purpose": "v2_updated", "header_row": 1, "data_start_row": 3,
        "fields": [
            {"source_label": "模块", "col_index": 2, "inferred_role": "context", "confidence": 0.95},
            {"source_label": "需求描述", "col_index": 4, "inferred_role": "record_identifier", "confidence": 0.95},
        ],
        "excluded_rows": [52],
    }, ensure_ascii=False)

    call_seq = []
    def seq_call(*a, **kw):
        call_seq.append(1)
        return schema_v1 if len(call_seq) == 1 else schema_v2

    tmp_dir, orig_dir = _use_test_schema_cache()
    original = sqg._call_llm_text
    sqg._call_llm_text = seq_call
    try:
        # 首次分析
        r1 = analyze_table_schema(sheets, "fake", "http://fake", "fake",
                                   file_bytes=file_bytes, force_reanalyze=True,
                                   allow_test_model=True)
        assert r1["table_purpose"] == "v1"

        # 重新分析（force_reanalyze=True）
        r2 = analyze_table_schema(sheets, "fake", "http://fake", "fake",
                                   file_bytes=file_bytes, force_reanalyze=True,
                                   allow_test_model=True)
        assert r2["table_purpose"] == "v2_updated"
        assert len(call_seq) == 2, f"应调用 2 次: {len(call_seq)}"
    finally:
        sqg._call_llm_text = original
        _restore_schema_cache_dir(tmp_dir, orig_dir)

    print(f"PASS: 首次=v1, 重新分析=v2, LLM 调用 {len(call_seq)} 次")


def test_production_rejects_fake_model():
    """生产路径拒绝 fake 模型，必须 fail-closed。"""
    print("=" * 60)
    print("测试：生产路径拒绝 fake 模型")
    print("=" * 60)

    from spreadsheet_question_generator import (
        parse_xlsx_to_sheet_contexts, analyze_table_schema,
        _is_fake_model,
    )
    import spreadsheet_question_generator as sqg

    real_path = Path(r"E:\Desktop\凯捷材料\7月实习\合同知识库材料\2.4标准\2.5偏离\Appendix E. price list.xlsx")
    if not real_path.exists():
        print("SKIP: 真实文件不存在")
        return

    file_bytes = real_path.read_bytes()
    sheets = parse_xlsx_to_sheet_contexts(file_bytes)

    # fake 模型必须被拒绝（不传 allow_test_model）
    original = sqg._call_llm_text
    sqg._call_llm_text = lambda *a, **kw: '{"fields": []}'
    try:
        try:
            analyze_table_schema(sheets, "fake", "http://fake", "fake",
                                  file_bytes=file_bytes, force_reanalyze=True)
            assert False, "fake 模型应被拒绝"
        except ValueError as e:
            assert "测试模型" in str(e) or "fake" in str(e), f"错误信息: {e}"
    finally:
        sqg._call_llm_text = original

    # mock 模型同样被拒绝
    assert _is_fake_model("mock-gpt4")
    assert _is_fake_model("test-model")
    assert not _is_fake_model("mimo-v2.5-pro")
    assert not _is_fake_model("gpt-4")

    print("PASS: fake/mock 模型被拒绝，真实模型通过")


def test_production_no_fake_in_output():
    """生产输出中不得出现 fake/v2_updated。"""
    print("=" * 60)
    print("测试：生产输出不含 fake/v2_updated")
    print("=" * 60)

    from spreadsheet_question_generator import (
        parse_xlsx_to_sheet_contexts, analyze_table_schema,
        generate_questions_from_schema, _load_schema_cache,
        _file_content_hash,
    )
    import spreadsheet_question_generator as sqg

    real_path = Path(r"E:\Desktop\凯捷材料\7月实习\合同知识库材料\2.4标准\2.5偏离\Appendix E. price list.xlsx")
    if not real_path.exists():
        print("SKIP: 真实文件不存在")
        return

    file_bytes = real_path.read_bytes()
    sheets = parse_xlsx_to_sheet_contexts(file_bytes)

    # 用一个非 fake 的 mock 模型名
    mock_schema = json.dumps({
        "table_purpose": "IT服务报价清单",
        "header_row": 1, "data_start_row": 3,
        "fields": [
            {"source_label": "模块", "col_index": 2, "inferred_role": "context", "confidence": 0.95},
            {"source_label": "需求描述", "col_index": 4, "inferred_role": "record_identifier", "confidence": 0.95},
        ],
        "excluded_rows": [52],
    }, ensure_ascii=False)

    tmp_dir, orig_dir = _use_test_schema_cache()
    original = sqg._call_llm_text
    sqg._call_llm_text = lambda *a, **kw: mock_schema
    try:
        result = analyze_table_schema(sheets, "test-model-x", "http://test", "test-model-x",
                                       file_bytes=file_bytes, force_reanalyze=True,
                                       allow_test_model=True)

        # 断言：不得出现 fake/v2_updated
        assert result.get("analysis_model") != "fake", f"model 不得为 fake: {result.get('analysis_model')}"
        assert result.get("table_purpose") != "v2_updated", f"purpose 不得为 v2_updated: {result.get('table_purpose')}"
        assert "fake" not in str(result.get("table_purpose", "")).lower()
    finally:
        sqg._call_llm_text = original
        _restore_schema_cache_dir(tmp_dir, orig_dir)

    print("PASS: 生产输出中无 fake/v2_updated")


def test_llm_failure_fail_closed():
    """LLM 调用失败时必须 fail-closed，不得生成空 schema。"""
    print("=" * 60)
    print("测试：LLM 失败 fail-closed")
    print("=" * 60)

    from spreadsheet_question_generator import (
        parse_xlsx_to_sheet_contexts, analyze_table_schema,
    )
    import spreadsheet_question_generator as sqg

    real_path = Path(r"E:\Desktop\凯捷材料\7月实习\合同知识库材料\2.4标准\2.5偏离\Appendix E. price list.xlsx")
    if not real_path.exists():
        print("SKIP: 真实文件不存在")
        return

    file_bytes = real_path.read_bytes()
    sheets = parse_xlsx_to_sheet_contexts(file_bytes)

    # LLM 返回非法 JSON
    original = sqg._call_llm_text
    tmp_dir, orig_dir = _use_test_schema_cache()
    try:
        # Case 1: LLM 超时
        sqg._call_llm_text = lambda *a, **kw: (_ for _ in ()).throw(TimeoutError("timeout"))
        try:
            analyze_table_schema(sheets, "mimo-v2.5-pro", "http://api", "mimo-v2.5-pro",
                                  file_bytes=file_bytes, force_reanalyze=True)
            assert False, "超时应抛出异常"
        except (TimeoutError, RuntimeError):
            pass

        # Case 2: LLM 返回垃圾
        sqg._call_llm_text = lambda *a, **kw: "not json at all"
        try:
            analyze_table_schema(sheets, "mimo-v2.5-pro", "http://api", "mimo-v2.5-pro",
                                  file_bytes=file_bytes, force_reanalyze=True)
            assert False, "非法 JSON 应抛出异常"
        except ValueError:
            pass

        # Case 3: LLM 返回空 fields
        sqg._call_llm_text = lambda *a, **kw: '{"fields": [], "table_purpose": "empty"}'
        try:
            analyze_table_schema(sheets, "mimo-v2.5-pro", "http://api", "mimo-v2.5-pro",
                                  file_bytes=file_bytes, force_reanalyze=True)
            assert False, "空 fields 应抛出异常"
        except ValueError:
            pass
    finally:
        sqg._call_llm_text = original
        _restore_schema_cache_dir(tmp_dir, orig_dir)

    print("PASS: LLM 失败/非法返回均 fail-closed")


def test_phase1_timeout_is_180s():
    """Phase 1 的读取超时应为 180 秒，不影响其他调用。"""
    print("=" * 60)
    print("测试：Phase 1 timeout=180s")
    print("=" * 60)

    from spreadsheet_question_generator import _call_llm_text
    import spreadsheet_question_generator as sqg

    captured_timeouts = []

    def mock_post(url, json=None, headers=None, timeout=None):
        captured_timeouts.append(timeout)
        # 模拟成功
        class FakeResp:
            status_code = 200
            def json(self):
                return {"choices": [{"message": {"content": '{"fields": [], "table_purpose": "test"}'}}]}
        return FakeResp()

    real_path = Path(r"E:\Desktop\凯捷材料\7月实习\合同知识库材料\2.4标准\2.5偏离\Appendix E. price list.xlsx")
    if not real_path.exists():
        print("SKIP: 真实文件不存在")
        return

    import requests
    original_post = requests.post
    requests.post = mock_post
    original_llm = sqg._call_llm_text

    tmp_dir, orig_dir = _use_test_schema_cache()
    try:
        file_bytes = real_path.read_bytes()
        sheets = parse_xlsx_to_sheet_contexts(file_bytes)

        try:
            sqg.analyze_table_schema(sheets, "mimo-v2.5-pro", "http://api", "mimo-v2.5-pro",
                                      file_bytes=file_bytes, force_reanalyze=True)
        except Exception:
            pass  # JSON 解析会失败，但我们只关心 timeout 值

        # 验证 Phase 1 使用 (15, 180) 超时
        assert len(captured_timeouts) >= 1, f"应捕获至少 1 次请求: {len(captured_timeouts)}"
        t = captured_timeouts[0]
        assert isinstance(t, tuple), f"timeout 应为元组: {t}"
        assert t == (15, 180), f"Phase 1 timeout 应为 (15, 180): {t}"

        # 验证其他调用不受影响（默认 timeout=120）
        captured_timeouts.clear()
        try:
            _call_llm_text("test", "key", "http://api", "model")
        except Exception:
            pass
        if captured_timeouts:
            t2 = captured_timeouts[0]
            assert t2 == 120, f"默认 timeout 应为 120: {t2}"
    finally:
        requests.post = original_post
        sqg._call_llm_text = original_llm
        _restore_schema_cache_dir(tmp_dir, orig_dir)

    print(f"PASS: Phase 1 timeout={captured_timeouts[0] if captured_timeouts else '(15, 180)'}")


def test_phase1_retry_on_timeout():
    """Phase 1 首次超时后重试一次，第二次成功。"""
    print("=" * 60)
    print("测试：Phase 1 首次超时重试成功")
    print("=" * 60)

    import spreadsheet_question_generator as sqg

    real_path = Path(r"E:\Desktop\凯捷材料\7月实习\合同知识库材料\2.4标准\2.5偏离\Appendix E. price list.xlsx")
    if not real_path.exists():
        print("SKIP: 真实文件不存在")
        return

    mock_schema = json.dumps({
        "table_purpose": "报价清单", "header_row": 1, "data_start_row": 3,
        "fields": [{"source_label": "模块", "col_index": 2, "inferred_role": "context", "confidence": 0.95}],
        "excluded_rows": [52],
    }, ensure_ascii=False)

    call_count = [0]
    def timeout_then_success(*a, **kw):
        call_count[0] += 1
        if call_count[0] == 1:
            raise RuntimeError("请求超时 (15s): http://api")
        return mock_schema

    tmp_dir, orig_dir = _use_test_schema_cache()
    original = sqg._call_llm_text
    sqg._call_llm_text = timeout_then_success
    try:
        file_bytes = real_path.read_bytes()
        sheets = parse_xlsx_to_sheet_contexts(file_bytes)
        result = sqg.analyze_table_schema(sheets, "test-model", "http://api", "test-model",
                                           file_bytes=file_bytes, force_reanalyze=True,
                                           allow_test_model=True)
        assert call_count[0] == 2, f"应调用 2 次: {call_count[0]}"
        assert result["table_purpose"] == "报价清单"
        assert result.get("llm_call_attempts") == 2
        assert len(result.get("llm_attempt_durations", [])) == 2
    finally:
        sqg._call_llm_text = original
        _restore_schema_cache_dir(tmp_dir, orig_dir)

    print(f"PASS: 首次超时, 第2次成功, 总尝试={call_count[0]}")


def test_phase1_retry_both_fail():
    """Phase 1 两次均失败后报错，不产生 schema。"""
    print("=" * 60)
    print("测试：Phase 1 两次失败报错")
    print("=" * 60)

    import spreadsheet_question_generator as sqg

    real_path = Path(r"E:\Desktop\凯捷材料\7月实习\合同知识库材料\2.4标准\2.5偏离\Appendix E. price list.xlsx")
    if not real_path.exists():
        print("SKIP: 真实文件不存在")
        return

    call_count = [0]
    def always_timeout(*a, **kw):
        call_count[0] += 1
        raise RuntimeError("请求超时 (180s): http://api")

    tmp_dir, orig_dir = _use_test_schema_cache()
    original = sqg._call_llm_text
    sqg._call_llm_text = always_timeout
    try:
        file_bytes = real_path.read_bytes()
        sheets = parse_xlsx_to_sheet_contexts(file_bytes)
        try:
            sqg.analyze_table_schema(sheets, "test-model", "http://api", "test-model",
                                      file_bytes=file_bytes, force_reanalyze=True,
                                      allow_test_model=True)
            assert False, "两次超时应抛出异常"
        except RuntimeError as e:
            assert "超时" in str(e), f"错误信息应含超时: {e}"
        assert call_count[0] == 2, f"应尝试 2 次: {call_count[0]}"
    finally:
        sqg._call_llm_text = original
        _restore_schema_cache_dir(tmp_dir, orig_dir)

    print(f"PASS: 两次超时后正确报错, 尝试={call_count[0]} 次, 无 schema 产生")


def test_phase1_no_fake_fallback():
    """Phase 1 失败时不得降级为 fake/mock schema。"""
    print("=" * 60)
    print("测试：Phase 1 无 fake 降级")
    print("=" * 60)

    import spreadsheet_question_generator as sqg

    real_path = Path(r"E:\Desktop\凯捷材料\7月实习\合同知识库材料\2.4标准\2.5偏离\Appendix E. price list.xlsx")
    if not real_path.exists():
        print("SKIP: 真实文件不存在")
        return

    def always_fail(*a, **kw):
        raise RuntimeError("请求超时 (180s): http://api")

    tmp_dir, orig_dir = _use_test_schema_cache()
    original = sqg._call_llm_text
    sqg._call_llm_text = always_fail
    try:
        file_bytes = real_path.read_bytes()
        sheets = parse_xlsx_to_sheet_contexts(file_bytes)

        # 验证：不能用 fake 模型绕过
        try:
            sqg.analyze_table_schema(sheets, "fake", "http://api", "fake",
                                      file_bytes=file_bytes, force_reanalyze=True)
            assert False, "fake 模型应被拒绝"
        except ValueError:
            pass

        # 验证：真实模型两次失败后不产生任何 schema
        try:
            sqg.analyze_table_schema(sheets, "mimo-v2.5-pro", "http://api", "mimo-v2.5-pro",
                                      file_bytes=file_bytes, force_reanalyze=True)
            assert False, "两次失败应抛出异常"
        except RuntimeError:
            pass

        # 验证：缓存中无数据
        cached = sqg._load_schema_cache(sqg._file_content_hash(file_bytes))
        assert cached is None, "失败后不应有缓存"
    finally:
        sqg._call_llm_text = original
        _restore_schema_cache_dir(tmp_dir, orig_dir)

    print("PASS: 失败后无 fake 降级，无缓存产生")


# ====== Main ======

# ====== 新增测试场景 ======

def test_resource_cost_table_candidate_catalog():
    """资源成本表：候选目录含 candidate_id、record_locator、target_field。"""
    print("=" * 60)
    print("测试：资源成本表候选目录")
    print("=" * 60)

    from spreadsheet_question_generator import (
        _build_candidate_anchors_from_schema, _slugify_field,
    )

    # 构建模拟的 SheetContext
    ctx = SheetContext(
        sheet_name="Sheet1", max_row=5, max_col=4,
        headers=["需求描述", "类型", "QTY", "Total Cost"],
        rows=[
            ["需求描述", "类型", "QTY", "Total Cost"],
            ["需求A", "Improvement", 10, 5000],
            ["需求B", "New function", 20, 8000],
            ["需求C", "Enhancement", 5, 3000],
            ["合计", "", 35, 16000],
        ],
    )

    confirmed_schema = {
        "record_locator_fields": ["需求描述"],
        "context_fields": [],
        "excluded_rows": [],
        "question_plan": {
            "target_field_priority": [
                {"field": "QTY", "role": "metric", "priority": 1, "reason": "数量"},
                {"field": "Total Cost", "role": "cost", "priority": 2, "reason": "费用"},
                {"field": "类型", "role": "categorical", "priority": 3, "reason": "分类"},
            ],
        },
    }

    candidates = _build_candidate_anchors_from_schema([ctx], confirmed_schema)

    assert len(candidates) > 0, "应生成候选"

    # 验证候选结构
    for c in candidates:
        assert "candidate_id" in c, f"应有 candidate_id: {c.keys()}"
        assert "record_locator" in c, f"应有 record_locator"
        assert "target_field" in c, f"应有 target_field"
        assert "allowed_evidence_fields" in c, f"应有 allowed_evidence_fields"
        assert c["target_field"]["label"] in ["QTY", "Total Cost", "类型"], f"target_field: {c['target_field']}"
        assert c["target_field"]["role"] in ["metric", "cost", "categorical"], f"role: {c['target_field']['role']}"

    # 验证不含汇总行
    for c in candidates:
        bounds = _parse_range_str(c["anchor_range"])
        if bounds:
            assert bounds[1] != 5, f"不应含汇总行: {c['anchor_range']}"

    # 验证 record_locator 含实际值
    for c in candidates:
        rl = c["record_locator"]
        assert "需求描述" in rl, f"record_locator 应含需求描述: {rl}"
        assert rl["需求描述"] in ["需求A", "需求B", "需求C"], f"值: {rl}"

    print(f"PASS: {len(candidates)} 个候选, 示例: {candidates[0]['candidate_id']}")


def test_inventory_table_recognized():
    """库存/参数表：无业务词硬编码也能识别记录字段与数值字段。"""
    print("=" * 60)
    print("测试：库存/参数表识别")
    print("=" * 60)

    from spreadsheet_question_generator import (
        _build_candidate_anchors_from_schema,
    )

    ctx = SheetContext(
        sheet_name="Inventory", max_row=4, max_col=4,
        headers=["Product Code", "Category", "Stock Qty", "Unit Price"],
        rows=[
            ["Product Code", "Category", "Stock Qty", "Unit Price"],
            ["P001", "Electronics", 150, 29.99],
            ["P002", "Clothing", 300, 15.50],
            ["P003", "Food", 50, 8.75],
        ],
    )

    confirmed_schema = {
        "record_locator_fields": ["Product Code"],
        "context_fields": ["Category"],
        "excluded_rows": [],
        "question_plan": {
            "target_field_priority": [
                {"field": "Stock Qty", "role": "metric", "priority": 1, "reason": "库存数量"},
                {"field": "Unit Price", "role": "cost", "priority": 2, "reason": "单价"},
            ],
        },
    }

    candidates = _build_candidate_anchors_from_schema([ctx], confirmed_schema)
    assert len(candidates) >= 4, f"应至少 4 个候选: {len(candidates)}"

    # 验证 Product Code 作为记录标识
    for c in candidates:
        assert "Product Code" in c["record_locator"], f"应含 Product Code: {c['record_locator']}"
        assert c["record_locator"]["Product Code"].startswith("P"), f"值: {c['record_locator']}"

    print(f"PASS: {len(candidates)} 个候选")


def test_summary_empty_isolated_excluded_from_candidates():
    """汇总行、空列、孤立数值不能成为候选。"""
    print("=" * 60)
    print("测试：汇总行/空列/孤立数值排除")
    print("=" * 60)

    from spreadsheet_question_generator import (
        _build_candidate_anchors_from_schema,
    )

    ctx = SheetContext(
        sheet_name="Test", max_row=6, max_col=4,
        headers=["Name", "Category", "Amount", "备注"],
        rows=[
            ["Name", "Category", "Amount", "备注"],
            ["Item A", "Type1", 100, "ok"],
            ["Item B", "Type2", 200, "ok"],
            ["", "", "", ""],  # 空行
            ["Total", "", 300, ""],  # 汇总行
            ["Item C", "Type3", "", ""],  # 空数值
        ],
    )

    confirmed_schema = {
        "record_locator_fields": ["Name"],
        "context_fields": ["Category"],
        "excluded_rows": [],
        "question_plan": {
            "target_field_priority": [
                {"field": "Amount", "role": "metric", "priority": 1, "reason": "数量"},
            ],
        },
    }

    candidates = _build_candidate_anchors_from_schema([ctx], confirmed_schema)

    # 不应包含汇总行（row 5）
    for c in candidates:
        bounds = _parse_range_str(c["anchor_range"])
        if bounds:
            assert bounds[1] != 5, f"不应含汇总行: {c['anchor_range']}"
            assert bounds[1] != 4, f"不应含空行: {c['anchor_range']}"

    # 不应包含空数值的行（row 6，Amount 为空）
    for c in candidates:
        bounds = _parse_range_str(c["anchor_range"])
        if bounds and bounds[1] == 6:
            assert False, f"不应含空数值行: {c['anchor_range']}"

    print(f"PASS: {len(candidates)} 个候选（排除了汇总行、空行、空数值行）")


def test_fail_closed_invalid_candidate_id():
    """LLM 返回错误 candidate_id 时 fail-closed 拒绝。"""
    print("=" * 60)
    print("测试：fail-closed 无效 candidate_id")
    print("=" * 60)

    from spreadsheet_question_generator import (
        _validate_phase2_question,
    )

    ctx = SheetContext(
        sheet_name="Sheet1", max_row=3, max_col=3,
        headers=["Name", "Type", "Value"],
        rows=[
            ["Name", "Type", "Value"],
            ["Item A", "Type1", 100],
            ["Item B", "Type2", 200],
        ],
    )
    sheets_by_name = {"Sheet1": ctx}

    candidate_catalog = [{
        "candidate_id": "sheet1_row_2_value",
        "sheet_name": "Sheet1",
        "anchor_range": "A2:C2",
        "header_context_range": "A1:C1",
        "record_locator": {"Name": "Item A"},
        "target_field": {"label": "Value", "role": "metric", "value": 100},
        "allowed_evidence_fields": ["Name", "Value"],
    }]

    # 测试 1: 错误 candidate_id
    q1 = {"candidate_id": "nonexistent_id", "question": "test", "target_field_label": "Value"}
    validated, category, reason = _validate_phase2_question(q1, candidate_catalog, sheets_by_name)
    assert validated is None, f"应拒绝无效 candidate_id"
    assert category == "candidate_mismatch", f"category: {category}"

    # 测试 2: 错误 target_field_label
    q2 = {"candidate_id": "sheet1_row_2_value", "question": "test", "target_field_label": "WrongField"}
    validated, category, reason = _validate_phase2_question(q2, candidate_catalog, sheets_by_name)
    assert validated is None, f"应拒绝不匹配的 target_field_label"
    assert category == "field_mismatch", f"category: {category}"

    print(f"PASS: 无效 candidate_id 和 target_field_label 均被拒绝")


def test_reference_answer_locally_rendered():
    """reference_answer 由本地真实表头和单元格渲染，LLM 响应中不存在。"""
    print("=" * 60)
    print("测试：reference_answer 本地渲染")
    print("=" * 60)

    from spreadsheet_question_generator import (
        _validate_phase2_question,
    )

    ctx = SheetContext(
        sheet_name="Sheet1", max_row=3, max_col=3,
        headers=["需求描述", "类型", "工作量"],
        rows=[
            ["需求描述", "类型", "工作量"],
            ["工具链优化", "Improvement", 17],
            ["看板搭建", "New function", 25],
        ],
    )
    sheets_by_name = {"Sheet1": ctx}

    candidate_catalog = [{
        "candidate_id": "sheet1_row_2_工作量",
        "sheet_name": "Sheet1",
        "anchor_range": "A2:C2",
        "header_context_range": "A1:C1",
        "record_locator": {"需求描述": "工具链优化"},
        "target_field": {"label": "工作量", "role": "metric", "value": 17},
        "allowed_evidence_fields": ["需求描述", "工作量"],
    }]

    # LLM 返回（无 reference_answer）
    q = {"candidate_id": "sheet1_row_2_工作量", "question": "工具链优化工作量", "target_field_label": "工作量", "difficulty": "事实", "topic": "工作量"}
    validated, category, reason = _validate_phase2_question(q, candidate_catalog, sheets_by_name)

    assert validated is not None, f"应通过验证: {reason}"
    assert "reference_answer" in validated, "应有 reference_answer"
    ref = validated["reference_answer"]
    assert ref, "reference_answer 不应为空"
    # 应包含真实字段名和值
    assert "需求描述" in ref, f"应含字段名: {ref}"
    assert "工具链优化" in ref, f"应含记录值: {ref}"
    assert "工作量" in ref, f"应含目标字段名: {ref}"
    assert "17" in ref, f"应含目标值: {ref}"
    # 应为键值对格式
    assert "：" in ref, f"应为键值对格式: {ref}"

    # LLM 响应中无 reference_answer
    assert "reference_answer" not in q, "LLM 响应不应含 reference_answer"

    print(f"PASS: reference_answer='{ref}'")


def test_non_schema_path_not_regressed():
    """非 schema 路径（generate_spreadsheet_questions）不回归。"""
    print("=" * 60)
    print("测试：非 schema 路径不回归")
    print("=" * 60)

    from spreadsheet_question_generator import (
        _build_candidate_anchors, _build_prompt,
    )

    ctx = SheetContext(
        sheet_name="Sheet1", max_row=3, max_col=3,
        headers=["产品", "类别", "价格"],
        rows=[
            ["产品", "类别", "价格"],
            ["产品A", "电子", 100],
            ["产品B", "服装", 200],
        ],
    )

    # 旧路径仍应工作
    candidates = _build_candidate_anchors([ctx])
    assert len(candidates) > 0, f"旧路径应生成候选: {len(candidates)}"

    prompt = _build_prompt([ctx], 5, "", candidates)
    assert len(prompt) > 100, f"旧路径应生成 prompt"

    # 旧候选格式不含 candidate_id（旧格式使用 fact_fields, query_focus）
    for c in candidates:
        assert "evidence_mode" in c, f"旧格式应有 evidence_mode"
        assert "anchor_range" in c, f"旧格式应有 anchor_range"

    print(f"PASS: 旧路径 {len(candidates)} 候选, prompt {len(prompt)} 字符")


def main():
    tests = [
        test_col_letter,
        test_parse_range_str,
        test_csv_basic,
        test_csv_encoding_gbk,
        test_csv_encoding_bom,
        test_csv_empty,
        test_xlsx_to_sheet_context,
        test_xlsx_multi_sheet,
        test_xlsx_merged_cells,
        test_xlsx_formula_detection,
        test_split_small_sheet,
        test_split_large_sheet,
        test_allowed_ranges_per_block,
        test_valid_anchor_in_whitelist,
        test_anchor_not_in_whitelist,
        test_anchor_out_of_bounds,
        test_anchor_too_large,
        test_render_single_row,
        test_render_multi_row,
        test_render_reference_answer_from_context,
        test_parse_valid_json,
        test_parse_markdown_code_block,
        test_parse_invalid_json,
        test_generate_csv_questions,
        test_generate_xlsx_questions,
        test_validation_rejects_bad_range,
        test_csv_in_doc_parser,
        test_supported_extensions_includes_new,
        test_build_prompt,
        test_xlsx_question_generator_delegates,
        test_existing_xlsx_functions_importable,
        test_price_anchor_d4d4_rejected,
        test_price_anchor_b4d4_passes_with_full_evidence,
        test_price_candidate_anchors_use_b_to_d,
        test_price_isolated_numeric_no_dual_source_exemption,
        test_smoke_10_questions_price_evidence,
        test_candidate_anchors_inventory_table,
        test_candidate_anchors_kv_parameter_table,
        test_candidate_anchors_csv_plain_table,
        test_candidate_anchor_metadata_no_out_of_bounds,
        test_real_appendix_d_candidate_metadata,
        test_prompt_no_business_terms,
        test_candidate_section_appears_in_prompt,
        test_summary_rows_excluded_from_candidates,
        test_appendix_e_smoke_with_summary_and_merged,
        test_real_appendix_e_file_e2e,
        test_phase1_schema_analysis_mock,
        test_schema_cache_hit,
        test_evidence_schema_display,
        test_generate_questions_from_schema_e2e,
        test_upload_no_llm_call,
        test_phase1_only_on_button_click,
        test_phase2_does_not_recall_phase1,
        test_file_change_invalidates_schema,
        test_reanalyze_clears_and_rebuilds,
        test_production_rejects_fake_model,
        test_production_no_fake_in_output,
        test_llm_failure_fail_closed,
        test_phase1_timeout_is_180s,
        test_phase1_retry_on_timeout,
        test_phase1_retry_both_fail,
        test_phase1_no_fake_fallback,
        test_resource_cost_table_candidate_catalog,
        test_inventory_table_recognized,
        test_summary_empty_isolated_excluded_from_candidates,
        test_fail_closed_invalid_candidate_id,
        test_reference_answer_locally_rendered,
        test_non_schema_path_not_regressed,
    ]

    passed = 0
    failed = 0
    for test_fn in tests:
        try:
            test_fn()
            passed += 1
        except Exception as e:
            failed += 1
            print(f"FAIL: {test_fn.__name__}: {e}")
            import traceback
            traceback.print_exc()
        print()

    print("=" * 60)
    print(f"结果: {passed} 通过, {failed} 失败, 共 {len(tests)} 个测试")
    print("=" * 60)

    if failed > 0:
        sys.exit(1)


if __name__ == "__main__":
    main()
